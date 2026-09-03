# -*- coding: utf-8 -*-
"""
Step D: qhzy-中药学 收尾
1. 0目录 追加 30.4 条目（对齐正文新增节）
2. 0封面 增补 MM-01 元数据卡
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中药学.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=6):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# ═══ 1. 0目录 追加 30.4 ═══
# 中药学 0目录 格式: 6列(章|节|小节|段|内容|?), 先探测
ws = wb['0目录']
print("0目录 max_column:", ws.max_column)
# 检查30章结尾
found = False
for r in range(1, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    if e and '30.3' in str(e):
        found = True
print("30.3 存在:", found)

append_rows('0目录', [
    ['', '', '', '', '【qhzy 增补·2026-08】以下为第30章新增节条目:', ''],
    ['30', '4', '', '', '30.4 岐黄智医分布式教学平台', ''],
    ['', '', '1', '', '30.4.1 2云+2硬+4台式机拓扑与数据流', ''],
    ['', '', '', '1', '30.4.1.1 节点分工', ''],
    ['', '', '', '2', '30.4.1.2 中药数据闭环', ''],
    ['', '', '2', '', '30.4.2 六者SOUL与中药学对接', ''],
    ['', '', '', '1', '30.4.2.1 药者(核心对接角色)', ''],
    ['', '', '', '2', '30.4.2.2 医者/患者/规者/法者', ''],
], ncols=6)
print("0目录 30.4 条目已追加")

# ═══ 2. 0封面 MM-01 元数据卡 ═══
ws = wb['0封面']
start = ws.max_row + 2
cover_rows = [
    ['', '', '【qhzy 增补·2026-08】教材元数据卡(分布式体系版)', '', ''],
    ['', '', '教材编号', 'MM-01', ''],
    ['', '', '教材名称', '中药学(岐黄智医版)', ''],
    ['', '', '所属领域', 'D02 中药学院', ''],
    ['', '', '教学层级', '本科·基础经典(第二层)', ''],
    ['', '', '学分/学时', '6 学分 / 160 学时(两学期)', ''],
    ['', '', '前置要求', 'CM-01 中医基础理论', ''],
    ['', '', '后续衔接', 'MM-02 方剂学 → CM-13 内科学', ''],
    ['', '', '分布式承载', '台式机T1-T4药材图像采集 · 浪潮硬②A800模型训练 · GitHub CI/CD门禁 · 华为云公网展示', ''],
    ['', '', '版本', 'v2.0-qhzy · 2026-08 · 由种子 v1.0 适应性修改', ''],
]
for i, row in enumerate(cover_rows):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=start + i, column=c, value=v)
print("0封面 MM-01 元数据卡已增补")

wb.save(DST)
print("Step D 完成")
