# -*- coding: utf-8 -*-
"""ch01 图注统一为四要素格式（a.提示语 b.线注释 c.符号注释 d.指针图段）
同步 Excel 图注行 + 正文 md 图注行"""
from pathlib import Path

XLSX = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
CH01 = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\ch01')

# 图注行号 → 新图注文本（基于试点版内容重组为 a./b./c./d. 四要素）
NEW_NOTES = {
    8: ('图注：a.对应正文 1.1、1.2 节（智能体/环境与网格世界的状态与动作）：'
        '网格世界重解释为医院垂直切面，智能体为患者/导诊。'
        'b.图中蓝色虚线箭头为示例路径：s1（1层大厅）→s2（1层走廊）→乘电梯↑→s5（2层走廊）'
        '→乘电梯↑→s8（3层走廊）→右转进入 s9（专家门诊），全程沿走廊+电梯，不穿越任何科室房间。'
        'c.符号：s1-s9 为状态（楼层-走廊-诊室位置）、s6 为禁区（手术区，进入即受罚 r=-1）、'
        's9 为目标（右手侧第 2 个诊室窗口，敲门进入或候诊叫号）、a1-a5 为上/下/左/右/原地。'
        'd.指针图段：指向电梯核 s2/s5/s8——患者上下楼必经垂直交通核，'
        '对应"必要时应乘电梯到目标楼层"的真实寻路逻辑。'),
    16: ('图注：a.对应正文 1.2 节（状态与动作）：四诊信息→治法方剂→疗效反馈的临床映射。'
         'b.图中三个圆角框自左至右为状态 S→动作 A→奖励 R，顶部箭头"辨证→论治"连接 S 与 A、'
         '"执行动作后收到 Rt+1"连接 A 与 R，底部橙色回路箭头为"复诊转归→新状态 St+1"。'
         'c.符号：S 为状态（望闻问切四诊总和）、A 为动作（八法及具体方剂）、R 为奖励（疗效反馈正/负/零）、'
         'Rt+1 为下一时刻奖励。'
         'd.指针图段：指向底部橙色回路——一次诊疗的转归成为下一次辨证的状态，'
         '正是"状态-动作-奖励"三要素的闭环。'),
    33: ('图注：a.对应正文 1.4 节（策略：辨证论治的决策规则）。'
         'b.图中左表为确定性策略表：每个状态（证候）恰好选择一个动作（方剂），覆盖全部 9 个状态；'
         '右图三条彩色轨迹起点不同、终点相同——蓝色 s1→s4→s7→s8→s9、绿色 s2→s5→s8→s9、'
         '橙色 s3→s6→s9（穿禁区手术区，示误）。'
         'c.符号：π(a|s) 为策略（给定状态 s 采取动作 a 的概率，确定性策略概率为 1）；'
         'a1-a5 为上/右/下/左/原地。'
         'd.指针图段：指向橙色穿禁区轨迹——策略若令患者穿越手术区即失败，'
         '最优策略必须避开禁区走廊规划，对应"患者绝不穿越科室房间"的约束。'),
    48: ('图注：a.对应正文 1.6 节（轨迹与回报）：诊疗全过程的奖励积累。'
         'b.图中蓝色箭头链 s1→s2→s5→s8→s9 为主链（初诊→辨证→首方→复诊→痊愈），'
         '箭头上方标注动作（a2/a3/a3/a2）、下方标注奖励（r=0/0/0/+1），末端绿色框汇总回报。'
         'c.符号：G 为回报（一条轨迹上所有奖励之和，主链 G=0+0+0+1=1）；'
         'γ 为折扣率（正文 1.6.2），用于无限长轨迹的折扣回报 Gt=Rt+1+γRt+2+…。'
         'd.指针图段：指向末端绿色回报框——整条诊疗链的累计疗效，'
         '正是"回报=全疗程健康结局"的量化表达。'),
    52: ('图注：a.对应正文 1.6.3 节（MDP 五元组收束后的双向映射）。'
         'b.图中外环五个浅绿盒为中医临床元素（四诊症状、辨证论治、治法方剂、疗效反馈、病邪传变），'
         '内环五个浅蓝盒为强化学习概念（状态 state、策略 policy、动作 action、奖励 reward、'
         '状态转移 state transition），红蓝径向双向箭头为"映射⇄反哺"。'
         'c.符号：S 为状态（四诊症状）、A 为动作（治法方剂）、R 为奖励（疗效反馈）、'
         'P 为状态转移（病邪传变）、π 为策略（辨证论治）。'
         'd.指针图段：指向任意一条红蓝双向箭头——中医概念与 RL 概念互为翻译，'
         '全书各章均以本映射环为语义基准。'),
}

from openpyxl import load_workbook

# 1) 更新 Excel
wb = load_workbook(XLSX)
ws = wb['1cmrl-conceptions']
for row, text in NEW_NOTES.items():
    ws.cell(row, 3, text)
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0,
                                        (len(text) // 44 + 1) * 14 + 6)
wb.save(XLSX)
print(f'[Excel] 更新 {len(NEW_NOTES)} 个图注行')

# 2) 同步 md（每个文件唯一的"图注："行替换为对应新文本）
md_map = {  # 图注行号 → md 文件（按图注出现顺序）
    8: '01-1.1.md', 16: '02-1.2.md', 33: '04-1.4.md', 48: '06-1.6.md',
    52: '06-1.6.md',  # 图1-6 也在 06-1.6.md（两个图注）
}
from collections import defaultdict
md_targets = defaultdict(list)
for row, fname in md_map.items():
    md_targets[fname].append(NEW_NOTES[row])

for fname, notes in md_targets.items():
    f = CH01 / fname
    text = f.read_text(encoding='utf-8')
    lines = text.splitlines()
    out = []
    ni = 0
    replaced = 0
    for ln in lines:
        if ln.strip().startswith('图注：') and ni < len(notes):
            out.append(notes[ni])
            ni += 1
            replaced += 1
        else:
            out.append(ln)
    f.write_text('\n'.join(out), encoding='utf-8')
    print(f'[md] {fname}: 替换 {replaced} 个图注行')
