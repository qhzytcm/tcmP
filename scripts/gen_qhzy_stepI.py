# -*- coding: utf-8 -*-
"""
Step I: 0目录 顺序重组 + 标题修正
问题：平台块(R943-R1020)物理上在体质块(R1042-R1062)之前，章顺序错乱；
      R943 标题缺"岐黄智医分布式"前缀；R1041 被误改；R1032 引用需更新。
方案：内存重组 0目录 → 表头+1-11章 | 12体质 | 13平台 | 说明文字 | 附篇
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中医基础理论.xlsx'
wb = openpyxl.load_workbook(DST)
ws = wb['0目录']

# 读取全部行（1..max_row）
rows = []
for r in range(1, ws.max_row + 1):
    rows.append([ws.cell(row=r, column=c).value for c in range(1, 6)])

def get(r):
    return rows[r - 1]

# 修正 R943 标题（平台块首行）
rows[942][4] = '第13章　岐黄智医分布式教学平台与评价体系'

# 修正 R1032 引用（知识树落地说明）
rows[1031][4] = rows[1031][4].replace('第12章独立成章', '第13章独立成章').replace('第12章', '第13章')

# 修正 R1040 引用（④ 本目录与规划教材对应）
rows[1039][4] = rows[1039][4].replace("第13章'教学平台与评价体系'", "第13章'岐黄智医分布式教学平台与评价体系'")

# 修正 R1041 增补说明（"第13章(体质)" → "第12章(体质)"）
rows[1040][4] = rows[1040][4].replace('以下为第13章(体质)', '以下为第12章(体质)')

# 提取块
platform_block = rows[942:1020]    # R943-R1020 平台块（78行）
note_block = rows[1023:1041]       # R1024-R1041 说明文字（18行）
tizhi_block = rows[1041:1062]      # R1042-R1062 体质块（21行）
fupian_block = rows[1062:1067]     # R1063-R1067 附篇（5行）
head_block = rows[:942]            # R1-R942 表头+1-11章

# 重组：head | tizhi | platform | note | fupian
new_rows = head_block + tizhi_block + platform_block + note_block + fupian_block

# 清空重写
for r in range(1, ws.max_row + 1):
    for c in range(1, 6):
        ws.cell(row=r, column=c, value=None)
for i, row in enumerate(new_rows):
    for c, v in enumerate(row, start=1):
        if v is not None:
            ws.cell(row=i + 1, column=c, value=v)

wb.save(DST)
print("Step I 完成, 新行数:", len(new_rows))

# 复核章顺序
wb2 = openpyxl.load_workbook(DST, read_only=True)
ws2 = wb2['0目录']
for i, row in enumerate(ws2.iter_rows(min_row=1, max_col=5), 1):
    cells = ['' if x.value is None else str(x.value).strip() for x in row]
    a, b, c, d, e = cells
    if e and len(e) < 30 and e.startswith('第') and '章' in e and not e.startswith('第13章 岐黄') or (e and e.startswith('第1') and '章' in e and len(e) < 30):
        if e.startswith('第') and '章' in e and len(e) < 30:
            print(f"R{i}: {e}")
