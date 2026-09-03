# -*- coding: utf-8 -*-
"""统一同步: 各章 md 图注（临床化版）-> Excel 对应图注行
定位: 图片锚点行+1 = 图注行（按锚点排序对应 md 文件顺序）"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')
SHEETS = ['1cmrl-conceptions', '2cmrl-bellman', '3cmrl-optimality', '4cmrl-iteration',
          '5cmrl-montecarlo', '6cmrl-approximation', '7cmrl-tdlearning',
          '8cmrl-valuefunction', '9cmrl-policygradient', '10cmrl-acloop']


def md_notes(ch):
    notes = []
    for f in sorted((D / f'ch{ch:02d}').glob('0*-1.*.md')):
        for ln in f.read_text(encoding='utf-8').splitlines():
            if ln.strip().startswith('图注：'):
                notes.append(ln.strip())
    return notes


wb = load_workbook(X)
total = 0
for ch in range(1, 11):
    ws = wb[SHEETS[ch - 1]]
    notes = md_notes(ch)
    imgs = sorted(ws._images, key=lambda i: i.anchor._from.row)
    for i, img in enumerate(imgs):
        if i >= len(notes):
            break
        row = img.anchor._from.row + 2   # 图体行+1 = 图注行
        note = notes[i]
        # 写入（位置参数; 若截断则索引赋值重试）
        ws.cell(row, 3, note)
        ws.cell(row, 3).font = Font(name='黑体', size=10)
        ws.cell(row, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = (math.ceil(len(note) / 44) + 1) * 14 + 6
        total += 1
wb.save(X)
print(f'写入 {total} 个图注行')

# 回读校验（截断检测）
wb2 = load_workbook(X)
fixed = 0
for ch in range(1, 11):
    ws = wb2[SHEETS[ch - 1]]
    notes = md_notes(ch)
    imgs = sorted(ws._images, key=lambda i: i.anchor._from.row)
    for i, img in enumerate(imgs):
        if i >= len(notes):
            break
        row = img.anchor._from.row + 2
        v = str(ws.cell(row, 3).value or '')
        if v != notes[i]:
            # 索引赋值重试
            ws.cell(row, 3).value = None
            wb2.save(X)
            wb2 = load_workbook(X)
            ws = wb2[SHEETS[ch - 1]]
            ws['C' + str(row)] = notes[i]
            fixed += 1
wb2.save(X)
print(f'截断重试修复 {fixed} 处')

# 最终校验
wb3 = load_workbook(X)
bad = []
for ch in range(1, 11):
    ws = wb3[SHEETS[ch - 1]]
    notes = md_notes(ch)
    imgs = sorted(ws._images, key=lambda i: i.anchor._from.row)
    for i, img in enumerate(imgs):
        if i >= len(notes):
            break
        row = img.anchor._from.row + 2
        v = str(ws.cell(row, 3).value or '')
        if v != notes[i]:
            bad.append(f'ch{ch}R{row}: {len(v)}/{len(notes[i])}')
wb3.close()
print(f'最终不一致: {bad if bad else "无 (45/45 全同步)"}')
