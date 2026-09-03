# -*- coding: utf-8 -*-
"""恢复内容简介（知识树前插 2 行）+ 图片锚点修正"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['封面页']

INTRO = ('本书以强化学习的数学理论为主线，完整移植到中医药临床与医院场景，三编 10 章：'
         '上编·提高认知（状态动作奖励/决策过程模型/统计估计随机逼近）；中编·确认信仰'
         '（价值估计方程/最优价值迭代/时序差分）；下编·追求幸福（奖励设计与四个最适度/'
         '函数近似/策略梯度/行动者-评价者与六者协同）。特色：中医场景全覆盖（45 幅插图）、'
         '数学与临床双向可验算、六者平台贯通、最适度≠最优化哲学贯穿。'
         '读者：中医药院校师生、临床医师、AI 研究者。')

# 1) R37（知识树图题前）插 2 行
ws.insert_rows(37, 2)

# 2) 内容简介标题 + 正文
ws.cell(37, 1, '内容简介')
ws.merge_cells(start_row=37, start_column=1, end_row=37, end_column=3)
c = ws.cell(37, 1)
c.font = Font(name='黑体', size=13, bold=True)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[37].height = 24

ws.cell(38, 1, INTRO)
ws.merge_cells(start_row=38, start_column=1, end_row=38, end_column=3)
c = ws.cell(38, 1)
c.font = Font(name='黑体', size=10)
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[38].height = (math.ceil(len(INTRO) / 95) + 1) * 14 + 4

# 3) 图片锚点 +2（原 38 -> 40）
for img in ws._images:
    img.anchor._from.row += 2

wb.save(X)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['封面页']
for r in (36, 37, 38, 39, 40, 41, 42):
    v = str(ws2.cell(r, 1).value or '')
    print(f'R{r}: ({len(v)}字) {v[:36]}')
anchors = [img.anchor._from.row + 1 for img in ws2._images]
print(f'图片锚点: {anchors}')
wb2.close()
