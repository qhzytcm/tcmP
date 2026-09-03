# -*- coding: utf-8 -*-
"""修复 ch6 R18 图注截断（30/634）——索引赋值重写"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')

note = ''
for ln in (D / 'ch06' / '02-1.2.md').read_text(encoding='utf-8').splitlines():
    if ln.strip().startswith('图注：'):
        note = ln.strip()
        break
print(f'md 图注: {len(note)} 字')

for attempt in range(3):
    wb = load_workbook(X)
    ws = wb['6cmrl-approximation']
    ws['C18'] = None
    wb.save(X)
    wb = load_workbook(X)
    ws = wb['6cmrl-approximation']
    ws['C18'] = note
    ws['C18'].font = Font(name='黑体', size=10)
    ws['C18'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[18].height = (math.ceil(len(note) / 44) + 1) * 14 + 6
    wb.save(X)
    wb = load_workbook(X)
    v = str(wb['6cmrl-approximation']['C18'].value or '')
    wb.close()
    print(f'尝试{attempt+1}: {len(v)}/{len(note)} {"✅" if v == note else "⚠"}')
    if v == note:
        break
