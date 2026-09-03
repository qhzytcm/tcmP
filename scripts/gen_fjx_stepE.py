# -*- coding: utf-8 -*-
"""
Step E: qhzy-方剂学 完善（深检后的两处小补充）
1. 0目录 学时汇总后补充 24.4 学时归属说明（平台章14学时含24.4）
2. 病证单元库 说明行补充方剂溯源注释（代表方可回落各论114首方）
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-方剂学.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=5):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# 1. 0目录：在【qhzy 增补】标记行前插入 24.4 学时归属
ws = wb['0目录']
target = None
for r in range(1, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    if e and '【qhzy 增补·2026-08】以下为第24章' in str(e):
        target = r
        break
if target:
    ws.insert_rows(target)
    ws.cell(row=target, column=1, value='24.4')
    ws.cell(row=target, column=5, value='24.4 岐黄智医分布式教学平台（学时含于平台章14学时内, 无需另计）')
    print(f"0目录 R{target} 已插入 24.4 学时归属说明")
else:
    print("!! 未找到增补标记行")

# 2. 病证单元库 补充方剂溯源注释
append_rows('病证单元库', [
    ['补充说明(qhzy v2.0)', '代表方列均可回落第5-23章各论114首方(如麻黄汤/天麻钩藤饮/六味地黄丸); 部分方为经方/时方别名变体(如"麦门冬汤"即"麦冬汤"), 均可在各论溯源; "同 BZ-xxx"为方剂同前条引用。', '', '', '', '', '', ''],
], ncols=8)

wb.save(DST)
print("Step E 完成")
