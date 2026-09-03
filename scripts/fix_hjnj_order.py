# -*- coding: utf-8 -*-
"""重排: 16.素问·运气七篇 移到 15 与 17 之间"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
names = wb.sheetnames
print(f"当前顺序: ...{names[names.index('15.素问·调经缪刺')-1] if '15.素问·调经缪刺' in names else '?'}...")

# 移动 16 到 15 之后
wb.move_sheet('16.素问·运气七篇', offset=-1)
wb.save(PATH)
names2 = wb.sheetnames
i15 = names2.index('15.素问·调经缪刺')
i16 = names2.index('16.素问·运气七篇')
i17 = names2.index('17.素问·医论杂篇')
print(f"15@{i15} 16@{i16} 17@{i17}")
print(f"顺序正确: {i16 == i15 + 1 and i17 == i16 + 1}")
