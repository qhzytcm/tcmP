# -*- coding: utf-8 -*-
"""修复: 第25章注释补「悬疑/非零和」标注"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
ws = wb['25.灵枢·津液清浊']

n = 0
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value == '注释' and row[2].value:
        v = str(row[2].value)
        if '非零和' not in v and '悬疑' not in v:
            # 补: 在结尾加"（异说并存，非零和局；难解处悬疑存疑）"
            new = v.rstrip('。') + '；诸家解说并见（非零和局），难解处悬疑存疑。'
            ws.cell(row=row[0].row, column=3, value=new)
            print(f"{row[1].value}: 已补 (len {len(v)}→{len(new)})")
            n += 1
        elif '非零和' not in v:
            new = v.rstrip('。') + '；并存（非零和局）。'
            ws.cell(row=row[0].row, column=3, value=new)
            print(f"{row[1].value}: 已补非零和 (len {len(v)}→{len(new)})")
            n += 1

wb.save(PATH)
print(f"修复 {n} 条")
