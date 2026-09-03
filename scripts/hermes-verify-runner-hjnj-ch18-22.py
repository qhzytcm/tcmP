# -*- coding: utf-8 -*-
"""runner: tempfile.mkstemp OS-safe 动态生成 hermes-verify- 脚本 → 运行 → 自动清理（ad-hoc）"""
import os, subprocess, sys, tempfile

BODY = r'''# -*- coding: utf-8 -*-
"""hermes-verify-hjnj-ch18-22.py — qhzy-黄帝内经 18-22章+重建修复 聚焦验证（ad-hoc, tempfile OS-safe）"""
import sys, re, os, json
import openpyxl

print(f"[hermes-verify] 脚本: {os.path.abspath(__file__)}")

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
fails, passes = [], []

def check(name, cond, detail=""):
    (passes if cond else fails).append(f"{name}{(' | ' + detail) if detail and not cond else ''}")

wb = openpyxl.load_workbook(PATH, read_only=True)

# 1. Sheet 顺序（15-16-17-18-19-20-21-22）
names = wb.sheetnames
seq = ['15.素问·调经缪刺','16.素问·运气七篇','17.素问·医论杂篇','18.灵枢·针道原穴',
       '19.灵枢·根结针法','20.灵枢·经脉经别','21.灵枢·经筋骨度','22.灵枢·营卫气血']
idx = [names.index(s) for s in seq]
check("15-22连续", idx == list(range(idx[0], idx[0]+8)), f"{idx}")
check("Sheet数=37", len(names) == 37, str(len(names)))

# 2. 无截断
ch = [s for s in names if re.match(r'^\d+\.', s)]
suspects = []
for s in ch:
    ws = wb[s]
    for row in ws.iter_rows(min_row=2, max_col=3):
        v = row[2].value
        if v and isinstance(v, str) and len(v) > 10:
            if not re.search(r'[。；：！？」』）】]$', v) and re.search(r'[以之而于与为者]$', v.strip()):
                suspects.append((s, row[0].row, len(v)))
check("无截断", not suspects, f"{len(suspects)}处")

# 3. 篇目统计
suwen = ling = 0
for s in ch:
    ws = wb[s]
    n = sum(1 for row in ws.iter_rows(min_row=2, max_col=3) if row[0].value == '篇名')
    if int(s.split('.')[0]) <= 17:
        suwen += n
    else:
        ling += n
check("素问81篇", suwen == 81, f"实际{suwen}")
check("灵枢21篇(1-21)", ling == 21, f"实际{ling}")

# 4. 18-22章结构
expected = {
    '18.灵枢·针道原穴': (4, ['九针十二原','本输','小针解','邪气脏腑病形']),
    '19.灵枢·根结针法': (5, ['根结','寿夭刚柔','官针','本神','终始']),
    '20.灵枢·经脉经别': (3, ['经脉','经别','经水']),
    '21.灵枢·经筋骨度': (5, ['经筋','骨度','五十营','营气','脉度']),
    '22.灵枢·营卫气血': (4, ['营卫生会','四时气','五邪','寒热病']),
}
for s, (n_pian, pnames) in expected.items():
    ws = wb[s]
    cats = {'篇名': 0, '原文': 0, '通读': 0, '注释': 0}
    got = []
    for row in ws.iter_rows(min_row=2, max_col=3):
        c = row[0].value
        if c in cats:
            cats[c] += 1
        if c == '篇名' and row[2].value:
            got.append(str(row[2].value))
    check(f"{s}篇名{n_pian}", cats['篇名'] == n_pian, f"实际{cats['篇名']}")
    check(f"{s}原文≥15", cats['原文'] >= 15, f"实际{cats['原文']}")
    check(f"{s}通读={n_pian}", cats['通读'] == n_pian, f"实际{cats['通读']}")
    check(f"{s}注释≥3", cats['注释'] >= 3, f"实际{cats['注释']}")
    for pn in pnames:
        check(f"{s}篇[{pn}]", any(pn in g for g in got))
    zt = "\n".join(str(row[2].value) for row in ws.iter_rows(min_row=2, max_col=3) if row[0].value == '注释' and row[2].value)
    check(f"{s}悬疑/非零和", '悬疑' in zt and '非零和' in zt)

# 5. 修复点
for s, code, kw in [
    ('20.灵枢·经脉经别', '20.1.1.13', '不盛不虚，以经取之'),
    ('22.灵枢·营卫气血', '22.1.1.11', '上焦如雾'),
    ('16.素问·运气七篇', '16.1.1.7', '太虚寥廓，肇基化元'),
]:
    ws = wb[s]
    found = any(row[1].value == code and row[2].value and kw in str(row[2].value)
                for row in ws.iter_rows(min_row=2, max_col=3))
    check(f"修复[{s} {code}]", found)

# 6. 16章全文版（299原文段/9篇）
ws = wb['16.素问·运气七篇']
n_or = sum(1 for row in ws.iter_rows(min_row=2, max_col=3) if row[0].value == '原文')
n_p = sum(1 for row in ws.iter_rows(min_row=2, max_col=3) if row[0].value == '篇名')
check("16章9篇", n_p == 9, f"实际{n_p}")
check("16章原文299", n_or == 299, f"实际{n_or}")

# 7. 进度
d = json.load(open(r'C:\Users\DELL\tcmP\scripts\progress\hjnj_writing_progress.json', encoding='utf-8'))
check("进度22/31", d['done'] == 22, str(d['done']))

wb.close()
print(f"\nPASS {len(passes)} | FAIL {len(fails)}")
for p in passes:
    print("  ✔", p)
for f in fails:
    print("  ✘", f)
sys.exit(1 if fails else 0)
'''

fd, tmp_path = tempfile.mkstemp(prefix='hermes-verify-hjnj-ch18-22-', suffix='.py', dir=tempfile.gettempdir())
with os.fdopen(fd, 'w', encoding='utf-8') as f:
    f.write(BODY)
print(f"[hermes-verify] 临时脚本: {tmp_path}")

PY = r'C:\Users\DELL\AppData\Local\hermes\hermes-agent\venv\Scripts\python.exe'
try:
    r = subprocess.run([PY, tmp_path], capture_output=True, text=True, encoding='utf-8', timeout=180)
    print(r.stdout)
    if r.stderr:
        print("STDERR:", r.stderr[-1200:])
    sys.exit(r.returncode)
finally:
    try:
        os.remove(tmp_path)
        print(f"[hermes-verify] 已清理: {tmp_path} 存在={os.path.exists(tmp_path)}")
    except OSError:
        pass
