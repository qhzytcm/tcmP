# -*- coding: utf-8 -*-
"""验证 v2 重插: media PNG=45/GIF=0 + 图题黑体居中 + 图注左顶格 + 锚点行"""
import zipfile
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
ok_all = True

with zipfile.ZipFile(X) as z:
    gifs = [n for n in z.namelist() if n.startswith('xl/media/') and n.endswith('.gif')]
    pngs = [n for n in z.namelist() if n.startswith('xl/media/') and n.endswith('.png')]
print(f'xl/media: PNG={len(pngs)} GIF={len(gifs)} '
      f'{"✅" if len(pngs) == 45 and not gifs else "⚠"}')

wb = load_workbook(X)
sheets = ['1cmrl-conceptions', '2cmrl-bellman', '3cmrl-optimality', '4cmrl-iteration',
          '5cmrl-montecarlo', '6cmrl-approximation', '7cmrl-tdlearning',
          '8cmrl-valuefunction', '9cmrl-policygradient', '10cmrl-acloop']
total = 0
for sn in sheets:
    ws = wb[sn]
    imgs = getattr(ws, '_images', None) or []
    total += len(imgs)
    bad = []
    for img in imgs:
        r = img.anchor._from.row + 1        # 图体行(1-based)
        cap = ws.cell(r - 1, 3)             # 图题行
        note = ws.cell(r + 1, 3)            # 图注行
        cap_txt = str(cap.value or '')
        note_txt = str(note.value or '')
        # 图题: 黑体+居中+以"图 "开头; 图注: 左对齐+以"图注："开头
        if not (cap_txt.startswith('图') and cap.font.name == '黑体'
                and cap.alignment.horizontal == 'center'):
            bad.append(f'行{r-1}图题[{cap.font.name}/{cap.alignment.horizontal}]"{cap_txt[:16]}"')
        if not (note_txt.startswith('图注：') and note.alignment.horizontal == 'left'):
            bad.append(f'行{r+1}图注[{note.alignment.horizontal}]"{note_txt[:16]}"')
    status = '✅' if not bad else '⚠'
    if bad:
        ok_all = False
    print(f'{sn}: {len(imgs)} 幅 {status}')
    for b in bad[:2]:
        print(f'    {b}')
wb.close()
print(f'合计: {total} 幅 {"✅ 45/45 全部规范" if total == 45 and ok_all else "⚠ 存在问题"}')
