# -*- coding: utf-8 -*-
"""清空 ch02 R33 重复图注行 → 重跑替换 ch02"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['2cmrl-bellman']
v32 = str(ws.cell(32, 3).value or '')
v33 = str(ws.cell(33, 3).value or '')
if v32 == v33 and v32.startswith('图注：'):
    ws.cell(33, 3, None)
    wb.save(X)
    print('R33 重复图注行已清空')
else:
    print(f'R32 != R33 或非图注, 未清空 (R32={len(v32)}字 R33={len(v33)}字)')
    wb.close()
    raise SystemExit(1)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['2cmrl-bellman']
n = sum(1 for r in range(1, ws2.max_row + 1)
        if str(ws2.cell(r, 3).value or '').strip().startswith('图注：'))
n_img = len(ws2._images)
wb2.close()
print(f'图注行数: {n}, 图数: {n_img}')
