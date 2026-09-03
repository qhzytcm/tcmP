# -*- coding: utf-8 -*-
"""ICD-11 标准术语对照表 -> 0映射2 sheet 追加（中医药主审 ICD-11 命名红线附录）"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['0映射2']

APP = [
    'ICD-11 标准命名红线附录（中医药术语定义与命名遵照 ICD-11 标准执行；tcmP 平台已内嵌 ICD-11 编码引擎，拓展 ragflow/FTS5）',
    '一、全书病证术语 × ICD-11 标准对照表（2026-01 MMS 版, icd11_mms.db 31,838 实体）',
    '中医病证 | 正文出现 | ICD-11 编码 | ICD-11 英文名 | 说明',
    '不寐 | 42 | 7A00 | Chronic insomnia | 标准编码',
    '心悸 | 9 | MC81.2 | Palpitations | 标准编码',
    '咳嗽 | 6 | MD12 | Cough | 标准编码（MD12 症状）',
    '感冒 | 15 | SA60 | Common cold disorder (TM1) | TM1 传统医学章节',
    '黄疸 | 0 | SA01 | Jaundice disorder (TM1) | TM1 传统医学章节',
    '胸痹 | 0 | BA40 | Angina pectoris | 西医对应编码',
    '眩晕 | 0 | MB48.0 | Vertigo | 标准编码',
    '泄泻 | 0 | ME05.1 | Diarrhoea | 标准编码',
    '便秘 | 0 | ME05.0 | Constipation | 标准编码',
    '水肿 | 0 | MG29 | Oedema | 标准编码',
    '喘证 | 0 | CA23 | Asthma | 标准编码',
    '虚劳 | 0 | MG22 | Fatigue | 标准编码（8E49 为病毒后疲劳）',
    '中风 | 0 | 8B20 | Stroke | 标准编码（8B21 缺血/8B22 出血细分）',
    '肺痨 | 0 | 1B10 | Tuberculosis of respiratory system | 标准编码',
    '消渴 | 0 | 5A10/5A11 | Diabetes mellitus | 5A10 1型/5A11 2型, 教学示例用通名',
    '郁证 | 0 | 6A70 | Depressive disorder | 标准编码（单次发作）',
    '头痛 | 0 | 8A81 等 | Headache disorders | 8A80 为偏头痛, 通名用 Headache disorders',
    '痹证 | 0 | FA2Z 等 | 关节痹痛类 | FA25 为痛风, 痹证为证候范围名',
    '胃脘痛 | 0 | 症状编码 | Epigastric pain | 症状类（非结构异常 DA90）',
    '汗证 | 0 | 症状编码 | Sweating disorder | 症状类（非 MG46 SIRS）',
    '心脾两虚/阴虚/阳虚/气虚/血虚/痰湿/瘀血 | 证候名 | — | 八纲/脏腑辨证 | ICD-11 无单病编码, 按辨证参考（随证治之映射）',
    '二、命名红线（审稿/写作强制）',
    '① 病证名定义与英文命名遵照 ICD-11（不寐=Chronic insomnia 7A00、感冒=Common cold disorder TM1 SA60 等）；',
    '② 中医病证优先 TM1 传统医学章节编码（SA00-SF57），无 TM1 条目的用西医对应编码（心悸 MC81.2）；',
    '③ 证候名（心脾两虚/阴虚/阳虚等）为辨证单元, ICD-11 无单病编码, 正文按辨证规范表述, 不得硬套西医病名；',
    '④ 教学示例涉及病证时, 英文命名须与 ICD-11 一致（正文术语表核对）；',
    '⑤ tcmP 平台对接：病证单元 DSU 经 /diag /bianzheng /semantic-search /rag 端点挂接 ICD-11 编码（FTS5+TF-IDF+RRF 检索, 内网零 token）。',
]
start = ws.max_row + 2
r = start
for line in APP:
    ws.cell(r, 1, line)
    r += 1
wb.save(X)
print(f'0映射2 追加 ICD-11 命名红线附录 {len(APP)} 行（R{start}-R{r-1}）')
