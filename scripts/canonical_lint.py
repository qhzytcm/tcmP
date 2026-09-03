# -*- coding: utf-8 -*-
"""canonical lint — cmrl 项目语法/规范检查
运行: python scripts/canonical_lint.py
覆盖: 全部生成脚本 + 工具脚本 py_compile; 图注行数=图数; 图题黑体
退出码: 0=全绿 1=有失败
"""
import py_compile
import sys
from pathlib import Path

ROOT = Path(r'C:\Users\DELL\tcmP')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')
X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
results = []


def check(name, ok, detail=''):
    results.append((name, ok, detail))
    print(f'{"PASS" if ok else "FAIL"}  {name}  {detail}')


# L1 全部 python 脚本语法编译
py_files = []
for base in (ROOT / 'scripts', D / 'figures', D / 'ch01'):
    if base.exists():
        py_files += list(base.glob('*.py'))
py_files = sorted(set(py_files))
bad = []
for p in py_files:
    try:
        py_compile.compile(str(p), doraise=True)
    except py_compile.PyCompileError as e:
        bad.append(f'{p.name}: {str(e)[:60]}')
check(f'L1 py_compile ({len(py_files)} 脚本)', not bad, f'异常={bad[:3] or "无"}')

# L2 各章图注行数 = 图数
from openpyxl import load_workbook
SHEETS = ['1cmrl-conceptions', '2cmrl-bellman', '3cmrl-optimality', '4cmrl-iteration',
          '5cmrl-montecarlo', '6cmrl-approximation', '7cmrl-tdlearning',
          '8cmrl-valuefunction', '9cmrl-policygradient', '10cmrl-acloop']
wb = load_workbook(X)
bad2 = []
for ch, sn in enumerate(SHEETS, 1):
    ws = wb[sn]
    n_notes = sum(1 for r in range(1, ws.max_row + 1)
                  if str(ws.cell(r, 3).value or '').strip().startswith('图注：'))
    n_img = len(ws._images)
    if n_notes != n_img:
        bad2.append(f'ch{ch}:{n_notes}/{n_img}')
check('L2 图注行数=图数 (10 章)', not bad2, f'异常={bad2 or "无"}')

# L3 全书图题行黑体
bad3 = []
for sn in SHEETS:
    ws = wb[sn]
    for r in range(1, ws.max_row + 1):
        c = ws.cell(r, 3)
        if str(c.value or '').startswith('图 ') and c.font.name != '黑体':
            bad3.append(f'{sn}R{r}')
check('L3 图题行黑体', not bad3, f'异常={bad3[:3] or "无"}')
wb.close()

fails = [r for r in results if not r[1]]
print(f'\n=== CANONICAL LINT: {"ALL PASS" if not fails else f"{len(fails)} FAILED"} '
      f'({len(results) - len(fails)}/{len(results)}) ===')
sys.exit(0 if not fails else 1)
