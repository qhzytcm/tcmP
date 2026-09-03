# -*- coding: utf-8 -*-
"""封面页 R41 图注补'自下而上'方向说明（树形读法）"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['封面页']
v = str(ws.cell(41, 1).value or '')
print(f'原图注: {len(v)} 字')
if '自下而上' not in v:
    add = '树形读法自下而上：须根在底部吸取数学养分，依次向上为根、干、枝，叶冠在顶部——读者沿树而上，逐级领悟。'
    v2 = v.rstrip() + add
    ws['A41'] = v2
    wb.save(X)
    print(f'追加后: {len(v2)} 字')
else:
    print('已含自下而上, 未改')
    wb.close()
    raise SystemExit(0)

# 回读校验
wb2 = load_workbook(X)
v3 = str(wb2['封面页'].cell(41, 1).value or '')
wb2.close()
print(f'回读: {len(v3)} 字 {"✅" if v3 == v2 else "⚠截断"}')
