# -*- coding: utf-8 -*-
"""
Step J: 0目录 纯内存重建（终极修复）
从种子 0目录 重建干净结构：
  head(R1-R941: 表头+1-11章) + 12体质块 + 13平台块(旧块改名+13.5-13.7) + 说明文字 + 附篇块
彻底规避 openpyxl insert/delete 行操作陷阱。
"""
import openpyxl, re

SEED = r'C:\Users\DELL\Desktop\textbooks\中医基础理论final.xlsx'
DST = r'C:\Users\DELL\Desktop\qhzy-中医基础理论.xlsx'

# 1. 读种子 0目录 全部行
swb = openpyxl.load_workbook(SEED, read_only=True)
sws = swb['0目录']
seed_rows = []
for row in sws.iter_rows(max_col=5):
    seed_rows.append([('' if c.value is None else str(c.value).strip()) for c in row])
swb.close()
print("种子 0目录 max_row:", len(seed_rows))

# 种子结构定位
head_end = None   # R941 结束（第12章标题之前）
ch12_start = None  # 旧第12章标题行（0-indexed）
note_start = None  # "本版目录的显著特征"
for i, r in enumerate(seed_rows):
    e = r[4]
    if e and e.startswith('第12章') and '教学平台' in e:
        ch12_start = i
    if e and e.startswith('本版目录的显著特征'):
        note_start = i
        break
print(f"旧第12章标题 @ {ch12_start + 1} | 显著特征 @ {note_start + 1}")
head_end = ch12_start  # 0-indexed 切片端点（不含）
old12_block = [r for r in seed_rows[ch12_start:note_start] if any(r)]  # 旧第12章块（含标题，去空行）
note_block = [r for r in seed_rows[note_start:] if any(r)]  # 说明文字（去空行）
print(f"head: {head_end} 行 | old12块: {len(old12_block)} 行 | note块: {len(note_block)} 行")

# 2. 旧12章块 → 13章块：标题改名 + 12.x→13.x
new13 = []
for r in old12_block:
    nr = r[:]
    e = nr[4]
    if '第12章' in e and '教学平台' in e:
        nr[4] = '第13章\u3000岐黄智医分布式教学平台与评价体系'
    else:
        m = re.match(r'^12\.', e)
        if m:
            nr[4] = '13.' + e[3:]
    new13.append(nr)

# 3. qhzy 增补块（与 Step D 相同的 12体质 + 附篇 + 13.5-13.7 内容）
tizhi_block = [
    ['12', '00', '00', '00', '第12章 体质学说与体质辨识'],
    ['', '01', '00', '00', '12.1 体质的概念与形成'],
    ['', '', '01', '00', '12.1.1 体质的基本概念'],
    ['', '', '', '01', '12.1.1.1 体质的定义与内涵'],
    ['', '', '', '02', '12.1.1.2 体质与证候、病机的关系'],
    ['', '', '02', '00', '12.1.2 体质的形成因素'],
    ['', '', '', '01', '12.1.2.1 先天因素(父母禀赋与遗传)'],
    ['', '', '', '02', '12.1.2.2 后天因素(饮食、起居、情志与地理)'],
    ['', '02', '00', '00', '12.2 体质的分类'],
    ['', '', '01', '00', '12.2.1 以《灵枢·阴阳二十五人》为纲'],
    ['', '', '', '01', '12.2.1.1 五行五形人'],
    ['', '', '', '02', '12.2.1.2 五音二十五人'],
    ['', '', '02', '00', '12.2.2 王琦九分法(现代参考)'],
    ['', '', '', '01', '12.2.2.1 九种基本体质'],
    ['', '', '', '02', '12.2.2.2 经典与九分法的映射'],
    ['', '03', '00', '00', '12.3 体质与发病、辨证论治'],
    ['', '', '01', '00', '12.3.1 体质与发病倾向'],
    ['', '', '02', '00', '12.3.2 体质与辨证论治(因人制宜)'],
    ['', '04', '00', '00', '12.4 体质辨识的AI实现(分布式平台落点)'],
    ['', '', '01', '00', '12.4.1 体质辨识量表数字化'],
    ['', '', '02', '00', '12.4.2 体质-病证关联推理'],
]
# 13.5-13.7 追加到 new13 末尾（旧块已有 13.1-13.4）
new13 += [
    ['', '05', '00', '00', '13.5 岐黄智医分布式网络编程体系'],
    ['', '', '01', '00', '13.5.1 三服务器+GitHub+四台式机拓扑'],
    ['', '', '02', '00', '13.5.2 分布式任务调度'],
    ['', '06', '00', '00', '13.6 六者SOUL与医圣成长引擎对接'],
    ['', '', '01', '00', '13.6.1 六者角色与教材内容的映射'],
    ['', '', '02', '00', '13.6.2 医圣成长引擎'],
    ['', '07', '00', '00', '13.7 教材-图谱-平台三角色闭环'],
    ['', '', '01', '00', '13.7.1 本章知识树→图谱注入'],
    ['', '', '02', '00', '13.7.2 使用数据→教材修订'],
]
fupian_block = [
    ['附', '00', '00', '00', '附篇 五运六气'],
    ['', '01', '00', '00', '附1 五运六气的基本概念'],
    ['', '02', '00', '00', '附2 六气'],
    ['', '03', '00', '00', '附3 运气与发病'],
    ['', '04', '00', '00', '附4 运气学说的学习与数字化'],
]
# qhzy 增补说明行
qhzy_note = [['', '', '', '', '【qhzy 增补·2026-08】以下为第12章(体质)、第13章(岐黄智医分布式教学平台)、附篇(五运六气)目录条目:']]

# 4. 重组：head + 空行 + 12体质 + 13平台 + 说明 + 附篇
new_toc = []
new_toc.extend(seed_rows[:head_end])          # R1-R941（含表头与空行）
new_toc.append([''] * 5)                       # 空行
new_toc.extend(tizhi_block)                    # 第12章体质（21行）
new_toc.append([''] * 5)                       # 空行
new_toc.extend(new13)                          # 第13章平台（旧块+13.5-13.7）
new_toc.append([''] * 5)                       # 空行
new_toc.extend(note_block)                     # 显著特征+校准说明（修正引用）
new_toc.append([''] * 5)                       # 空行
new_toc.extend(qhzy_note)                      # qhzy 增补说明
new_toc.extend(fupian_block)                   # 附篇（5行）

# 修正说明文字中的引用：知识树落地 第12章→第13章；④ 引用
for i, r in enumerate(new_toc):
    e = r[4]
    if e and '知识树落地' in e and '第12章' in e:
        r[4] = e.replace('第12章独立成章', '第13章独立成章').replace('第12章', '第13章')
    if e and '本目录与规划教材的对应' in e:
        r[4] = e.replace("第12章'教学平台与评价体系'", "第13章'岐黄智医分布式教学平台与评价体系'").replace('第12章', '第13章')

# 5. 写入 qhzy（替换原 0目录 内容）
wb = openpyxl.load_workbook(DST)
ws = wb['0目录']
# 清空现有
for r in range(1, ws.max_row + 1):
    for c in range(1, 6):
        ws.cell(row=r, column=c, value=None)
for i, row in enumerate(new_toc):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=i + 1, column=c, value=v)
wb.save(DST)
print("Step J 完成, 新 0目录 行数:", len(new_toc))

# 6. 复核
wb2 = openpyxl.load_workbook(DST, read_only=True)
ws2 = wb2['0目录']
print("\n=== 章级顺序复核 ===")
for i, row in enumerate(ws2.iter_rows(min_row=1, max_col=5), 1):
    cells = ['' if x.value is None else str(x.value).strip() for x in row]
    e = cells[4]
    if e and len(e) < 30 and e.startswith('第') and '章' in e:
        print(f"R{i}: {e}")
print("\n=== 13.x 顺序抽查（R964-R990） ===")
cnt = 0
for i, row in enumerate(ws2.iter_rows(min_row=955, max_col=5), 955):
    e = row[4].value
    if e and isinstance(e, str) and e.strip().startswith('13.'):
        print(f"R{i}: {e.strip()[:45]}")
        cnt += 1
    if cnt >= 22:
        break
wb2.close()
