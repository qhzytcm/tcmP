# -*- coding: utf-8 -*-
"""强制清空 ch02 R33（索引赋值）+ 验证"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['2cmrl-bellman']
v33 = str(ws.cell(33, 3).value or '')
print(f'清空前 R33: ({len(v33)}字) {v33[:30]}')
ws['C33'] = None
wb.save(X)

wb2 = load_workbook(X)
ws2 = wb2['2cmrl-bellman']
v33b = str(ws2.cell(33, 3).value or '')
n_notes = sum(1 for r in range(1, ws2.max_row + 1)
              if str(ws2.cell(r, 3).value or '').strip().startswith('图注：'))
n_img = len(ws2._images)
wb2.close()
print(f'清空后 R33: ({len(v33b)}字)')
print(f'图注行数: {n_notes}, 图数: {n_img}')
