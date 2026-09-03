# -*- coding: utf-8 -*-
"""
Step H: 修复 0目录 第12/13章重复冲突
问题：种子旧块 R943"第12章 教学平台与评价体系"(12.1-12.4) 未改名，
      与 qhzy 新增 R1033"第12章 体质" 及 R1054"第13章" 冲突，且 13.1-13.4 重复。
方案：
  1. 旧块章标题改"第13章 岐黄智医分布式教学平台与评价体系"，12.x→13.x
  2. 删除新增重复块 R1054-R1067（13.1-13.7），其中 13.5-13.7 内容并入旧块尾部
  3. 旧说明文字中"第12章'教学平台'"引用改为第13章
"""
import openpyxl, re

DST = r'C:\Users\DELL\Desktop\qhzy-中医基础理论.xlsx'
wb = openpyxl.load_workbook(DST)
ws = wb['0目录']

# 1. 收集 13.5-13.7 新内容（来自 R1059-R1067，先读后删）
new_135 = []  # (a,b,c,d,e)
for r in range(1059, 1068):
    cells = [ws.cell(row=r, column=c).value for c in range(1, 6)]
    new_135.append(['' if v is None else str(v).strip() for v in cells])
print("捕获 13.5-13.7 行数:", len([x for x in new_135 if x[4]]))

# 2. 旧块章标题改名 + 12.x→13.x（R943-R1011）
for r in range(943, 1012):
    e = ws.cell(row=r, column=5).value
    if e is None:
        continue
    e = str(e)
    ne = e
    if '第12章' in ne and '教学平台' in ne:
        ne = ne.replace('第12章', '第13章')
    # 标题 12.x → 13.x
    for pat in [r'^12\.(\d+)\.(\d+)\.(\d+)\s', r'^12\.(\d+)\.(\d+)\s', r'^12\.(\d+)\s']:
        m = re.match(pat, ne)
        if m:
            ne = re.sub(r'^12\.', '13.', ne)
            break
    if ne != e:
        ws.cell(row=r, column=5, value=ne)
print("旧块已改名 12.x→13.x")

# 3. 删除新增重复块 R1054-R1067（13.1-13.7 完整块）
ws.delete_rows(1054, 14)
print("已删除重复第13章块 (R1054-R1067)")

# 4. 在旧 13.4 块尾部（原 R1011 位置，删除后仍为 R1011 之后的空行处）插入 13.5-13.7
#    删除后旧块尾部行号不变（删除发生在更后面），找到最后一个 13.4 标题行
last_134 = None
for r in range(940, 1030):
    e = ws.cell(row=r, column=5).value
    if e and isinstance(e, str) and e.startswith('13.4'):
        last_134 = r
print("13.4 最后一行:", last_134)
# 找到其后第一个空行
insert_at = last_134 + 1
while insert_at <= ws.max_row and ws.cell(row=insert_at, column=5).value not in (None, ''):
    insert_at += 1
print("插入位置:", insert_at)

# 在 insert_at 插入空行（13.5 有 9 行内容：标题+2节+6段）
rows_needed = 0
for x in new_135:
    if any(x):
        rows_needed += 1
ws.insert_rows(insert_at, rows_needed)
for i, x in enumerate([y for y in new_135 if any(y)]):
    for c in range(1, 6):
        if x[c - 1]:
            ws.cell(row=insert_at + i, column=c, value=x[c - 1])
print("13.5-13.7 已插入")

# 5. 旧说明文字引用修正（R1015-R1031 区域，删除后行号略变，按内容查找）
for r in range(940, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    if e and isinstance(e, str):
        ne = e
        if '第12章' in ne and '教学平台' in ne:
            ne = ne.replace('第12章', '第13章')
            ws.cell(row=r, column=5, value=ne)
            print(f"说明引用修正 R{r}: {ne[:50]}")

wb.save(DST)
print("Step H 完成")

# 复核
wb2 = openpyxl.load_workbook(DST, read_only=True)
ws2 = wb2['0目录']
chapters = []
for row in ws2.iter_rows(min_row=1, max_col=5):
    a, b, c, d, e = ['' if x.value is None else str(x.value).strip() for x in row]
    if e and ('第' in e) and ('章' in e) and not e.startswith(('①', '②', '③', '④', 'a.', 'b.', 'c.', 'd.', 'e.', '如果', '病证', 'AI', '知识', '纯', '本版', '【', '含章')):
        if len(e) < 40 and ('第' in e[:3]):
            chapters.append(e)
print("章级标题清单:")
for c in chapters:
    print("  ", c)
