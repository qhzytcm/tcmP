# -*- coding: utf-8 -*-
"""测试: read_only vs 常规模式读取 6.2.1.7"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'

# 常规模式
wb = openpyxl.load_workbook(PATH)
ws = wb['6.素问·经脉血气']
v = ws.cell(row=19, column=3).value
print(f"常规模式: len={len(v)} {v[:60]}")
wb.close()

# read_only 模式
wb2 = openpyxl.load_workbook(PATH, read_only=True)
ws2 = wb2['6.素问·经脉血气']
for row in ws2.iter_rows(min_row=19, max_row=19, max_col=3):
    v2 = row[2].value
    print(f"read_only: len={len(v2)} {str(v2)[:60]}")
wb2.close()
