# -*- coding: utf-8 -*-
"""Step F: 病证单元库去重 + 补齐 3 个新 BZU（凑齐 32 个唯一编码）"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中医基础理论.xlsx'
wb = openpyxl.load_workbook(DST)
ws = wb['病证单元库']

# 1. 收集所有 BZ-* 行，删除重复编码（保留首次出现）
seen = {}
to_delete = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and isinstance(v, str) and v.startswith('BZ-'):
        code = v
        if code in seen:
            to_delete.append(r)
        else:
            seen[code] = r

print("重复行(将删除):", [(r, ws.cell(row=r, column=1).value) for r in to_delete])
for r in sorted(to_delete, reverse=True):
    ws.delete_rows(r)

# 2. 追加 3 个新 BZU（唯一编码）
start = ws.max_row + 1
while start > 1:
    has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, 9))
    if has:
        break
    start -= 1

new_rows = [
    ['BZ-B08-Z11', '眩晕 MB51', '气滞血瘀证 Z11', '瘀血阻络, 清窍失养', '眩晕头痛、痛处固定、唇甲紫暗、舌紫暗有瘀斑、脉涩', '活血化瘀、通窍止眩', '通窍活血汤', '赤芍、川芎、桃仁、红花、老葱、生姜、红枣、麝香(代用: 白芷)'],
    ['BZ-B12-Z05', '泄泻 ME05', '脾胃气虚证 Z5', '脾胃虚弱, 运化失职, 清浊不分', '大便溏泄、食少腹胀、神疲乏力、面色萎黄、舌淡苔白', '健脾益气、渗湿止泻', '参苓白术散', '人参、白术、茯苓、甘草、山药、莲子肉、薏苡仁、砂仁、桔梗、扁豆'],
    ['BZ-B13-Z12', '便秘 ME06', '胃阴虚证 Z12', '胃阴亏耗, 津液不足, 肠失濡润', '大便干结、口干舌燥、饥不欲食、舌红少津、脉细数', '滋阴增液、润肠通便', '增液汤', '玄参、麦冬、生地'],
]
for i, row in enumerate(new_rows):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=start + i, column=c, value=v)

wb.save(DST)

# 3. 复核唯一编码数
wb2 = openpyxl.load_workbook(DST, read_only=True)
ws2 = wb2['病证单元库']
codes = set()
for row in ws2.iter_rows(max_col=1):
    v = row[0].value
    if v and isinstance(v, str) and v.startswith('BZ-'):
        codes.add(v)
print("唯一 BZU 编码数:", len(codes))
print("清单:")
for c in sorted(codes):
    print("  ", c)
