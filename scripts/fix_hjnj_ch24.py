# -*- coding: utf-8 -*-
"""修复: 第24章注释补「非零和」标注（异说并存处）"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
ws = wb['24.灵枢·脏腑身形']

# 每条注释补「非零和」: 在含"或异/或有出入/存异"处补"并存（非零和局）"
# 具体: 在"悬疑存疑"前/后补, 或在结尾补
fixes = {
    '24.1.1.10': None,  # 自动处理
    '24.2.1.4': None,
    '24.3.1.4': None,
    '24.4.1.7': None,
    '24.5.1.8': None,
    '24.6.1.8': None,
    '24.7.1.14': None,
}

n = 0
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value == '注释' and row[1].value in fixes:
        v = str(row[2].value)
        if '非零和' not in v:
            # 在结尾"（通行）"后补 或 在"悬疑存疑"处补
            if v.endswith('（通行）。'):
                new = v[:-4] + '；并存（非零和局）。'
            elif '悬疑存疑' in v:
                new = v.replace('悬疑存疑', '悬疑存疑（异说并存，非零和局）')
            else:
                new = v.rstrip('。') + '；并存（非零和局）。'
            ws.cell(row=row[0].row, column=3, value=new)
            print(f"{row[1].value}: 已补非零和 (len {len(v)}→{len(new)})")
            n += 1

wb.save(PATH)
print(f"修复 {n} 条")
