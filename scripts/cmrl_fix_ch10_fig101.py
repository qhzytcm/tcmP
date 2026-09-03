# -*- coding: utf-8 -*-
"""ch10 fig10-1 布局修正: 图题R5/图体R6(图片锚)/图注R7
insert_rows(6,1) + 锚点>=6 全部+1 + R6 清空（图片覆盖行）"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['10cmrl-acloop']

# 1) 插入 1 行 R6 前
ws.insert_rows(6, 1)

# 2) 全部锚点 >= 6 的 +1（fig10-1 锚 R5 -> R6, 其余 +1）
for img in ws._images:
    if img.anchor._from.row + 1 >= 6:
        img.anchor._from.row += 1

# 3) R6 清空（图体行, 图片覆盖）
ws.cell(6, 3, None)
wb.save(X)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['10cmrl-acloop']
anchors = sorted(img.anchor._from.row + 1 for img in ws2._images)
cap5 = str(ws2.cell(5, 3).value or '')
note7 = str(ws2.cell(7, 3).value or '')
n_notes = sum(1 for r in range(1, ws2.max_row + 1)
              if str(ws2.cell(r, 3).value or '').strip().startswith('图注：'))
wb2.close()
print(f'锚点: {anchors}')
print(f'图题 R5: {cap5[:18]}')
print(f'图注 R7: {len(note7)}字')
print(f'图注行数: {n_notes}')
