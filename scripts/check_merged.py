# -*- coding: utf-8 -*-
"""检查 6.素问·经脉血气 合并单元格"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
ws = wb['6.素问·经脉血气']

print(f"合并区域数: {len(ws.merged_cells.ranges)}")
for rng in ws.merged_cells.ranges:
    print(f"  {rng}")
    if rng.min_row <= 19 <= rng.max_row and rng.min_col <= 3 <= rng.max_col:
        print(f"  ★ R19C3 在合并区域内!")

# 也检查 R19 单元格属性
c = ws.cell(row=19, column=3)
print(f"\nR19C3: value={repr(c.value)[:50]}")
print(f"  data_type={c.data_type}, number_format={c.number_format}")
print(f"  行高: {ws.row_dimensions[19].height}")
wb.close()
