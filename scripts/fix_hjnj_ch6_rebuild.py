# -*- coding: utf-8 -*-
"""
终极修复: 用脚本源重建第6章 sheet（绕开 R19C3 单元格损坏）
gen_hjnj_ch06.py 数据(含通读补丁) + 注释标注补丁 + 表头
"""
import importlib.util, re
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'

# 1. 从 gen_hjnj_ch06.py 提取 rows（脚本源）
with open(r'C:\Users\DELL\tcmP\scripts\gen_hjnj_ch06.py', encoding='utf-8') as f:
    src = f.read()
# 替换保存调用防止干扰（仅此一处）
src = src.replace('wb.save(DST)', 'pass  # 拦截')
ns = {}
exec(compile(src, '<gen>', 'exec'), ns)
rows = ns['rows']
print(f"脚本源 rows: {len(rows)} 行")

# 2. 检查 6.2.1.7 在脚本源中的完整性
for cat, code, text in rows:
    if code == '6.2.1.7':
        print(f"脚本源 6.2.1.7: len={len(text)} {'✔' if '肾欲坚' in text else '✘'}")
        break

# 3. 打开工作簿, 删除旧第6章, 重建
wb = openpyxl.load_workbook(PATH)
SHEET = '6.素问·经脉血气'
idx = wb.sheetnames.index(SHEET)
del wb[SHEET]
ws = wb.create_sheet(SHEET, idx)  # 插回原位置

ws.cell(row=1, column=1, value='类型')
ws.cell(row=1, column=2, value='四级编码')
ws.cell(row=1, column=3, value='内容（原文/注释）')
for r, (cat, code, text) in enumerate(rows, 2):
    ws.cell(row=r, column=1, value=cat)
    ws.cell(row=r, column=2, value=code)
    ws.cell(row=r, column=3, value=text)

# 4. 补注释标注（悬疑/非零和, 与全书一致）
n = 0
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value == '注释' and row[2].value:
        v = str(row[2].value)
        tail = []
        if '悬疑' not in v:
            tail.append('难解处悬疑存疑')
        if '非零和' not in v:
            tail.append('异说并存（非零和局）')
        if tail:
            ws.cell(row=row[0].row, column=3, value=v.rstrip('。') + '；' + '；'.join(tail) + '。')
            n += 1
print(f"补注释标注: {n} 条")

wb.save(PATH)
wb.close()

# 5. 验证
wb2 = openpyxl.load_workbook(PATH, read_only=True)
ws2 = wb2['6.素问·经脉血气']
for row in ws2.iter_rows(min_row=2, max_col=3):
    if row[1].value == '6.2.1.7':
        v = str(row[2].value)
        print(f"验证 6.2.1.7: len={len(v)} {'✔完整' if '肾欲坚' in v else '✘仍截断'}")
wb2.close()
print(f"Sheet 位置: {openpyxl.load_workbook(PATH).sheetnames.index('6.素问·经脉血气')}")
