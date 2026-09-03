# -*- coding: utf-8 -*-
"""
决定性重建 v2: 全簿读取 + 正文章用 gen_hjnj_ch01~22.py 源覆盖 + 写入全新文件 + 替换
（解决 sharedStrings 截断问题: 20.1.1.13 / 22.1.1.11 等）
"""
import os, re, shutil
import openpyxl

SRC = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
DST = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版)_rebuilt2.xlsx'

# 1. 读取原文件所有 sheet 数据
wb_old = openpyxl.load_workbook(SRC, read_only=True)
all_data = {}
for s in wb_old.sheetnames:
    ws = wb_old[s]
    rows = []
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if any(v is not None for v in vals):
            rows.append(vals)
    all_data[s] = rows
wb_old.close()
print(f"读取 {len(all_data)} 个 sheet")

# 2. 从 gen 脚本源提取正文章完整文本（01-22）
def extract_rows(src_text):
    src_text = src_text.replace("wb.save(DST)", "pass  # 拦截")
    src_text = src_text.replace("wb.save(PATH)", "pass  # 拦截")
    ns = {}
    exec(compile(src_text, '<gen>', 'exec'), ns)
    return ns.get('SHEET'), ns.get('rows', [])

script_fixes = {}
for i in range(1, 23):
    src_path = rf'C:\Users\DELL\tcmP\scripts\gen_hjnj_ch{i:02d}.py'
    if not os.path.exists(src_path):
        continue
    with open(src_path, encoding='utf-8') as f:
        src_text = f.read()
    try:
        sheet_name, rows = extract_rows(src_text)
        if sheet_name:
            script_fixes[sheet_name] = [[cat, code, text] for cat, code, text in rows]
            print(f"  脚本覆盖: {sheet_name} ({len(rows)} 行)")
    except Exception as e:
        print(f"  ✘ {src_path}: {e}")

# 3. 创建新工作簿, 按规范顺序写入
wb_new = openpyxl.Workbook()
wb_new.remove(wb_new.active)

ORDER = ['0改写说明','0分布式体系','平台建设与依赖','平台支撑完整性','0历史',
         '0映射1','0映射2','0映射3','0封面','0目录',
         'ICD11统一命名','病证单元库','多智能体协同','兼容性思维流','可复用Skills']
ch_sheets = sorted([s for s in all_data if re.match(r'^\d+\.', s)],
                   key=lambda x: int(x.split('.')[0]))
ORDER += ch_sheets

for s in ORDER:
    if s not in all_data:
        print(f"  ✘ 缺 sheet: {s}")
        continue
    ws = wb_new.create_sheet(s)
    if s in script_fixes:
        # 正文章: 脚本源 + 表头
        ws.cell(row=1, column=1, value='类型')
        ws.cell(row=1, column=2, value='四级编码')
        ws.cell(row=1, column=3, value='内容（原文/注释）')
        for r, row in enumerate(script_fixes[s], 2):
            for c, v in enumerate(row, start=1):
                if v:
                    ws.cell(row=r, column=c, value=v)
    else:
        for r, row in enumerate(all_data[s], 1):
            for c, v in enumerate(row, start=1):
                if v:
                    ws.cell(row=r, column=c, value=v)

wb_new.save(DST)
print(f"\n重建完成: {DST}")

# 4. 验证修复点
wb_check = openpyxl.load_workbook(DST, read_only=True)
for s, code, kw in [
    ('20.灵枢·经脉经别', '20.1.1.13', '不盛不虚，以经取之'),
    ('22.灵枢·营卫气血', '22.1.1.11', '上焦如雾'),
]:
    ws = wb_check[s]
    for row in ws.iter_rows(min_row=2, max_col=3):
        if row[1].value == code:
            v = str(row[2].value)
            print(f"{code} len={len(v)} {'✔完整' if kw in v else '✘仍截断'}")
wb_check.close()

# 5. 替换
shutil.move(DST, SRC)
print(f"已替换: {SRC}")
