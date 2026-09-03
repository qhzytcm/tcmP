# -*- coding: utf-8 -*-
"""差异测试: C20(空) vs C19(先清空再写) vs 新sheet"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
FULL = '病在肾，愈在春；春不愈，甚于长夏；长夏不死，持于秋；起于冬。禁犯焠㶼热食温炙衣。肾病者，夜半慧，四季甚，下晡静。肾欲坚，急食苦以坚之，用苦补之，咸写之。'

wb = openpyxl.load_workbook(PATH)
ws = wb['6.素问·经脉血气']

# 1. 写 C20（空）
ws.cell(row=20, column=3, value=FULL)
print("已写 C20")

# 2. 清空 C19 再写
ws.cell(row=19, column=3, value=None)
ws.cell(row=19, column=3, value=FULL)
print("已清空并重写 C19")

wb.save(PATH)
wb.close()

wb2 = openpyxl.load_workbook(PATH)
ws2 = wb2['6.素问·经脉血气']
for r in [19, 20]:
    v = ws2.cell(row=r, column=3).value
    ok = '✔完整' if v and len(v) == len(FULL) else '✘截断'
    print(f"R{r}: len={len(v) if v else 0} {ok}")
    if v:
        print(f"  {v[:40]}")
wb2.close()
