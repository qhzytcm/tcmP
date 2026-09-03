# -*- coding: utf-8 -*-
"""canonical test — cmrl 教材全书统一验证入口（项目规范测试命令）
运行: python scripts/canonical_test.py   (Anaconda python: 含 openpyxl/matplotlib)
覆盖: 终检 3/3 / 正文 verify / 图注 45 同步+e引导+字数 / 字号规范 / 知识树树形 / 封面页
退出码: 0=全绿 1=有失败
"""
import re
import subprocess
import sys
from pathlib import Path

ROOT = Path(r'C:\Users\DELL\tcmP')
TEXTBOOK = Path(r'C:\Users\DELL\textbook-project')
SCRIPTS = TEXTBOOK / 'scripts'
D = TEXTBOOK / 'drafts' / 'cmrl'
X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
SHEETS = ['1cmrl-conceptions', '2cmrl-bellman', '3cmrl-optimality', '4cmrl-iteration',
          '5cmrl-montecarlo', '6cmrl-approximation', '7cmrl-tdlearning',
          '8cmrl-valuefunction', '9cmrl-policygradient', '10cmrl-acloop']

results = []


def check(name, ok, detail=''):
    results.append((name, ok, detail))
    print(f'{"PASS" if ok else "FAIL"}  {name}  {detail}')


# ---------- T1 全书图表终检 3/3 ----------
out = subprocess.run([sys.executable, str(SCRIPTS / 'cmrl_fig_final_check.py')],
                     capture_output=True, text=True, encoding='utf-8', timeout=600)
check('T1 图表终检 3/3 (45/45 哈希)', out.returncode == 0 and '通过 (3/3)' in out.stdout,
      f'rc={out.returncode}')

# ---------- T2 正文 verify ----------
out2 = subprocess.run([sys.executable, str(SCRIPTS / 'cmrl_verify.py'), '--all'],
                      capture_output=True, text=True, encoding='utf-8', timeout=600)
check('T2 正文 verify --all', out2.returncode == 0 and 'CMRL VERIFY: PASS' in out2.stdout,
      f'rc={out2.returncode}')

# ---------- T3 图注 45/45 同步 + e 引导 + 字数 ----------
from openpyxl import load_workbook
wb = load_workbook(X)
total = 0
no_e = []
over = []
mismatch = []
for ch in range(1, 11):
    notes = []
    for f in sorted((D / f'ch{ch:02d}').glob('0*-1.*.md')):
        for ln in f.read_text(encoding='utf-8').splitlines():
            if ln.strip().startswith('图注：'):
                notes.append(ln.strip())
    ws = wb[SHEETS[ch - 1]]
    imgs = sorted(ws._images, key=lambda i: i.anchor._from.row)
    for i, img in enumerate(imgs):
        if i >= len(notes):
            break
        row = img.anchor._from.row + 2
        v = str(ws.cell(row, 3).value or '')
        total += 1
        if v[:30] != notes[i][:30]:
            mismatch.append(f'ch{ch}R{row}')
    for n in notes:
        if '阅读引导' not in n:
            no_e.append(f'ch{ch}')
        if len(n) > 650:
            over.append(f'ch{ch}:{len(n)}')
check('T3 图注 45/45 md↔Excel 一致', total == 45 and not mismatch,
      f'共{total} 不一致={mismatch[:3] or "无"}')
check('T4 图注均含 e.阅读引导', not no_e, f'缺={no_e[:3] or "无"}')
check('T5 图注字数 ≤650', not over, f'超限={over[:3] or "无"}')

# ---------- T6 字号规范 (脚本最小字号 >=9) ----------
small = []
for p in (D / 'figures').glob('gen_ch*_figures_feynman.py'):
    t = p.read_text(encoding='utf-8', errors='ignore')
    sizes = [int(x) for x in re.findall(r'fontsize[=:]\s*(\d+)', t)]
    if sizes and min(sizes) < 9:
        small.append(f'{p.name}:{min(sizes)}')
check('T6 字号规范: 脚本最小字号 ≥9', not small, f'异常={small[:3] or "无"}')

# ---------- T7 知识树树形 ----------
t7 = Path(r'C:\Users\DELL\tcmP\scripts\cmrl_knowledge_tree.py').read_text(encoding='utf-8')
check('T7 知识树: 树形结构(主干/主枝/叶冠/须底)',
      'Polygon' in t7 and 'Ellipse' in t7 and 'box(50, 3.5, XU' in t7, '')

# ---------- T8 封面页 ----------
ws_cov = wb['封面页']
n_img = len(ws_cov._images)
anchors = [img.anchor._from.row + 1 for img in ws_cov._images]
note_cov = str(ws_cov.cell(41, 1).value or '')
cap_cov = str(ws_cov.cell(39, 1).value or '')
check('T8 封面页: 树形图锚R40 + 图题 + 自下而上',
      n_img == 1 and anchors == [40] and cap_cov.startswith('图 K-1')
      and '自下而上' in note_cov,
      f'图={n_img} 锚={anchors} 图注{len(note_cov)}字')
wb.close()

fails = [r for r in results if not r[1]]
print(f'\n=== CANONICAL TEST: {"ALL PASS" if not fails else f"{len(fails)} FAILED"} '
      f'({len(results) - len(fails)}/{len(results)}) ===')
sys.exit(0 if not fails else 1)
