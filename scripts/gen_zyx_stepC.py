# -*- coding: utf-8 -*-
"""
Step C: qhzy-中药学 引擎Sheet 增补 + 0改写说明升级 + 平台建设与依赖增补
1. ICD11统一命名: 增补平台db校准列（错码修正）
2. 病证单元库: 增补 脏腑/六经/舌脉 三列（26 BZU）
3. 多智能体协同: 新增 node-coordinator（8→9）
4. 兼容性思维流: 新增第7局（算力调度）
5. 可复用Skills: 新增 tcm-distributed-node-orchestration（6→7）
6. 0改写说明: 版本 v1.0 → v2.0-qhzy
7. 平台建设与依赖: 追加 E 节 分布式适配
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中药学.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=6):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# ═══ 1. ICD11统一命名 校准 ═══
ws = wb['ICD11统一命名']
hdr_row = None
for r in range(1, 6):
    v1 = ws.cell(row=r, column=1).value
    v2 = ws.cell(row=r, column=2).value
    if v1 and '平台病码' in str(v1):
        hdr_row = r
        break
    if v2 and '统一命名' in str(v2):
        hdr_row = r
        break
print("ICD11表头行:", hdr_row)
calib = {
    'B1': ('CA00', 'Acute nasopharyngitis', '✅ 一致'),
    'B2': ('MD12', 'Cough', '✅ 一致'),
    'B3': ('CA23', 'Asthma', '✅ 一致'),
    'B4': ('MD11.5', 'Dyspnoea', '🔧 细化: MD11.5(原MD11大类)'),
    'B5': ('MC81.2', 'Palpitations', '🔧 细化: MC81.2(原MC81大类)'),
    'B6': ('BA40', 'Angina pectoris', '✅ 一致'),
    'B7': ('7A00', 'Chronic insomnia', '✅ 一致'),
    'B8': ('MB48.0', 'Vertigo', '🔧 修正: 原MB51=上肢瘫痪; 正确MB48.0 Vertigo'),
    'B9': ('8B11', 'Cerebral ischaemic stroke', '🔧 修正: 原8B20=未特指卒中; 缺血性=8B11'),
    'B10': ('MD81', 'Abdominal pain', '🔧 修正: 原MD90=恶心呕吐; 腹痛=MD81'),
    'B11': ('MD90', 'Nausea or vomiting', '🔧 修正: 原ME08=胃肠胀气; 恶心呕吐=MD90'),
    'B12': ('DD91.2', 'Functional diarrhoea', '🔧 修正: 原ME05=排便习惯改变; 功能性腹泻=DD91.2'),
    'B13': ('DD91.1', 'Functional constipation', '🔧 修正: 原ME06=慢性肠炎; 功能性便秘=DD91.1'),
    'B14': (None, '(待WHO-API回填)', '⏳ pending_who_api'),
    'B15': ('MG27', 'Oedema', '⚠️ db中MG27=出血未归类, 水肿码待复核'),
    'B16': ('5A11', 'Type 2 diabetes mellitus', '✅ 一致'),
    'B17': (None, '(待WHO-API回填)', '⏳ pending_who_api'),
    'B18': ('GC00', 'Cystitis', '✅ 一致'),
    'B19': ('ME84.2', 'Low back pain', '🔧 修正: 原ME83=风湿; 下背痛=ME84.2'),
    'B20': ('8A81', 'Tension-type headache', '🔧 细化: 原8A80=偏头痛; 头痛按型8A80/8A81/8A83'),
}
if hdr_row:
    ws.cell(row=hdr_row, column=6, value='平台db校准(icd11_mms.db 2026-01)')
    n = 0
    for r in range(hdr_row + 1, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        if code and str(code).strip() in calib:
            new_code, title, status = calib[str(code).strip()]
            ws.cell(row=r, column=6, value=f"{status} | db权威: {title}" + (f" | 建议码: {new_code}" if new_code else ""))
            n += 1
    print(f"ICD11 校准列写入 {n} 行")

# ═══ 2. 病证单元库 增补 3 列 ═══
ws = wb['病证单元库']
bzu_hdr = None
for r in range(1, 6):
    v = ws.cell(row=r, column=1).value
    if v and str(v).strip() == 'BZU编码':
        bzu_hdr = r
        break
print("BZU表头行:", bzu_hdr)
if bzu_hdr:
    ws.cell(row=bzu_hdr, column=9, value='脏腑定位(zangfu)')
    ws.cell(row=bzu_hdr, column=10, value='六经定位(six_channels)')
    ws.cell(row=bzu_hdr, column=11, value='舌脉(tongue_pulse)')
    zf_map = {
        'BZ-B01-Z01': ('肺·卫表', '太阳病', '舌淡红苔薄白, 脉浮紧'),
        'BZ-B01-Z02': ('肺·卫表', '太阳病', '舌尖红苔薄黄, 脉浮数'),
        'BZ-B02-Z07': ('肺·脾', '太阴病', '舌淡苔白腻, 脉滑'),
        'BZ-B02-Z15': ('肝·肺', '不适用', '舌红苔薄黄, 脉弦数'),
        'BZ-B03-Z07': ('肺·脾', '太阴病', '舌淡苔白滑, 脉滑'),
        'BZ-B04-Z07': ('肺·脾', '太阴病', '舌淡苔白腻, 脉滑'),
        'BZ-B05-Z08': ('心·脾', '太阴病', '舌淡苔白, 脉细弱'),
        'BZ-B05-Z14': ('心·血脉', '不适用', '舌紫暗, 脉涩'),
        'BZ-B06-Z14': ('心·血脉', '不适用', '舌紫暗有瘀斑, 脉涩'),
        'BZ-B06-Z11': ('心·肝', '不适用', '舌紫暗, 脉弦涩'),
        'BZ-B06-Z05': ('心·脾', '太阴病', '舌淡紫, 脉虚涩'),
        'BZ-B07-Z08': ('心·脾', '太阴病', '舌淡苔薄白, 脉细弱'),
        'BZ-B08-Z04': ('肝·肾', '不适用', '舌红苔黄, 脉弦数有力'),
        'BZ-B08-Z05': ('脾·心', '太阴病', '舌淡苔白, 脉细弱'),
        'BZ-B09-Z04': ('肝·肾', '不适用', '舌红苔黄, 脉弦有力'),
        'BZ-B10-Z03': ('肝·胃', '少阳病', '舌淡红苔薄白, 脉弦'),
        'BZ-B10-Z05': ('脾·胃', '太阴病', '舌淡苔白, 脉虚弱'),
        'BZ-B10-Z12': ('胃', '不适用', '舌红少津, 脉细数'),
        'BZ-B11-Z03': ('肝·胃', '少阳病', '舌淡红苔薄白, 脉弦'),
        'BZ-B12-Z06': ('脾·大肠', '不适用', '舌红苔黄腻, 脉滑数'),
        'BZ-B13-Z05': ('脾·肺', '太阴病', '舌淡苔白, 脉虚'),
        'BZ-B15-Z09': ('肾·脾', '少阴病', '舌淡胖苔白滑, 脉沉迟'),
        'BZ-B16-Z10': ('肾', '少阴病', '舌红少苔, 脉细数'),
        'BZ-B18-Z06': ('膀胱·下焦', '不适用', '舌红苔黄腻, 脉滑数'),
        'BZ-B19-Z13': ('肾·腰府', '不适用', '舌淡苔白腻, 脉沉缓'),
        'BZ-B20-Z04': ('肝', '不适用', '舌红苔薄黄, 脉弦'),
    }
    for r in range(bzu_hdr + 1, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        if code and str(code).strip() in zf_map:
            zf, six, tp = zf_map[str(code).strip()]
            ws.cell(row=r, column=9, value=zf)
            ws.cell(row=r, column=10, value=six)
            ws.cell(row=r, column=11, value=tp)
    print("BZU 三列已补齐")

# ═══ 3. 多智能体协同 node-coordinator ═══
append_rows('多智能体协同', [
    ['9', 'node-coordinator 分布式节点协调器(新增)', '分布式任务调度、就近调度、算力负载均衡(含A800 GPU调度)、CI/CD门禁对接', '不产出教材正文, 无节点则任务悬置', '算力←2云+2硬+4台式机(A800训练); 版本←GitHub', '任务(章节编写/药材图像采集/模型训练/图谱注入) → 就近节点调度 + 状态回传', '只读任务队列与节点状态; 对接GitHub Actions', '任何节点不得"赢家通吃"; 调度须给出就近理由; A800训练任务与轻量任务错峰'],
    ['协作协议(qhzy增补): ⑤任务分发一律经 node-coordinator: 章节编写→台式机/浪潮硬②; 药材图像采集→台式机T1-T4; 模型训练(A800)→浪潮硬②; 批量生成→华为硬①容器; 公网服务→华为云①; 版本门禁→GitHub。⑥节点状态60s心跳上报, 调度决策留痕。', '', '', '', '', '', '', ''],
], ncols=8)

# ═══ 4. 兼容性思维流 第7局 ═══
append_rows('兼容性思维流', [
    ['7', '本地 vs 云端 vs 台式机(算力调度, 含A800)', '台式机抢服务器算力/没有GPU就没用', '就近调度、所长相加: 轻量任务就近台式机, 批量上华为硬①容器, A800重训练上浪潮硬②, 公网服务上华为云①', '算力池化: 2云(华为云+GitHub) + 2硬(华为硬+浪潮硬含A800) + 4台式机 各司其职', '0分布式体系 Sheet + node-coordinator 调度协议'],
    ['', '', '', '', '', ''],
    ['永久协议⑤(qhzy增补): 分布式节点间同样适用"兼容表示优先"——节点冲突先找就近调度/错峰方案, 不得宣布某节点"无用"而淘汰之。', '', '', '', '', ''],
], ncols=6)

# ═══ 5. 可复用Skills 新增 ═══
append_rows('可复用Skills', [
    ['tcm-distributed-node-orchestration', '岐黄智医分布式节点编排(中药学版): 华为云①/GitHub/华为硬①/浪潮硬②(含A800)/四台式机的任务就近调度、药材图像采集、A800模型训练调度、CI/CD门禁对接。', '用户要求"分布式部署中药学平台/药材图像采集/A800模型训练/多节点调度"时使用。', '① 读0分布式体系Sheet拓扑; ② 任务类型→节点映射(图像采集→台式机, 模型训练A800→浪潮硬②, 批量→华为硬①, 公网→华为云①); ③ 对接GitHub Actions门禁; ④ 心跳监控与故障降级; ⑤ A800训练任务错峰调度。', '浪潮地址192.168.0.102单址; 台式机IP待定不臆造; 药材图像数据须脱敏并标注来源; A800算力任务须声明GPU占用。', 'deploy-v3.sh健康检查200; 节点状态表与实机一致; A800训练日志可复盘。'],
], ncols=6)

# ═══ 6. 0改写说明 升级 ═══
ws = wb['0改写说明']
for r in range(1, ws.max_row + 1):
    for c in range(1, 5):
        v = ws.cell(row=r, column=c).value
        if v and isinstance(v, str):
            nv = v
            if 'v1.0' in nv and '2026-08-07' in nv:
                nv = nv.replace('v1.0', 'v2.0-qhzy')
            if nv != v:
                ws.cell(row=r, column=c, value=nv)
append_rows('0改写说明', [
    ['版本变更', 'v2.0-qhzy (2026-08-08): ① 新增0分布式体系Sheet(2云+2硬+4台式机拓扑, 浪潮硬含A800 GPU); ② 第30章新增30.4节(岐黄智医分布式教学平台+中药数据闭环); ③ ICD-11命名表增补平台db校准列(修正错码MB48.0/8B11/MD81/MD90/DD91.x/ME84.2); ④ 病证单元库增补脏腑/六经/舌脉三列; ⑤ 多智能体协同新增node-coordinator(A800训练调度); ⑥ 兼容性思维流新增第7局; ⑦ 可复用Skills新增分布式编排规格。'],
], ncols=2)

# ═══ 7. 平台建设与依赖 E 节 ═══
append_rows('平台建设与依赖', [
    ['E. 岐黄智医分布式适配(qhzy v2.0增补)', '', '', '', '', ''],
    ['节点', '中药学平台承载', '数据流', '', '', ''],
    ['华为云服务器 114.115.211.254', '公网入口: 看板/SageAPI/PWA', '智能鉴别/药性预测API对外', '', '', ''],
    ['GitHub qhzytcm/tcmP', '版本+CI/CD门禁(G1/G2/G3)', '推送→门禁→部署', '', '', ''],
    ['华为硬服务器 192.168.1.12', 'Gateway调度+容器批量生成', '287味药正文批量生成', '', '', ''],
    ['浪潮硬服务器 192.168.0.102 (A800)', '计算后端: 图谱+Embedding+图像识别/药性预测模型训练', '药材图像/成分数据训练', '', '', ''],
    ['台式机T1-T4', 'L4作者/审稿 + 药材饮片图像采集', '图像/指纹图谱→浪潮A800训练', '', '', ''],
], ncols=6)

wb.save(DST)
print("Step C 完成")
print("Sheet 数:", len(wb.sheetnames))
