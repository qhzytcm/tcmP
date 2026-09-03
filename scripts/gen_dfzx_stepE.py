# -*- coding: utf-8 -*-
"""
Step E: qhzy-东方哲学概论 完善（深检后的两处小补充）
1. 0目录 补充 6.4 学时归属说明（全书30学时含6.4, 无需另计）
2. 7附件C 伦理规范与分布式采集衔接注释（台式机T1-T4采集数据遵循本规范）
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-东方哲学概论.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=3):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# 1. 0目录 学时归属说明
append_rows('0目录', [
    ['6.4 岐黄智医分布式教学平台（qhzy 增补·学时含于全书30学时内, 无需另计）'],
], ncols=1)
print("0目录 6.4 学时归属已补充")

# 2. 7附件C 衔接注释
append_rows('7附件', [
    ['C. 人机闭环实验伦理规范 · 分布式采集衔接(qhzy v2.0增补)'],
    ['分布式采集 | 台式机T1-T4采集的七日感知/节律追踪/口述史访谈数据, 一律遵循本规范: 知情同意(采集前说明用途)、数据脱敏(去除可识别身份字段)、治未病式风险预防(发现异常指标即提示人工复核而非自动干预)、全程日志可审计(对接0分布式体系 node-coordinator 调度留痕)。'],
], ncols=1)
print("7附件C 分布式衔接已补充")

wb.save(DST)
print("Step E 完成")
