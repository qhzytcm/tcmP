# -*- coding: utf-8 -*-
"""修复 ch10 图题行字体（Times New Roman -> 黑体, 对齐保持 center）"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['10cmrl-acloop']
fixed = 0
for r in range(1, ws.max_row + 1):
    c = ws.cell(r, 3)
    v = str(c.value or '')
    if v.startswith('图 ') and c.font.name != '黑体':
        c.font = Font(name='黑体', size=12, bold=False)
        c.alignment = Alignment(horizontal='center', vertical='center')
        fixed += 1
wb.save(X)
print(f'ch10 修复 {fixed} 个图题行字体')

# 校验: 全书图题行字体扫描
wb2 = load_workbook(X)
bad = []
for sn in wb2.sheetnames:
    if not sn[0].isdigit():
        continue
    ws2 = wb2[sn]
    for r in range(1, ws2.max_row + 1):
        c = ws2.cell(r, 3)
        v = str(c.value or '')
        if v.startswith('图 ') and c.font.name != '黑体':
            bad.append(f'{sn}R{r}:{c.font.name}')
wb2.close()
print(f'全书非黑体图题行: {bad if bad else "无"}')
