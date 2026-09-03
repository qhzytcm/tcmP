# -*- coding: utf-8 -*-
"""
Step D: qhzy-中医诊断学 收尾
1. 0目录 追加 14.4/14.5 条目（对齐正文新增节）
2. 0封面 增补 CM-02 元数据卡
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中医诊断学.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=5):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# ═══ 1. 0目录 追加 14.4/14.5 ═══
# 诊断学 0目录 格式: 章|节|小节|段|内容（数字序号，非00格式）
append_rows('0目录', [
    ['', '', '', '', '【qhzy 增补·2026-08】以下为第14章新增节条目:'],
    ['14', '4', '', '', '14.4 岐黄智医分布式教学平台'],
    ['', '', '1', '', '14.4.1 2云+2硬+4台式机拓扑与数据流'],
    ['', '', '', '1', '14.4.1.1 节点分工'],
    ['', '', '', '2', '14.4.1.2 诊断数据闭环'],
    ['', '', '2', '', '14.4.2 六者SOUL与诊断学对接'],
    ['', '', '', '1', '14.4.2.1 医者(核心枢纽)'],
    ['', '', '', '2', '14.4.2.2 患者/药者/规者'],
    ['14', '5', '', '', '14.5 四诊客观化数据闭环'],
    ['', '', '1', '', '14.5.1 望诊图像数据库'],
    ['', '', '', '1', '14.5.1.1 舌象/面色标注流水线'],
    ['', '', '2', '', '14.5.2 脉象与问诊数据'],
    ['', '', '', '1', '14.5.2.1 脉象波形库与十问歌数字化'],
], ncols=5)
print("0目录 14.4/14.5 条目已追加")

# ═══ 2. 0封面 增补 CM-02 元数据卡 ═══
ws = wb['0封面']
start = ws.max_row + 2
cover_rows = [
    ['', '', '【qhzy 增补·2026-08】教材元数据卡(分布式体系版)', '', ''],
    ['', '', '教材编号', 'CM-02', ''],
    ['', '', '教材名称', '中医诊断学(岐黄智医版)', ''],
    ['', '', '所属领域', 'D01 中医学院', ''],
    ['', '', '教学层级', '本科·基础经典(第二层)', ''],
    ['', '', '学分/学时', '5 学分', ''],
    ['', '', '前置要求', 'CM-01 中医基础理论', ''],
    ['', '', '后续衔接', 'MM-01 临床中药学 → MM-02 方剂学 → CM-13 内科学', ''],
    ['', '', '分布式承载', '台式机T1-T4四诊数据采集 · GitHub CI/CD门禁 · 华为云公网展示 · 浪潮硬②图谱/模型', ''],
    ['', '', '版本', 'v2.0-qhzy · 2026-08 · 由种子 v1.0 适应性修改', ''],
]
for i, row in enumerate(cover_rows):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=start + i, column=c, value=v)
print("0封面 CM-02 元数据卡已增补")

wb.save(DST)
print("Step D 完成")
