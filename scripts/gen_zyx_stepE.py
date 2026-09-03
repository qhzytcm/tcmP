# -*- coding: utf-8 -*-
"""
Step E: qhzy-中药学 完善（检查发现的两处小不足）
1. 0目录 学时汇总后补充 30.4 学时归属说明（平台章18学时含30.4）
2. 病证单元库 说明行补充药名变体溯源注释（生地/丹皮等简称）
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中药学.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=6):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# 1. 0目录：在【qhzy 增补】行前插入 30.4 学时归属
ws = wb['0目录']
# 定位【qhzy 增补】行，在其上方插入一行学时说明
target = None
for r in range(1, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    if e and '【qhzy 增补·2026-08】以下为第30章' in str(e):
        target = r
        break
if target:
    ws.insert_rows(target)
    ws.cell(row=target, column=1, value='30.4')
    ws.cell(row=target, column=5, value='30.4 岐黄智医分布式教学平台（学时含于下学期平台章18学时内, 无需另计）')
    print(f"0目录 R{target} 已插入 30.4 学时归属说明")
else:
    print("!! 未找到增补标记行")

# 2. 病证单元库 说明行补充药名变体注释
append_rows('病证单元库', [
    ['补充说明(qhzy v2.0)', '主要药物列溯源: 个别药名为本草简称变体(如"生地"=鲜/干地黄、"丹皮"=牡丹皮、"芍药"=白芍/赤芍, 均可在各论287味药中溯源); "同 BZ-xxx"为药物同前条引用。', '', '', '', '', '', ''],
], ncols=8)

wb.save(DST)
print("Step E 完成")
