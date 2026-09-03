# -*- coding: utf-8 -*-
"""检查 0cmrl目录 sheet 表头与四列结构"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X, read_only=True)
ws = wb['0cmrl目录']
rows = list(ws.iter_rows(max_row=9, values_only=True))
wb.close()
print('总行数:', ws.max_row)
for i, r in enumerate(rows[:9]):
    print(i, [str(c)[:12] if c is not None else '' for c in r])
