# -*- coding: utf-8 -*-
"""序2 题目修正: 远景 -> 愿景（文档 + 封面页 Excel）"""
from pathlib import Path
from openpyxl import load_workbook

# 1) 文档
f = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\book\08-序二-智能主审序.md')
t = f.read_text(encoding='utf-8')
n_doc = t.count('远景')
t2 = t.replace('远景', '愿景')
f.write_text(t2, encoding='utf-8')
print(f'文档修正: {n_doc} 处 远景 -> 愿景')

# 2) Excel 封面页
X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['封面页']
n_xl = 0
for r in range(18, 26):
    v = ws.cell(r, 1).value
    if v and '远景' in str(v):
        ws.cell(r, 1, str(v).replace('远景', '愿景'))
        n_xl += 1
wb.save(X)
print(f'Excel 修正: {n_xl} 处')

# 3) 校验
wb2 = load_workbook(X)
v19 = str(wb2['封面页'].cell(19, 1).value or '')
v22 = str(wb2['封面页'].cell(22, 1).value or '')
wb2.close()
print(f'R19 题目: {v19}')
print(f'R22 序文含愿景: {"愿景" in v22}  含远景: {"远景" in v22}')
doc_ok = '愿景' in (Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\book\08-序二-智能主审序.md')
                    .read_text(encoding='utf-8'))
print(f'文档校验: {"OK" if doc_ok else "FAIL"}')
