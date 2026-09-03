# -*- coding: utf-8 -*-
"""重写 ch02 R33 + ch05 R22 图注（截断修复, 从 md 提取完整文本）"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')

TARGETS = [
    ('2cmrl-bellman', 33, 2, '04-1.4.md'),
    ('5cmrl-montecarlo', 22, 5, '03-1.3.md'),
]

for sn, row, ch, fname in TARGETS:
    note = ''
    for ln in (D / f'ch{ch:02d}' / fname).read_text(encoding='utf-8').splitlines():
        if ln.strip().startswith('图注：'):
            note = ln.strip()
            break
    wb = load_workbook(X)
    ws = wb[sn]
    ws.cell(row, 3, note)
    ws.cell(row, 3).font = Font(name='黑体', size=10)
    ws.cell(row, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[row].height = (math.ceil(len(note) / 44) + 1) * 14 + 6
    wb.save(X)
    wb2 = load_workbook(X)
    v = str(wb2[sn].cell(row, 3).value or '')
    wb2.close()
    print(f'{sn} R{row}: {len(v)}/{len(note)} {"✅" if v == note else "⚠截断"}')
