# -*- coding: utf-8 -*-
"""
第16章 重建 init: 删除旧 16.素问·运气七篇, 创建空 sheet + 表头
后续 part1~part4 追加各篇全文
"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)

SHEET = '16.素问·运气七篇'
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)
ws.cell(row=1, column=1, value='类型')
ws.cell(row=1, column=2, value='四级编码')
ws.cell(row=1, column=3, value='内容（原文/注释）')
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 120

wb.save(PATH)
print(f"已重建: {SHEET}")
print(f"Sheet 顺序: {wb.sheetnames}")
