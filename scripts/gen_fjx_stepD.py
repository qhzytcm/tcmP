# -*- coding: utf-8 -*-
"""
Step D: qhzy-方剂学 收尾
1. 0目录 追加 24.4 条目（对齐正文新增节）
2. 0封面 增补 MM-02 元数据卡
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-方剂学.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=5):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# ═══ 1. 0目录 追加 24.4 ═══
ws = wb['0目录']
print("0目录 max_column:", ws.max_column)
append_rows('0目录', [
    ['', '', '', '', '【qhzy 增补·2026-08】以下为第24章新增节条目:'],
    ['24', '4', '', '', '24.4 岐黄智医分布式教学平台'],
    ['', '', '1', '', '24.4.1 2云+2硬+4台式机拓扑与数据流'],
    ['', '', '', '1', '24.4.1.1 节点分工'],
    ['', '', '', '2', '24.4.1.2 方剂数据闭环'],
    ['', '', '2', '', '24.4.2 六者SOUL与方剂学对接'],
    ['', '', '', '1', '24.4.2.1 医者(核心对接角色)'],
    ['', '', '', '2', '24.4.2.2 药者/患者/规者/法者'],
], ncols=5)
print("0目录 24.4 条目已追加")

# ═══ 2. 0封面 MM-02 元数据卡 ═══
ws = wb['0封面']
start = ws.max_row + 2
cover_rows = [
    ['', '', '【qhzy 增补·2026-08】教材元数据卡(分布式体系版)', '', ''],
    ['', '', '教材编号', 'MM-02', ''],
    ['', '', '教材名称', '方剂学(岐黄智医版)', ''],
    ['', '', '所属领域', 'D02 中药学院', ''],
    ['', '', '教学层级', '本科·基础经典(第二层)', ''],
    ['', '', '学分/学时', '4 学分 / 90 学时(单学期)', ''],
    ['', '', '前置要求', 'MM-01 临床中药学(前置链: CM-01→CM-02→MM-01→MM-02)', ''],
    ['', '', '后续衔接', 'CM-13 内科学 → 临床各科', ''],
    ['', '', '分布式承载', '台式机T1-T4方证病案采集 · 浪潮硬②A800组方模型训练 · GitHub CI/CD门禁 · 华为云公网展示', ''],
    ['', '', '版本', 'v2.0-qhzy · 2026-08 · 由种子 v1.0 适应性修改', ''],
]
for i, row in enumerate(cover_rows):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=start + i, column=c, value=v)
print("0封面 MM-02 元数据卡已增补")

wb.save(DST)
print("Step D 完成")
