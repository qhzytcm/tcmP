# -*- coding: utf-8 -*-
"""封面页 sheet: 追加序一区（内容简介上方, 书籍顺序）
插入 8 行: 标题1 + 序文5段 + 空2; 知识树图片锚点手动 +8"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['封面页']

XU1 = [
    '序一（中医药主审）',
    '机器强化智能，疑难病例易解',
    ('中医药传承数千年，辨证论治的智慧一以贯之：望闻问切以察其态，辨证立法以定其则，'
     '遣方用药以应其变，随证治之、效不更方、中病即止——这一套"观察—判断—决策—反馈"'
     '的循环，千百年来靠师承口传与临床积累代代相续。疑难病例之所以难，正在于证候复杂多变：'
     '表里同病、寒热错杂、虚实夹杂、多脏同病，医者要在纷繁的四诊信息中抓住主证，'
     '在众多的治法方剂中作出抉择。'),
    ('本书把这一古老的临床智慧，与强化学习的数学理论做了认真的、严格的对照：'
     '四诊合参对应状态观测，治法方剂对应动作空间，疗效反馈对应奖励信号，'
     '病邪传变对应状态转移，辨证论治对应策略学习。这不是修辞层面的比附，'
     '而是可以落地计算的结构同构——当疑难病例的诊疗过程被写成"状态—动作—奖励"的序列，'
     '价值迭代便可在万千证候组合中反复评估改进，策略梯度便可在探索与利用之间寻得平衡，'
     '行动者-评价者架构便让"医者决策"与"规者质控"各司其职。'),
    ('机器强化智能，正是这样的图景：机器在足够算力与真实数据之上，反复强化对中医药场景的认知，'
     '把名老中医的经验转化为可验算、可传承、可进化的决策模型；疑难病例易解，则是这份智能的落点'
     '——辨证不再只是个人的冥思，而是人机协同下对病机层次的层层剖析；'
     '选方不再只是经验的直觉，而是对全疗程回报的审慎权衡。'),
    ('作为本书的中医药审校，我在审读全稿的过程中坚持了几条红线：其一，中医概念不可被算法概念偷换'
     '——肝不是控制器，经络不是数据传输线，一切映射都须以中医本体为准；其二，方剂之用于教学示例，'
     '功效主治、证型归属必须符合统编教材口径，归脾汤治心脾两虚、银翘散治风热表证、'
     '麻黄汤治风寒表实，一字不可含糊；其三，疗效数值不可夸大——有效率的表达须有证据等级支撑。'
     '审稿中发现的若干辨证表述与数值口径问题，均已逐一修正统一。'),
    ('愿读者以本书为阶，既通强化学习之器，亦守辨证论治之道，'
     '在"机器强化智能"的时代潮流中，让疑难病例多一分"易解"的可能。'),
    '中医药主审  2026 年 8 月',
]

# 1) 插入 8 行（R10 前）
ws.insert_rows(10, 8)

# 2) 写序一区
r = 10
# 标题行
ws.cell(r, 1, XU1[0])
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=13, bold=True)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 24
r += 1
# 题目行
ws.cell(r, 1, XU1[1])
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=12)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 22
r += 1
# 序文段落（每段合并 A-C, 左顶格 wrap）
for para in XU1[2:6]:
    ws.cell(r, 1, para)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1)
    c.font = Font(name='黑体', size=10)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[r].height = (math.ceil(len(para) / 95) + 1) * 14 + 4
    r += 1
# 落款
ws.cell(r, 1, XU1[6])
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=10)
c.alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 20
r += 1
# 空行 2（r, r+1 自动留空）

# 3) 知识树图片锚点 +8
for img in ws._images:
    img.anchor._from.row += 8

wb.save(X)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['封面页']
title = str(ws2.cell(10, 1).value or '')
topic = str(ws2.cell(11, 1).value or '')
anchors = [img.anchor._from.row + 1 for img in ws2._images]
intro = str(ws2.cell(19, 1).value or '')
cap = str(ws2.cell(21, 1).value or '')
wb2.close()
print(f'序一标题 R10: {title}')
print(f'序一题目 R11: {topic}')
print(f'内容简介 R19: {intro}')
print(f'知识树图题 R21: {cap[:16]}')
print(f'封面页图片锚点: {anchors}')
