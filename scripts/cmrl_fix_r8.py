# -*- coding: utf-8 -*-
"""修复 ch01 R8 图注截断（32/325）——多重尝试"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')

# md 图注（ch01 第 1 幅）
note = ''
for ln in (D / 'ch01' / '01-1.1.md').read_text(encoding='utf-8').splitlines():
    if ln.strip().startswith('图注：'):
        note = ln.strip()
        break
print(f'md 图注: {len(note)} 字')

# 尝试 1: 清空重写
for attempt in range(3):
    wb = load_workbook(X)
    ws = wb['1cmrl-conceptions']
    ws['C8'] = None
    wb.save(X)
    wb = load_workbook(X)
    ws = wb['1cmrl-conceptions']
    ws['C8'] = note
    ws['C8'].font = Font(name='黑体', size=10)
    ws['C8'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[8].height = (math.ceil(len(note) / 44) + 1) * 14 + 6
    wb.save(X)
    wb = load_workbook(X)
    v = str(wb['1cmrl-conceptions']['C8'].value or '')
    wb.close()
    print(f'尝试{attempt+1}: {len(v)}/{len(note)} {"✅" if v == note else "⚠截断"}')
    if v == note:
        break
