# -*- coding: utf-8 -*-
"""
Step K: 0目录 终极重建 —— 新建 Sheet 替换旧 Sheet（纯内存，绝无行操作残留）
结构: 表头+1-11章(R1-R941) | 12体质块 | 13平台块(旧块改名+13.5-13.7) | 说明文字 | 附篇块
"""
import openpyxl, re

SEED = r'C:\Users\DELL\Desktop\textbooks\中医基础理论final.xlsx'
DST = r'C:\Users\DELL\Desktop\qhzy-中医基础理论.xlsx'

# 1. 读种子 0目录（read_only → 内存 list）
swb = openpyxl.load_workbook(SEED, read_only=True)
sws = swb['0目录']
seed_rows = [['' if c.value is None else str(c.value).strip() for c in row] for row in sws.iter_rows(max_col=5)]
swb.close()
print("种子行数:", len(seed_rows))

# 2. 定位
ch12_idx = None   # 旧"第12章 教学平台"行索引
note_idx = None   # "本版目录的显著特征"行索引
for i, r in enumerate(seed_rows):
    if r[4].startswith('第12章') and '教学平台' in r[4]:
        ch12_idx = i
    if r[4].startswith('本版目录的显著特征'):
        note_idx = i
        break
print(f"ch12_idx={ch12_idx}(R{ch12_idx+1}) note_idx={note_idx}(R{note_idx+1})")

head = seed_rows[:ch12_idx]                       # 表头+1-11章（含空行）
old12 = [r for r in seed_rows[ch12_idx:note_idx] if any(r)]   # 旧12章块（去空行）
notes = [r for r in seed_rows[note_idx:] if any(r)]           # 说明文字（去空行）
print(f"head={len(head)} old12={len(old12)} notes={len(notes)}")

# 3. old12 → 13章（标题改名 + 12.x→13.x）
new13 = []
for r in old12:
    nr = r[:]
    e = nr[4]
    if '第12章' in e and '教学平台' in e:
        nr[4] = '第13章\u3000岐黄智医分布式教学平台与评价体系'
    elif re.match(r'^12\.', e):
        nr[4] = '13.' + e[3:]
    new13.append(nr)

# 4. qhzy 增补块
tizhi = [
    ['12','00','00','00','第12章 体质学说与体质辨识'],
    ['','01','00','00','12.1 体质的概念与形成'],
    ['','','01','00','12.1.1 体质的基本概念'],
    ['','','','01','12.1.1.1 体质的定义与内涵'],
    ['','','','02','12.1.1.2 体质与证候、病机的关系'],
    ['','','02','00','12.1.2 体质的形成因素'],
    ['','','','01','12.1.2.1 先天因素(父母禀赋与遗传)'],
    ['','','','02','12.1.2.2 后天因素(饮食、起居、情志与地理)'],
    ['','02','00','00','12.2 体质的分类'],
    ['','','01','00','12.2.1 以《灵枢·阴阳二十五人》为纲'],
    ['','','','01','12.2.1.1 五行五形人'],
    ['','','','02','12.2.1.2 五音二十五人'],
    ['','','02','00','12.2.2 王琦九分法(现代参考)'],
    ['','','','01','12.2.2.1 九种基本体质'],
    ['','','','02','12.2.2.2 经典与九分法的映射'],
    ['','03','00','00','12.3 体质与发病、辨证论治'],
    ['','','01','00','12.3.1 体质与发病倾向'],
    ['','','02','00','12.3.2 体质与辨证论治(因人制宜)'],
    ['','04','00','00','12.4 体质辨识的AI实现(分布式平台落点)'],
    ['','','01','00','12.4.1 体质辨识量表数字化'],
    ['','','02','00','12.4.2 体质-病证关联推理'],
]
new13 += [
    ['','05','00','00','13.5 岐黄智医分布式网络编程体系'],
    ['','','01','00','13.5.1 三服务器+GitHub+四台式机拓扑'],
    ['','','02','00','13.5.2 分布式任务调度'],
    ['','06','00','00','13.6 六者SOUL与医圣成长引擎对接'],
    ['','','01','00','13.6.1 六者角色与教材内容的映射'],
    ['','','02','00','13.6.2 医圣成长引擎'],
    ['','07','00','00','13.7 教材-图谱-平台三角色闭环'],
    ['','','01','00','13.7.1 本章知识树→图谱注入'],
    ['','','02','00','13.7.2 使用数据→教材修订'],
]
fupian = [
    ['附','00','00','00','附篇 五运六气'],
    ['','01','00','00','附1 五运六气的基本概念'],
    ['','02','00','00','附2 六气'],
    ['','03','00','00','附3 运气与发病'],
    ['','04','00','00','附4 运气学说的学习与数字化'],
]
qhzy_note = [['','','','','【qhzy 增补·2026-08】以下为第12章(体质)、第13章(岐黄智医分布式教学平台)、附篇(五运六气)目录条目:']]

# 说明文字引用修正
for r in notes:
    e = r[4]
    if '知识树落地' in e and '第12章' in e:
        r[4] = e.replace('第12章独立成章', '第13章独立成章').replace('第12章', '第13章')
    if '本目录与规划教材的对应' in e:
        r[4] = e.replace("第12章'教学平台与评价体系'", "第13章'岐黄智医分布式教学平台与评价体系'").replace('第12章', '第13章')

# 5. 组装（含空行分隔）
new_toc = []
new_toc += head
new_toc.append([''] * 5)
new_toc += tizhi
new_toc.append([''] * 5)
new_toc += new13
new_toc.append([''] * 5)
new_toc += notes
new_toc.append([''] * 5)
new_toc += qhzy_note
new_toc += fupian
print("组装总行数:", len(new_toc))

# 6. 新建 Sheet 替换（删除旧 0目录，新建同名）
wb = openpyxl.load_workbook(DST)
idx = wb.sheetnames.index('0目录')
del wb['0目录']
ws = wb.create_sheet('0目录', index=idx)
for i, row in enumerate(new_toc):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=i + 1, column=c, value=v)
wb.save(DST)
print("Step K 完成")

# 7. 独立复核
wb2 = openpyxl.load_workbook(DST, read_only=True)
ws2 = wb2['0目录']
print("新 max_row:", ws2.max_row)
print("\n=== 章级顺序 ===")
for i, row in enumerate(ws2.iter_rows(min_row=1, max_col=5), 1):
    cells = ['' if x.value is None else str(x.value).strip() for x in row]
    e = cells[4]
    if e and len(e) < 30 and e.startswith('第') and '章' in e and not e.startswith('第13章 岐黄'):
        print(f"R{i}: {e}")
print("\n=== 13.x 前12行 ===")
cnt = 0
for i, row in enumerate(ws2.iter_rows(min_row=950, max_col=5), 950):
    e = row[4].value
    if e and isinstance(e, str) and e.strip().startswith('13.'):
        print(f"R{i}: {e.strip()[:45]}")
        cnt += 1
    if cnt >= 12:
        break
wb2.close()
