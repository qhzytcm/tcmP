# -*- coding: utf-8 -*-
"""决定性定位: C19 写不同文本 vs C20 写焠㶼文本"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
T1 = '测试' * 40  # 80字普通文本
T2 = '病在肾，愈在春；春不愈，甚于长夏；长夏不死，持于秋；起于冬。禁犯焠㶼热食温炙衣。肾病者，夜半慧，四季甚，下晡静。肾欲坚，急食苦以坚之，用苦补之，咸写之。'

wb = openpyxl.load_workbook(PATH)
ws = wb['6.素问·经脉血气']
ws.cell(row=19, column=3, value=T1)  # C19 普通文本
ws.cell(row=20, column=3, value=T2)  # C20 焠㶼文本
wb.save(PATH)
wb.close()

wb2 = openpyxl.load_workbook(PATH)
ws2 = wb2['6.素问·经脉血气']
v19 = ws2.cell(row=19, column=3).value
v20 = ws2.cell(row=20, column=3).value
print(f"C19 普通文本: len={len(v19)} {'✔' if len(v19) == len(T1) else '✘截断'}")
print(f"C20 焠㶼文本: len={len(v20)} {'✔' if len(v20) == len(T2) else '✘截断'}")
wb2.close()
