# -*- coding: utf-8 -*-
"""runner: tempfile.mkstemp OS-safe 动态生成 hermes-verify- 脚本 → 运行 → 自动清理（ad-hoc）"""
import os, subprocess, sys, tempfile

BODY = r'''# -*- coding: utf-8 -*-
"""hermes-verify-hjnj-full.py — qhzy-黄帝内经 全书31章 完成性聚焦验证（ad-hoc, tempfile OS-safe）"""
import sys, re, os, json
import openpyxl

print(f"[hermes-verify] 脚本: {os.path.abspath(__file__)}")

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
fails, passes = [], []

def check(name, cond, detail=""):
    (passes if cond else fails).append(f"{name}{(' | ' + detail) if detail and not cond else ''}")

wb = openpyxl.load_workbook(PATH, read_only=True)

# 1. 结构: 46 Sheet = 15 引擎 + 31 正文章
names = wb.sheetnames
ch = [s for s in names if re.match(r'^\d+\.', s)]
check("总Sheet=46", len(names) == 46, str(len(names)))
check("正文章31", len(ch) == 31, str(len(ch)))
nums = [int(s.split('.')[0]) for s in ch]
check("1-31连续", nums == list(range(1, 32)), str(nums))
EXPECT_HEAD = ['0改写说明','0分布式体系','平台建设与依赖','平台支撑完整性','0历史',
               '0映射1','0映射2','0映射3','0封面','0目录',
               'ICD11统一命名','病证单元库','多智能体协同','兼容性思维流','可复用Skills']
check("引擎15连续", names[:15] == EXPECT_HEAD, str(names[:15]))

# 2. 篇目: 素问81 + 灵枢81 = 162
suwen = ling = 0
for s in ch:
    ws = wb[s]
    n = sum(1 for row in ws.iter_rows(min_row=2, max_col=3) if row[0].value == '篇名')
    if int(s.split('.')[0]) <= 17:
        suwen += n
    else:
        ling += n
check("素问81篇", suwen == 81, f"实际{suwen}")
check("灵枢81篇", ling == 81, f"实际{ling}")

# 3. 全书总量
tot = {'原文': 0, '通读': 0, '注释': 0}
for s in ch:
    ws = wb[s]
    for row in ws.iter_rows(min_row=2, max_col=3):
        c = row[0].value
        if c in tot and row[2].value:
            tot[c] += 1
check(f"原文{tot['原文']}(≥1700)", tot['原文'] >= 1700, f"实际{tot['原文']}")
check(f"通读{tot['通读']}(=142)", tot['通读'] == 142, f"实际{tot['通读']}")
check(f"注释{tot['注释']}(≥160)", tot['注释'] >= 160, f"实际{tot['注释']}")

# 4. 无截断
suspects = []
for s in ch:
    ws = wb[s]
    for row in ws.iter_rows(min_row=2, max_col=3):
        v = row[2].value
        if v and isinstance(v, str) and len(v) > 10:
            if not re.search(r'[。；：！？」』）】]$', v) and re.search(r'[以之而于与为者]$', v.strip()):
                suspects.append((s, row[0].row))
check("无截断", not suspects, f"{len(suspects)}处")

# 5. 每章悬疑/非零和
bad_ch = []
for s in ch:
    ws = wb[s]
    zt = "\n".join(str(row[2].value) for row in ws.iter_rows(min_row=2, max_col=3) if row[0].value == '注释' and row[2].value)
    if not (zt and '悬疑' in zt and '非零和' in zt):
        bad_ch.append(s)
check("31章悬疑/非零和齐", not bad_ch, str(bad_ch))

# 6. 关键名句抽查（首/中/末篇）
yw1 = "\n".join(str(row[2].value) for row in wb['1.素问·养生总纲'].iter_rows(min_row=2, max_col=3) if row[0].value == '原文')
yw16 = "\n".join(str(row[2].value) for row in wb['16.素问·运气七篇'].iter_rows(min_row=2, max_col=3) if row[0].value == '原文')
yw31 = "\n".join(str(row[2].value) for row in wb['31.灵枢·九针痈疽'].iter_rows(min_row=2, max_col=3) if row[0].value == '原文')
for name, yw, ks in [
    ('ch1', yw1, ['法于阴阳，和于术数', '上古之人，其知道者']),
    ('ch16', yw16, ['甲己之岁，土运统之', '诸风掉眩，皆属于肝', '亢则害，承乃制']),
    ('ch31', yw31, ['九针者，天地之大数也', '发于足指，名曰脱痈', '急斩之，不则死矣']),
]:
    for k in ks:
        check(f"{name}[{k[:10]}]", k in yw)

# 7. 0系列未破坏
t1 = "\n".join(str(c.value) for row in wb['0映射1'].iter_rows() for c in row if c.value)
t2 = "\n".join(str(c.value) for row in wb['0映射2'].iter_rows() for c in row if c.value)
t3 = "\n".join(str(c.value) for row in wb['0映射3'].iter_rows() for c in row if c.value)
tt = "\n".join(str(c.value) for row in wb['0目录'].iter_rows() for c in row if c.value)
tc = "\n".join(str(c.value) for row in wb['0封面'].iter_rows() for c in row if c.value)
check("0映射1邹纯朴", '邹纯朴' in t1)
check("0映射2翟双庆", '翟双庆' in t2)
check("0映射3李梢", '李梢' in t3)
check("0目录素问灵枢", '素问81篇' in tt and '灵枢81篇' in tt)
check("0封面三篇文字", '通透内经内涵' in tc and '人机闭环迭代' in tc and '理会原则精神' in tc)

# 8. 进度
d = json.load(open(r'C:\Users\DELL\tcmP\scripts\progress\hjnj_writing_progress.json', encoding='utf-8'))
check("进度31/31", d['done'] == 31, str(d['done']))

wb.close()
print(f"\nPASS {len(passes)} | FAIL {len(fails)}")
for p in passes:
    print("  ✔", p)
for f in fails:
    print("  ✘", f)
sys.exit(1 if fails else 0)
'''

fd, tmp_path = tempfile.mkstemp(prefix='hermes-verify-hjnj-full-', suffix='.py', dir=tempfile.gettempdir())
with os.fdopen(fd, 'w', encoding='utf-8') as f:
    f.write(BODY)
print(f"[hermes-verify] 临时脚本: {tmp_path}")

PY = r'C:\Users\DELL\AppData\Local\hermes\hermes-agent\venv\Scripts\python.exe'
try:
    r = subprocess.run([PY, tmp_path], capture_output=True, text=True, encoding='utf-8', timeout=180)
    print(r.stdout)
    if r.stderr:
        print("STDERR:", r.stderr[-1200:])
    sys.exit(r.returncode)
finally:
    try:
        os.remove(tmp_path)
        print(f"[hermes-verify] 已清理: {tmp_path} 存在={os.path.exists(tmp_path)}")
    except OSError:
        pass
