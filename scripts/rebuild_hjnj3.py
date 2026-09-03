# -*- coding: utf-8 -*-
"""
决定性重建 v3: 全簿读取 → 写入全新文件 → 替换
（彻底解决 sharedStrings 截断: 6.2.1.7 等长文本写入被截断的问题）
"""
import os, re, shutil
import openpyxl

SRC = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
DST = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版)_rebuilt3.xlsx'

# 1. 读取原文件所有 sheet 数据（当前内容为准）
wb_old = openpyxl.load_workbook(SRC, read_only=True)
all_data = {}
for s in wb_old.sheetnames:
    ws = wb_old[s]
    rows = []
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if any(v is not None for v in vals):
            rows.append(vals)
    all_data[s] = rows
wb_old.close()
print(f"读取 {len(all_data)} 个 sheet")

# 2. 修复 6.2.1.7（脚本源完整文本）
FULL62 = '病在肾，愈在春；春不愈，甚于长夏；长夏不死，持于秋；起于冬。禁犯焠㶼热食温炙衣。肾病者，夜半慧，四季甚，下晡静。肾欲坚，急食苦以坚之，用苦补之，咸写之。'
fixed = False
for i, row in enumerate(all_data['6.素问·经脉血气']):
    if len(row) >= 2 and row[1] == '6.2.1.7':
        print(f"修复前: {str(row[2])[:40]}... (len={len(str(row[2]))})")
        row[2] = FULL62
        fixed = True
        print(f"修复后: {row[2][:40]}... (len={len(row[2])})")
print(f"6.2.1.7 修复: {fixed}")

# 3. 创建新工作簿按原顺序写入
wb_new = openpyxl.Workbook()
wb_new.remove(wb_new.active)
for s in all_data:
    ws = wb_new.create_sheet(s)
    for r, row in enumerate(all_data[s], 1):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=r, column=c, value=v)

wb_new.save(DST)
print(f"重建完成: {DST}")

# 4. 验证修复点
wb_check = openpyxl.load_workbook(DST, read_only=True)
ws = wb_check['6.素问·经脉血气']
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[1].value == '6.2.1.7':
        v = str(row[2].value)
        print(f"验证: len={len(v)} {'✔完整' if '肾欲坚' in v else '✘仍截断'}")
wb_check.close()

# 5. 替换
shutil.move(DST, SRC)
print(f"已替换: {SRC}")
