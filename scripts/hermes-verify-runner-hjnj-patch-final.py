# -*- coding: utf-8 -*-
"""runner: tempfile.mkstemp OS-safe 动态生成 hermes-verify- 脚本 → 运行 → 自动清理（ad-hoc）"""
import os, subprocess, sys, tempfile

BODY = r'''# -*- coding: utf-8 -*-
"""hermes-verify-hjnj-patch-final.py — qhzy-黄帝内经 平台适配补丁 聚焦验证（ad-hoc, tempfile OS-safe）"""
import sys, os, sqlite3
import openpyxl

print(f"[hermes-verify] 脚本: {os.path.abspath(__file__)}")

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
fails, passes = [], []

def check(name, cond, detail=""):
    (passes if cond else fails).append(f"{name}{(' | ' + detail) if detail and not cond else ''}")

wb = openpyxl.load_workbook(PATH, read_only=True)

# 1. ICD11 平台db校准列
ws = wb['ICD11统一命名']
check("F列表头", ws.cell(row=2, column=6).value == '平台db校准(entities表)', str(ws.cell(row=2, column=6).value))
cal = {}
for row in ws.iter_rows(min_row=3, max_col=6):
    if row[2].value:
        cal[str(row[2].value).strip()] = str(row[5].value) if row[5].value else ''
check("校准列17行", len(cal) == 17, f"实际{len(cal)}")
check("8B11 ischaemic", 'ischaemic' in cal.get('8B11', ''), cal.get('8B11', ''))
check("MD81校准", 'Abdominal or pelvic pain' in cal.get('MD81', ''), cal.get('MD81', ''))
check("DD91.2校准", 'Functional diarrhoea' in cal.get('DD91.2', ''), cal.get('DD91.2', ''))

# 2. 病证单元三列
ws = wb['病证单元库']
hdr = [ws.cell(row=2, column=c).value for c in (8, 9, 10)]
check("三列表头", hdr == ['分子靶标(molecular_targets)', 'ICD11桥接码', '六经归属'], str(hdr))
n_ext = sum(1 for row in ws.iter_rows(min_row=3, max_col=10) if row[0].value and str(row[0].value).startswith('BZ-HJ') and row[9].value)
check("三列17行", n_ext == 17, f"实际{n_ext}")

# 3. db 命中
conn = sqlite3.connect(r'C:\Users\DELL\tcmP\data\icd11_mms.db')
cur = conn.cursor()
for code in ['MB48.0', '8B11', 'MD81', 'DD91.2', 'ME84.2', 'ME06', 'MD12', '5A11', 'MG30']:
    cur.execute("SELECT title FROM entities WHERE code=?", (code,))
    row = cur.fetchone()
    check(f"db[{code}]", row is not None)
conn.close()

# 4. build_document 17/17
sys.path.insert(0, r'C:\Users\DELL\tcmP\scripts')
from tcm_embed import build_document
ws = wb['病证单元库']
ok = total = 0
for row in ws.iter_rows(min_row=3, max_col=7):
    vals = ['' if c.value is None else str(c.value).strip() for c in row]
    if not vals[0] or not vals[0].startswith('BZ-HJ'):
        continue
    total += 1
    code, name, yuanwen, bingji, zangfu, linchuang, luodian = vals
    dsu = {
        "disease_side": {"disease_name": name, "icd11_code": "", "key_symptoms": [bingji]},
        "syndrome_side": {"syndrome_name": f"{name}（内经）", "pattern_type": "内经病证",
                          "zangfu": zangfu, "six_channels": "", "key_symptoms": []},
        "clinical": {"recommended_formula": "", "notes": f"溯源: {yuanwen}; 病机: {bingji}; 临床: {linchuang}"},
    }
    try:
        doc = build_document(dsu)
        if doc and len(str(doc)) > 20:
            ok += 1
    except Exception:
        pass
check("build_document 17/17", ok == total == 17, f"{ok}/{total}")

# 5. 正文未破坏
tot = {'原文': 0, '通读': 0, '注释': 0}
for s in ['1.素问·养生总纲','2.素问·阴阳学说','3.素问·藏象','4.素问·诊法总论','5.素问·脉学']:
    ws = wb[s]
    for row in ws.iter_rows(min_row=2, max_col=3):
        if row[0].value in tot:
            tot[row[0].value] += 1
check("正文原文160", tot['原文'] == 160, str(tot['原文']))
check("正文通读20", tot['通读'] == 20, str(tot['通读']))
check("正文注释31", tot['注释'] == 31, str(tot['注释']))

# 6. 序言与引擎未破坏
tc = "\n".join(str(c.value) for row in wb['0封面'].iter_rows() for c in row if c.value)
check("0封面序一", '序一 · 通透内经内涵' in tc and '翟双庆' in tc)
check("0封面序二", '序二 · 人机闭环迭代' in tc and '李梢' in tc)
check("0封面简介", '内容简介 · 理会原则精神' in tc and '邹纯朴' in tc)
te = "\n".join(str(c.value) for row in wb['0分布式体系'].iter_rows() for c in row if c.value)
check("0分布式A800", 'A800' in te and 'node-coordinator' in te)
tk = "\n".join(str(c.value) for row in wb['可复用Skills'].iter_rows() for c in row if c.value)
check("Skills6规格", 'tcm-huangdi-neijing-seed' in tk and 'tcm-distributed-node-orchestration' in tk)

wb.close()
print(f"\nPASS {len(passes)} | FAIL {len(fails)}")
for p in passes:
    print("  ✔", p)
for f in fails:
    print("  ✘", f)
sys.exit(1 if fails else 0)
'''

fd, tmp_path = tempfile.mkstemp(prefix='hermes-verify-hjnj-patch-final-', suffix='.py', dir=tempfile.gettempdir())
with os.fdopen(fd, 'w', encoding='utf-8') as f:
    f.write(BODY)
print(f"[hermes-verify] 临时脚本: {tmp_path}")

PY = r'C:\Users\DELL\AppData\Local\hermes\hermes-agent\venv\Scripts\python.exe'
try:
    r = subprocess.run([PY, tmp_path], capture_output=True, text=True, encoding='utf-8', timeout=180)
    print(r.stdout)
    if r.stderr:
        print("STDERR:", r.stderr[-1200:])
    sys.exit(r.returncode)
finally:
    try:
        os.remove(tmp_path)
        print(f"[hermes-verify] 已清理: {tmp_path} 存在={os.path.exists(tmp_path)}")
    except OSError:
        pass
