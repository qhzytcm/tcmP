# -*- coding: utf-8 -*-
"""
决定性重建 v4（最终方案）: 正文章全部用脚本源重建（绕开 sharedStrings 损坏）
- ch01-15,17-31: gen_hjnj_chNN.py 源
- ch16: part1-4 追加脚本（init + part1~4）
- 引擎/0系列: 保留当前文件数据（含序言/补注/ICD11校准）
"""
import os, re, shutil
import openpyxl

SRC = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
DST = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版)_rebuilt4.xlsx'

# 1. 读取引擎/0系列 sheet（保留当前）
wb_old = openpyxl.load_workbook(SRC, read_only=True)
engine_data = {}
for s in wb_old.sheetnames:
    if not re.match(r'^\d+\.', s):
        ws = wb_old[s]
        rows = []
        for row in ws.iter_rows():
            vals = [c.value for c in row]
            if any(v is not None for v in vals):
                rows.append(vals)
        engine_data[s] = rows
wb_old.close()
print(f"引擎/0系列: {len(engine_data)} 个 sheet")

# 2. 从脚本源提取正文章 rows
def extract_from_script(src_path, name):
    with open(src_path, encoding='utf-8') as f:
        src = f.read()
    src = src.replace('wb.save(DST)', 'pass  # 拦截')
    src = src.replace('wb.save(PATH)', 'pass  # 拦截')
    ns = {}
    exec(compile(src, f'<{name}>', 'exec'), ns)
    return ns.get('SHEET'), ns.get('rows', [])

ch_data = {}  # sheet_name -> rows
# ch01-15, 17-31
for i in list(range(1, 16)) + list(range(17, 32)):
    src_path = rf'C:\Users\DELL\tcmP\scripts\gen_hjnj_ch{i:02d}.py'
    if not os.path.exists(src_path):
        print(f"  ✘ 缺脚本: {src_path}")
        continue
    try:
        sheet_name, rows = extract_from_script(src_path, f'ch{i:02d}')
        if sheet_name:
            ch_data[sheet_name] = [[cat, code, text] for cat, code, text in rows]
            print(f"  ✔ ch{i:02d} {sheet_name} ({len(rows)} 行)")
    except Exception as e:
        print(f"  ✘ ch{i:02d}: {e}")

# ch16: part1-4 追加
for part in ['init', 'part1', 'part2', 'part3', 'part4']:
    src_path = rf'C:\Users\DELL\tcmP\scripts\gen_hjnj_ch16_{part}.py'
    if not os.path.exists(src_path):
        continue
    if part == 'init':
        continue  # init 只删建
    try:
        sheet_name, rows = extract_from_script(src_path, f'ch16_{part}')
        if part == 'part1':
            ch_data['16.素问·运气七篇'] = [[cat, code, text] for cat, code, text in rows]
            print(f"  ✔ ch16 part1 ({len(rows)} 行)")
        else:
            ch_data['16.素问·运气七篇'].extend([[cat, code, text] for cat, code, text in rows])
            print(f"  ✔ ch16 {part} (+{len(rows)} 行)")
    except Exception as e:
        print(f"  ✘ ch16 {part}: {e}")
print(f"正文章: {len(ch_data)} 个 sheet, ch16 共 {len(ch_data.get('16.素问·运气七篇', []))} 行")

# 3. 创建新工作簿
wb_new = openpyxl.Workbook()
wb_new.remove(wb_new.active)

ORDER = list(engine_data.keys()) + sorted(ch_data.keys(), key=lambda x: int(x.split('.')[0]))
for s in ORDER:
    if s in ch_data:
        ws = wb_new.create_sheet(s)
        ws.cell(row=1, column=1, value='类型')
        ws.cell(row=1, column=2, value='四级编码')
        ws.cell(row=1, column=3, value='内容（原文/注释）')
        for r, row in enumerate(ch_data[s], 2):
            for c, v in enumerate(row, start=1):
                if v:
                    ws.cell(row=r, column=c, value=v)
    else:
        ws = wb_new.create_sheet(s)
        for r, row in enumerate(engine_data[s], 1):
            for c, v in enumerate(row, start=1):
                if v:
                    ws.cell(row=r, column=c, value=v)

wb_new.save(DST)
print(f"\n重建完成: {DST}")

# 4. 验证修复点（含 6.2.1.7）
wb_check = openpyxl.load_workbook(DST, read_only=True)
checks = [
    ('6.素问·经脉血气', '6.2.1.7', '肾欲坚'),
    ('16.素问·运气七篇', '16.1.1.7', '太虚寥廓，肇基化元'),
    ('20.灵枢·经脉经别', '20.1.1.13', '不盛不虚，以经取之'),
    ('22.灵枢·营卫气血', '22.1.1.11', '上焦如雾'),
]
for s, code, kw in checks:
    ws = wb_check[s]
    found = False
    for row in ws.iter_rows(min_row=2, max_col=3):
        if row[1].value == code and row[2].value:
            v = str(row[2].value)
            found = kw in v
            print(f"{s} {code}: len={len(v)} {'✔' if found else '✘'}")
            break
    if not found:
        print(f"{s} {code}: ✘ 未找到")
wb_check.close()

# 5. 替换
shutil.move(DST, SRC)
print(f"已替换: {SRC}")
