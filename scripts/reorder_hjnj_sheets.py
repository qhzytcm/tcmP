# -*- coding: utf-8 -*-
"""重排 Sheet 顺序: 对齐前5本 qhzy 模式（0系列 → 引擎 → 正文章）"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)

ORDER = [
    '0改写说明', '0分布式体系', '平台建设与依赖', '平台支撑完整性', '0历史',
    '0映射1', '0映射2', '0映射3', '0封面', '0目录',
    'ICD11统一命名', '病证单元库', '多智能体协同', '兼容性思维流', '可复用Skills',
    '1.素问·养生总纲', '2.素问·阴阳学说', '3.素问·藏象', '4.素问·诊法总论', '5.素问·脉学',
]

# 逐个移动到目标位置
for i, name in enumerate(ORDER):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=-(wb.sheetnames.index(name) - i))

wb.save(PATH)
print(f"重排后 Sheet 顺序:")
for i, s in enumerate(wb.sheetnames, 1):
    print(f"  {i:2d}. {s}")
