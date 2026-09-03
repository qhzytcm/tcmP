# -*- coding: utf-8 -*-
"""修复 Excel 图注行截断：从 md 提取完整图注重写 + 回读校验（单行写入模式）"""
import re
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')

# 需要修复的 (sheet, 行号, 章, md 文件)
TARGETS = [
    ('4cmrl-iteration', 18, 4, '02-1.2.md'),
    ('5cmrl-montecarlo', 51, 5, '06-1.6.md'),
    ('8cmrl-valuefunction', 16, 8, '02-1.2.md'),
]

# 从 md 提取完整图注行
notes = {}
for sn, row, ch, fname in TARGETS:
    f = D / f'ch{ch:02d}' / fname
    for ln in f.read_text(encoding='utf-8').splitlines():
        if ln.strip().startswith('图注：'):
            notes[(sn, row)] = ln.strip()
            break

wb = load_workbook(X)
for sn, row, ch, fname in TARGETS:
    full = notes.get((sn, row), '')
    if not full:
        print(f'[MISS] {sn} R{row} 未找到 md 图注')
        continue
    ws = wb[sn]
    ws.cell(row, 3, full)   # 单行写入
    # 行高按长度重设
    n_rows = max(1, (len(full) // 44) + 1)
    ws.row_dimensions[row].height = n_rows * 14 + 6
    print(f'[写] {sn} R{row}: {len(full)} 字')
wb.save(X)

# 回读校验: 长度一致即未截断
wb = load_workbook(X)
ok = True
for sn, row, ch, fname in TARGETS:
    v = str(wb[sn].cell(row, 3).value or '')
    full = notes.get((sn, row), '')
    match = v == full
    ok &= match
    print(f'[校验] {sn} R{row}: Excel={len(v)} md={len(full)} '
          f'{"✅一致" if match else "⚠截断"}')
wb.close()
print(f'修复结果: {"全部一致" if ok else "仍有截断"}')
