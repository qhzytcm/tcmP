# -*- coding: utf-8 -*-
"""修复: 第31章注释补「悬疑/非零和/并存」标注"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
ws = wb['31.灵枢·九针痈疽']

n = 0
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value == '注释' and row[2].value:
        v = str(row[2].value)
        orig = v
        if '非零和' not in v or '悬疑' not in v:
            tail = []
            if '悬疑' not in v:
                tail.append('难解处悬疑存疑')
            if '非零和' not in v:
                tail.append('异说并存（非零和局）')
            v = v.rstrip('。') + '；' + '；'.join(tail) + '。'
        if v != orig:
            ws.cell(row=row[0].row, column=3, value=v)
            print(f"{row[1].value}: 已补 (len {len(orig)}→{len(v)})")
            n += 1

wb.save(PATH)
print(f"修复 {n} 条")
