# -*- coding: utf-8 -*-
"""终极测试: 常规模式修复 6.2.1.7 → 保存 → 读回"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
FULL = '病在肾，愈在春；春不愈，甚于长夏；长夏不死，持于秋；起于冬。禁犯焠㶼热食温炙衣。肾病者，夜半慧，四季甚，下晡静。肾欲坚，急食苦以坚之，用苦补之，咸写之。'

wb = openpyxl.load_workbook(PATH)
ws = wb['6.素问·经脉血气']

# 找 6.2.1.7 并打印周围信息
target_row = None
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[1].value == '6.2.1.7':
        target_row = row[0].row
        print(f"找到 R{target_row}: len={len(str(row[2].value))}")
        # 检查编码单元格类型
        print(f"  code type: {type(row[1].value)}, value repr: {repr(row[1].value)}")
        break

if target_row:
    ws.cell(row=target_row, column=3, value=FULL)
    print(f"已写入 R{target_row} len={len(FULL)}")
    wb.save(PATH)
    print("已保存")
wb.close()

# 读回验证
wb2 = openpyxl.load_workbook(PATH)
v = wb2['6.素问·经脉血气'].cell(row=target_row, column=3).value
print(f"读回: len={len(v)} {'✔完整' if len(v) == len(FULL) else '✘截断'}")
print(f"内容: {v[:60]}")
wb2.close()
