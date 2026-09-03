# -*- coding: utf-8 -*-
"""修正 fig10-1 锚点 R5 -> R6（图体行=图题 R5 下方）"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['10cmrl-acloop']
for img in ws._images:
    if img.anchor._from.row + 1 == 5:
        img.anchor._from.row += 1
        print('fig10-1 锚点 R5 -> R6')
wb.save(X)

wb2 = load_workbook(X)
ws2 = wb2['10cmrl-acloop']
anchors = sorted(img.anchor._from.row + 1 for img in ws2._images)
wb2.close()
print(f'锚点: {anchors}')
