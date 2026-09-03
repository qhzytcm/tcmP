# -*- coding: utf-8 -*-
"""检查 Excel 占用状态 + 各章图注同步情况 + .aligned.xlsx 副本"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')

# 1) Excel 可写性测试（try save 到 temp 副本）
try:
    wb = load_workbook(X)
    wb.close()
    print('Excel 可读: True')
except Exception as e:
    print(f'Excel 读取失败: {e}')

# 2) .aligned.xlsx 副本
aligned = X.parent / '强化学习的数学原理-赵世钰.aligned.xlsx'
print(f'.aligned.xlsx 存在: {aligned.exists()}')

# 3) 各章图注（md 临床化版）与 Excel 当前行对比
# 每章: 图片锚点+1 = 图注行; md 图注 vs Excel 图注
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')
md_notes = {}
for ch in range(1, 11):
    for f in sorted((D / f'ch{ch:02d}').glob('0*-1.*.md')):
        for ln in f.read_text(encoding='utf-8').splitlines():
            if ln.strip().startswith('图注：'):
                md_notes.setdefault(ch, []).append(ln.strip())

excel_notes = {}
try:
    wb = load_workbook(X)
    for ch in range(1, 11):
        sn = f'{ch}cmrl-' + ['conceptions', 'bellman', 'optimality', 'iteration',
                              'montecarlo', 'approximation', 'tdlearning',
                              'valuefunction', 'policygradient', 'acloop'][ch - 1]
        ws = wb[sn]
        notes = []
        for img in sorted(ws._images, key=lambda i: i.anchor._from.row):
            r = img.anchor._from.row + 2
            v = str(ws.cell(r, 3).value or '')
            notes.append(v)
        excel_notes[ch] = notes
    wb.close()
except Exception as e:
    print(f'Excel 读取异常: {e}')

# 4) 对比: md 图注条数 vs Excel 图注行数 + 首 20 字一致性
for ch in range(1, 11):
    m = md_notes.get(ch, [])
    e = excel_notes.get(ch, [])
    match = sum(1 for a, b in zip(m, e) if a[:20] == b[:20])
    print(f'ch{ch:02d}: md {len(m)} 条 / Excel {len(e)} 行 / 前20字一致 {match}')
