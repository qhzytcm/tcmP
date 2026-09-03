# -*- coding: utf-8 -*-
"""
补缺填失: 
① ICD11统一命名 增补「平台db校准」列（对齐中基v3.1模式, 英文术语以db为准）
② 病证单元库 增补 分子靶标/ICD11桥接码/六经归属 三列（对齐平台DSU schema）
"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)

# ═══ ① ICD11统一命名 增补「平台db校准」列 ═══
ws = wb['ICD11统一命名']
# 表头在第2行: 平台病码|内经病证/概念|ICD-11码|英文术语|状态依据
# 增补 F列「平台db校准」: db权威术语(entities表)
CAL = {
    'MB48.0': 'db: Vertigo ✔',
    '8B11': 'db: Cerebral ischaemic stroke（非 infarction）✔',
    'MD81': 'db: Abdominal or pelvic pain ✔',
    'DD91.2': 'db: Functional diarrhoea ✔',
    'ME84.2': 'db: Low back pain ✔',
    'ME06': 'db: Chronic enteritis of uncertain aetiology（痹证对应存疑, 待临床转化卷）',
    'MD12': 'db: Cough ✔',
    '5A11': 'db: Type 2 diabetes mellitus ✔',
    'MG30': 'db: Chronic pain ✔',
    'MB2?': '待查 db（pending）',
    '1F4A?': '待查 db（pending）',
    'MF5A?': '待查 db（pending）',
    'ME10.1?': '待查 db（pending）',
    '1B54?': '待查 db（pending）',
    '6A60?': '待查 db（pending）',
    'MG43?': '待查 db（pending）',
    '—': '内经专论, 无直接对应（概念桥接）',
}
ws.cell(row=2, column=6, value='平台db校准(entities表)')
n = 0
for row in ws.iter_rows(min_row=3, max_col=3):
    code_cell = row[2]
    if code_cell.value:
        code = str(code_cell.value).strip()
        cal = CAL.get(code, '待核')
        ws.cell(row=code_cell.row, column=6, value=cal)
        n += 1
ws.column_dimensions['F'].width = 50
print(f"ICD11 平台db校准列: {n} 行")

# ═══ ② 病证单元库 增补三列 ═══
ws = wb['病证单元库']
# 表头第2行: 编码|病证名|内经篇目溯源|病机要点|脏腑归属|临床对应|平台落点
# 增补 H分子靶标 I ICD11桥接码 J 六经归属
BZU_EXT = {
    'BZ-HJ-01': ('(风证: 靶标待临床卷)', 'MB48.0/MD12 等按证候', '—'),
    'BZ-HJ-02': ('(痹证: 炎症因子待填)', 'ME06', '—'),
    'BZ-HJ-03': ('(痿证: 神经肌肉待填)', 'MB2?', '阳明经'),
    'BZ-HJ-04': ('(厥证: 循环待填)', 'MG30/MD12', '—'),
    'BZ-HJ-05': ('(咳: 气道炎症待填)', 'MD12', '太阴/阳明'),
    'BZ-HJ-06': ('(疟: 疟原虫待填)', '1F4A?', '少阳'),
    'BZ-HJ-07': ('(热病: 炎症因子待填)', '—', '六经'),
    'BZ-HJ-08': ('(水肿: 水钠代谢待填)', 'MF5A?', '少阴/太阴'),
    'BZ-HJ-09': ('(胀: 胃肠动力待填)', '—', '太阴/阳明'),
    'BZ-HJ-10': ('(消渴: 胰岛素抵抗待填)', '5A11', '—'),
    'BZ-HJ-11': ('(黄疸: 胆红素代谢待填)', 'ME10.1?', '太阴'),
    'BZ-HJ-12': ('(积聚: 增殖信号待填)', '—', '厥阴'),
    'BZ-HJ-13': ('(痛: 痛觉通路待填)', 'MG30', '诸经'),
    'BZ-HJ-14': ('(痈疽: 感染免疫待填)', '1B54?', '诸经'),
    'BZ-HJ-15': ('(癫狂: 神经递质待填)', '6A60?', '—'),
    'BZ-HJ-16': ('(梦: 睡眠神经待填)', '—', '—'),
    'BZ-HJ-17': ('(虚: 代谢免疫待填)', 'MG43?', '诸经'),
}
ws.cell(row=2, column=8, value='分子靶标(molecular_targets)')
ws.cell(row=2, column=9, value='ICD11桥接码')
ws.cell(row=2, column=10, value='六经归属')
n = 0
for row in ws.iter_rows(min_row=3, max_col=1):
    code_cell = row[0]
    if code_cell.value:
        code = str(code_cell.value).strip()
        if code in BZU_EXT:
            a, b, c = BZU_EXT[code]
            ws.cell(row=code_cell.row, column=8, value=a)
            ws.cell(row=code_cell.row, column=9, value=b)
            ws.cell(row=code_cell.row, column=10, value=c)
            n += 1
ws.column_dimensions['H'].width = 26
ws.column_dimensions['I'].width = 16
ws.column_dimensions['J'].width = 12
print(f"病证单元库 增补三列: {n} 行")

wb.save(PATH)
print(f"补缺完成: {PATH}")
