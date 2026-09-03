# -*- coding: utf-8 -*-
"""修复: 第27章注释补「悬疑/非零和」标注"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
ws = wb['27.灵枢·色诊禁服']

n = 0
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value == '注释' and row[2].value:
        v = str(row[2].value)
        orig = v
        if '非零和' not in v:
            if '悬疑存疑' in v:
                v = v.replace('悬疑存疑', '悬疑存疑（异说并存，非零和局）')
            else:
                v = v.rstrip('。') + '；并存（非零和局）。'
        if '悬疑' not in v:
            # 悬疑已在部分注释中补, 其余补尾注
            v = v.rstrip('。') + '；难解处悬疑存疑。'
        if v != orig:
            ws.cell(row=row[0].row, column=3, value=v)
            print(f"{row[1].value}: 已补 (len {len(orig)}→{len(v)})")
            n += 1

wb.save(PATH)
print(f"修复 {n} 条")
