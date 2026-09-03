# -*- coding: utf-8 -*-
"""ch05 R22 图注重写（完整 502 字）+ 回读校验"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
f = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\ch05\03-1.3.md')
note = ''
for ln in f.read_text(encoding='utf-8').splitlines():
    if ln.strip().startswith('图注：'):
        note = ln.strip()
        break
print(f'md 图注: {len(note)} 字')

wb = load_workbook(X)
ws = wb['5cmrl-montecarlo']
ws.cell(22, 3, note)
ws.cell(22, 3).font = Font(name='黑体', size=10)
ws.cell(22, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[22].height = (math.ceil(len(note) / 44) + 1) * 14 + 6
wb.save(X)

wb2 = load_workbook(X)
v = str(wb2['5cmrl-montecarlo'].cell(22, 3).value or '')
wb2.close()
print(f'R22 回读: {len(v)}/{len(note)} {"✅完整" if v == note else "⚠截断"}')
