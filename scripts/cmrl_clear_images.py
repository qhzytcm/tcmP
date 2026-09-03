# -*- coding: utf-8 -*-
"""清空章节 sheet 的浮动图（仅 1cmrl-10cmrl; 封面页等附属 sheet 不动）"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
total = 0
for sn in wb.sheetnames:
    if not sn[0].isdigit():   # 仅章节 sheet（1cmrl-... ~ 10cmrl-...）
        continue
    ws = wb[sn]
    imgs = getattr(ws, '_images', None) or []
    if imgs:
        total += len(imgs)
        ws._images.clear()
wb.save(X)
print(f'已清空章节 sheet 浮动图 {total} 幅（附属 sheet 保留）')
