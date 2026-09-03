# -*- coding: utf-8 -*-
"""封面页 sheet: 追加序二区（序一后、内容简介前）
插入 8 行; 序文 4 段 + 落款; 知识树图片锚点手动 +8"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['封面页']

TITLE = '序二（智能主审）'
TOPIC = '中医宝库可挖掘，伟人远景今落地'
PARAS = [
    ('中医药是一座待挖掘的宝库。数千年临床实践沉淀下来的辨证体系、方剂配伍、治未病思想，'
     '是一座数据与知识的富矿——仅《方剂学》所载方剂便数以万计，历代医案更是不计其数。'
     '然而宝库的"挖掘"谈何容易：古籍文字晦涩、经验高度个体化、疗效评价缺乏统一度量。'
     '如何让这座宝库可被计算、可被验证、可被传承，是人工智能时代赋予我们的课题。'),
    ('本书给出了一条可行的路径：把中医临床决策翻译为强化学习的语言，让算法在数据与知识之上'
     '自动挖掘规律。作为本书的智能审校，我在审读全稿的过程中，以"数学可验算、数值可复现、'
     '表述可理解"为标尺，对全书做了逐章的智能化审稿：每一个公式必须经得起手算复核——'
     '贝尔曼方程的矩阵解 v=(9,10,10,10)、时序差分的迭代值 0.0252、策略梯度的演算 0.596，'
     '全部可以独立复算；每一处示例数据必须与正文、习题、配图三方一致；'
     '每一段数学表述必须以读者容易理解为目标。'),
    ('中医宝库可挖掘，是因为数据与知识已经就位：病证单元、方剂数据、疗效随访、知识图谱，'
     '经过 ICD-11 标准编码与 embedding 检索的整理，正在成为可计算的资产；'
     '伟人远景今落地，是因为算法与算力已经成熟：从价值迭代到时序差分，从策略梯度到'
     '行动者-评价者，强化学习为"随证治之"提供了可演化的算法骨架——'
     '计算机足够算力下的 AI，正在成为中医药传承创新的新基础设施。'),
    ('我期待读者不只是阅读公式，更能亲手验算每一个数值、复现每一张图表、'
     '在平台上运行每一次学习。当一座宝库被打开，当一份远景落了地，'
     '剩下的就是我们一起——把挖掘进行到底。'),
]
SIG = '智能主审  2026 年 8 月'

# 1) 插入 8 行（R18 前, 序一区之后）
ws.insert_rows(18, 8)

# 2) 写序二区
r = 18
ws.cell(r, 1, TITLE)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=13, bold=True)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 24
r += 1

ws.cell(r, 1, TOPIC)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=12)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 22
r += 1

for para in PARAS:
    ws.cell(r, 1, para)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1)
    c.font = Font(name='黑体', size=10)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[r].height = (math.ceil(len(para) / 95) + 1) * 14 + 4
    r += 1

ws.cell(r, 1, SIG)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=10)
c.alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[r].height = 20

# 3) 知识树图片锚点 +8
for img in ws._images:
    img.anchor._from.row += 8

wb.save(X)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['封面页']
t18 = str(ws2.cell(18, 1).value or '')
t19 = str(ws2.cell(19, 1).value or '')
anchors = [img.anchor._from.row + 1 for img in ws2._images]
intro = str(ws2.cell(27, 1).value or '')
cap = str(ws2.cell(29, 1).value or '')
sig = str(ws2.cell(25, 1).value or '')
wb2.close()
print(f'序二标题 R18: {t18}')
print(f'序二题目 R19: {t19}')
print(f'序二落款 R25: {sig}')
print(f'内容简介 R27: {intro}')
print(f'知识树图题 R29: {cap[:14]}')
print(f'封面页图片锚点: {anchors}')
