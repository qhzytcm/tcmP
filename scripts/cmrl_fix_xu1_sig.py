# -*- coding: utf-8 -*-
"""补写序一落款行 R17（XU1[7] 漏写）"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['封面页']
ws.cell(17, 1, '中医药主审  2026 年 8 月')
ws.merge_cells(start_row=17, start_column=1, end_row=17, end_column=3)
c = ws.cell(17, 1)
c.font = Font(name='黑体', size=10)
c.alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[17].height = 20
wb.save(X)

wb2 = load_workbook(X)
v = str(wb2['封面页'].cell(17, 1).value or '')
wb2.close()
print(f'落款 R17: {v} {"✅" if "中医药主审" in v else "⚠"}')
