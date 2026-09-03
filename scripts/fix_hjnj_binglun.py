# -*- coding: utf-8 -*-
"""
修复: 16.8 拆分 — 刺法论(16.8) + 本病论(16.9), 补本病论核心原文
使素问 81 篇篇名齐全（66-74: 天元纪/五运行/六微旨/气交变/五常政/六元正纪/至真要大/刺法/本病 = 9 篇）
"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
ws = wb['16.素问·运气七篇']

# 1. 修改 16.8 篇名（去掉"与本病论"）
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value == '篇名' and row[2].value and '16.8' in str(row[1].value):
        ws.cell(row=row[0].row, column=3, value='遗篇·刺法论（素问第72篇）')
        print(f"16.8 篇名已改: {ws.cell(row=row[0].row, column=3).value}")
    # 通读也更新
    if row[0].value == '通读' and row[1].value == '16.8.0':
        ws.cell(row=row[0].row, column=3, value='通读：遗篇·刺法论主"升降不前，气交有变"之刺法预防（刺法可以却病全神），提出"正气存内，邪不可干，避其毒气"之疫病防治名句，虽为后世补遗（宋代刘温舒补入），然其"刺法预防、扶正却邪"之思想与运气病机相发明。')
        print("16.8 通读已改")

# 2. 定位 16.8 注释行（最后一行），在其后插入 16.9 本病论
last_row = ws.max_row
# 找到 16.8 注释行号
insert_at = None
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value == '注释' and row[1].value == '16.8.1.9':
        insert_at = row[0].row
        break
if insert_at is None:
    # 找 16.8 最后一行
    for row in ws.iter_rows(min_row=2, max_col=3):
        if row[1].value and str(row[1].value).startswith('16.8'):
            insert_at = row[0].row
print(f"16.8 注释行: R{insert_at}")

NEW_ROWS = [
    ['篇名', '16.9', '遗篇·本病论（素问第73篇）'],
    ['通读', '16.9.0', '通读：遗篇·本病论论"气交失守"之本病原委——"天地气乱，民病暴至"，详述人气不足、天气虚之"三虚"相感致暴病（暴亡/暴厥/暴热等），提出"神游失守"与刺法扶正之治，为疫病与运气失守之专论。'],
    ['原文', '16.9.1.1', '黄帝问曰：天元九窒，余已知之，愿闻气交，何名失守？岐伯曰：谓其上下升降，迁正退位，各有经论，上下各有不前，故名失守也。是故气交失易位，气交乃变，异恒常矣。'],
    ['原文', '16.9.1.2', '帝曰：愿闻气交遇会胜抑之由，变成民病，轻重何如？岐伯曰：胜相会，抑伏使然。是故辰戌之岁，木气升之，主逢天柱，胜而不前。又遇庚戌，金运先天，中运胜之，忽然不前。木运升天，金乃抑之，升而不前，即清生风少，肃杀于春，露霜复降，草木乃萎。民病温疫早发，咽嗌乃干，两胁满，肢节皆痛。'],
    ['原文', '16.9.1.3', '帝曰：愿卒闻之。岐伯曰：人气不足，天气如虚，人神失守，神光不聚，邪鬼干人，致有夭亡。'],
    ['原文', '16.9.1.4', '帝曰：其发何如？岐伯曰：木疫之至，民病四肢不举，两胁满，善太息，或目黄，或爪甲枯，皆木运不及，肝气内郁也。'],
    ['注释', '16.9.1.5', '"人气不足，天气如虚，人神失守"："三虚"（人虚+天虚+失守）致疫之论——①以正气不足、运气失守相感释（通行），为"正虚邪凑"在疫病之发挥；②遗篇"鬼干"之语，或释为古代疫病观之表述（非鬼神迷信），或从现代流行病学相参。并存（非零和局）。又本篇与《刺法论》互为表里（一论刺法预防，一论病发之由）。'],
]

# 插入（从 insert_at+1 起逐行插入）
for i, row_data in enumerate(NEW_ROWS):
    r = insert_at + 1 + i
    ws.insert_rows(r)
    ws.cell(row=r, column=1, value=row_data[0])
    ws.cell(row=r, column=2, value=row_data[1])
    ws.cell(row=r, column=3, value=row_data[2])
    print(f"插入 R{r}: {row_data[0]} {row_data[1]}")

wb.save(PATH)
print(f"完成: {PATH}")
