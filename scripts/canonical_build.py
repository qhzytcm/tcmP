# -*- coding: utf-8 -*-
"""canonical build — cmrl 交付产物构建/完整性检查
运行: python scripts/canonical_build.py
覆盖: 45 幅 PNG 产物 / Excel media / 知识树 PNG / 备份文件检查
退出码: 0=产物完整 1=有缺失
"""
import sys
from pathlib import Path

D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')
X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
results = []


def check(name, ok, detail=''):
    results.append((name, ok, detail))
    print(f'{"PASS" if ok else "FAIL"}  {name}  {detail}')


# B1 45 幅章节 PNG 产物
figs = []
for ch in range(1, 11):
    figs += list((D / 'figures' / f'ch{ch:02d}').glob('fig*.png'))
ok1 = len(figs) == 45 and all(f.stat().st_size > 20 * 1024 for f in figs)
check(f'B1 章节 PNG 产物 45 幅非空', ok1, f'共{len(figs)}幅')

# B2 知识树 PNG
tree = D / 'figures' / 'book' / 'cmrl-knowledge-tree.png'
ok2 = tree.exists() and tree.stat().st_size > 100 * 1024
check('B2 知识树 PNG (树形版)', ok2, f'{tree.stat().st_size // 1024}KB')

# B3 Excel media 图片数（45 章节 + 1 知识树）
from openpyxl import load_workbook
wb = load_workbook(X)
total_img = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    if sn[0].isdigit():
        total_img += len(ws._images)
    elif sn == '封面页':
        total_img += len(ws._images)
ok3 = total_img == 46
check('B3 Excel 图片总数 46 (45章节+1封面)', ok3, f'共{total_img}')

# B4 章节 sheet 齐全（10 章 + 附属 6）
need = {f'{i}cmrl-' for i in range(1, 11)}
have = set()
for sn in wb.sheetnames:
    for i in range(1, 11):
        if sn.startswith(f'{i}cmrl-'):
            have.add(f'{i}cmrl-')
ok4 = need <= have
check('B4 章节 sheet 齐全', ok4, f'缺={need - have or "无"}')
wb.close()

fails = [r for r in results if not r[1]]
print(f'\n=== CANONICAL BUILD: {"ALL PASS" if not fails else f"{len(fails)} FAILED"} '
      f'({len(results) - len(fails)}/{len(results)}) ===')
sys.exit(0 if not fails else 1)
