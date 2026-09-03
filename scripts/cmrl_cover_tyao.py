# -*- coding: utf-8 -*-
"""封面页 sheet: 追加内容提要区（内容简介前）
插入 8 行: 标题1 + 韵文6 + 正文版1; 知识树图片锚点手动 +8"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['封面页']

TITLE = '内容提要（主编）'
VERSE = [
    '中医场景强化学习，状态动作奖励齐；',
    '四诊合参是状态，方剂动作疗效励。',
    '贝尔曼方程做账本，价值迭代求最优；',
    '时序差分走一步，Q学习向最优走。',
    '行动者配评价者，六者协同幸福绕；',
    '最适而非最优化，四个适度身心安。',
    '提高认知开篇立，确认信仰中编牢；',
    '追求幸福下编成，随证治之算法照。',
]
BODY = ('本书把强化学习的数学理论完整讲进中医药场景：四诊合参是状态，治法方剂是动作，'
        '疗效反馈是奖励，辨证论治是策略。上编提高认知（1-3 章：状态动作奖励/决策过程模型/'
        '统计估计随机逼近）；中编确认信仰（4-6 章：贝尔曼账本/价值策略迭代/时序差分 Q 学习）；'
        '下编追求幸福（7-10 章：四个最适度/函数近似/策略梯度/行动者-评价者与六者协同）。'
        '全书以"最适度而非最优化"为哲学主线，知识树须根干枝叶俱全：读者沿树而上，'
        '无需跳跃，不必自悟，例题皆可手算，插图皆可复看，明理过渡丝滑，衔接逻辑顺溜。')

# 1) 插入 8 行（R27 内容简介前）
ws.insert_rows(27, 8)

# 2) 写内容提要区
r = 27
ws.cell(r, 1, TITLE)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=13, bold=True)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 24
r += 1
for line in VERSE:
    ws.cell(r, 1, line)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(r, 1)
    c.font = Font(name='黑体', size=10)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 18
    r += 1
ws.cell(r, 1, BODY)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws.cell(r, 1)
c.font = Font(name='黑体', size=10)
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[r].height = (math.ceil(len(BODY) / 95) + 1) * 14 + 4

# 3) 知识树图片锚点 +8
for img in ws._images:
    img.anchor._from.row += 8

wb.save(X)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['封面页']
t27 = str(ws2.cell(27, 1).value or '')
v28 = str(ws2.cell(28, 1).value or '')
anchors = [img.anchor._from.row + 1 for img in ws2._images]
intro = str(ws2.cell(36, 1).value or '')
cap = str(ws2.cell(38, 1).value or '')
wb2.close()
print(f'内容提要标题 R27: {t27}')
print(f'韵文首行 R28: {v28}')
print(f'内容简介 R36: {intro}')
print(f'知识树图题 R38: {cap[:14]}')
print(f'封面页图片锚点: {anchors}')
