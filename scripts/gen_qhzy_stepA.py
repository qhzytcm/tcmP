# -*- coding: utf-8 -*-
"""
Step A: 复制种子《中医基础理论final.xlsx》为 qhzy 底稿
- 保留全部 25 个 Sheet 的单元格值（含样式类信息：合并单元格尽量保留）
- 重排 Sheet 顺序为 qhzy 27-Sheet 规范结构
"""
import openpyxl
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\DELL\Desktop\textbooks\中医基础理论final.xlsx'
DST = r'C:\Users\DELL\Desktop\qhzy-中医基础理论.xlsx'

src_wb = openpyxl.load_workbook(SRC)
dst_wb = openpyxl.Workbook()
dst_wb.remove(dst_wb.active)

# 目标顺序：前 8 个 0 系列 + 12 章正文 + 附篇 + 引擎 Sheet
ORDER = [
    '0改写说明', '0历史', '0映射1', '0映射2', '0映射3', '0封面', '0目录',
    '1绪论', '2哲思', '3脏腑', '4精神', '5经络', '6体窍',
    '7病因', '8发病', '9病机', '10治则', '11养康', '12平台度量',
    'ICD11统一命名', '病证单元库', '多智能体协同', '兼容性思维流', '可复用Skills',
]

RENAME = {'12平台度量': '13平台度量'}  # 平台章升级为第13章

def copy_sheet(src_ws, dst_wb, name):
    """复制工作表：单元格值 + 合并单元格 + 列宽（尽力保留）"""
    dst_ws = dst_wb.create_sheet(title=name)
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    for mc in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mc))
    # 列宽
    for col_idx, dim in src_ws.column_dimensions.items():
        if dim.width:
            letter = col_idx if isinstance(col_idx, str) else get_column_letter(col_idx)
            dst_ws.column_dimensions[letter].width = dim.width
    return dst_ws

for name in ORDER:
    if name in src_wb.sheetnames:
        target = RENAME.get(name, name)
        copy_sheet(src_wb[name], dst_wb, target)
        print(f"✔ 复制: {name} → {target} ({src_wb[name].max_row}行)")
    else:
        print(f"✘ 缺失: {name}")

dst_wb.save(DST)
print(f"\n底稿已保存: {DST}")
print(f"Sheet 数: {len(dst_wb.sheetnames)}")
for s in dst_wb.sheetnames:
    print("  ", s)
