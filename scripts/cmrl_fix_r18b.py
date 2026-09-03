# -*- coding: utf-8 -*-
"""ch6 fig6-2 图块下移修复（R18 截断）: 图题R17/图体R18(锚)/图注R19"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')

note = ''
for ln in (D / 'ch06' / '02-1.2.md').read_text(encoding='utf-8').splitlines():
    if ln.strip().startswith('图注：'):
        note = ln.strip()
        break

wb = load_workbook(X)
ws = wb['6cmrl-approximation']

# 1) 图题 R16 -> R17, R16 清空
cap16 = str(ws.cell(16, 3).value or '')
print(f'图题 R16: {cap16[:22]}')
ws.cell(17, 3, cap16)
ws.cell(17, 3).font = Font(name='黑体', size=12)
ws.cell(17, 3).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[17].height = 22
ws.cell(16, 3, None)

# 2) 插入 1 行 R19 前
ws.insert_rows(19, 1)

# 3) 锚点 >= 17 的 +1（fig6-2 R17->R18, fig6-3/6-4 后续 +1）
for img in ws._images:
    if img.anchor._from.row + 1 >= 17:
        img.anchor._from.row += 1

# 4) R18 清空（图体行, 图片覆盖）
ws.cell(18, 3, None)

# 5) 图注写 R19（新空行）
ws.cell(19, 3, note)
ws.cell(19, 3).font = Font(name='黑体', size=10)
ws.cell(19, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[19].height = (math.ceil(len(note) / 44) + 1) * 14 + 6

wb.save(X)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['6cmrl-approximation']
v19 = str(ws2.cell(19, 3).value or '')
anchors = sorted(img.anchor._from.row + 1 for img in ws2._images)
cap17 = str(ws2.cell(17, 3).value or '')
wb2.close()
print(f'图注 R19: {len(v19)}/{len(note)} {"✅" if v19 == note else "⚠截断"}')
print(f'图题 R17: {cap17[:20]}')
print(f'锚点: {anchors}')
