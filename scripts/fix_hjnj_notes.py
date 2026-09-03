# -*- coding: utf-8 -*-
"""
修复: 删除 5 条错位补注, 按正确篇内位置重插（编码续接篇内注释序号）
"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)

# 待修复的错位行: (sheet, 行号, 篇名)
BAD = {
    '2.素问·阴阳学说': [42],
    '3.素问·藏象': [36, 37],
    '4.素问·诊法总论': [43, 44],
}

# 删除错误行（从大到小删避免行号错位）
for s, rows in BAD.items():
    ws = wb[s]
    for r in sorted(rows, reverse=True):
        ws.delete_rows(r)
    print(f"{s}: 已删除 {len(rows)} 行错位注释")

# 正确的补注内容（篇名 → 注释文本）
NOTES = {
    '阴阳离合论': '异说并存（非零和局）：开阖枢之"枢"——①张介宾以枢为"枢轴转动之要"；②有注家以"枢"为少阳、少阴经脉居表里之间、主转输出入之机；今人有以"阀门-枢纽"模型释之者，与古说并存。',
    '移精变气论': '异说并存（非零和局）："祝由"与"移精变气"之关系——①王冰谓"祝说病由"即移易精气以变其气；②有学者认为移精变气为导引行气之术而祝由为祝祷之法，二者有别；两说并存，不改原文。',
    '玉版论要': '异说并存（非零和局）："揆度奇恒"——①王冰以揆度为"度病之深浅"、奇恒为"言奇病"；②有学者释"奇恒"为"异于恒常之脉证"，与"揆度"合为诊法总纲；措辞略异而义理相通，并存教学。',
    '五脏生成': '异说并存（非零和局）："诸脉者皆属于目"——①以"目为宗脉之所聚"释（肝开窍于目，诸经皆上注于目）；②有注家以"属于"训"统属于"，言脉气皆会于目系；二说并存。',
    '五脏别论': '异说并存（非零和局）："气口亦太阴也"——①王冰以气口为手太阴肺经之脉，肺朝百脉故独主五脏；②有学者强调"胃为水谷之海"为气口之本，五脏六腑之气味皆出于胃而变见于气口；一重肺、一重胃，两说并存。',
}

def insert_after(ws, row_idx, cat, code, text):
    ws.insert_rows(row_idx + 1)
    ws.cell(row=row_idx + 1, column=1, value=cat)
    ws.cell(row=row_idx + 1, column=2, value=code)
    ws.cell(row=row_idx + 1, column=3, value=text)

ch_sheets = ['1.素问·养生总纲','2.素问·阴阳学说','3.素问·藏象','4.素问·诊法总论','5.素问·脉学']
added = 0
for s in ch_sheets:
    ws = wb[s]
    # 扫描: 篇名行 → 下一篇名前最后一行
    # 收集所有行 (行号, 类型, 编码, 篇名)
    data = []
    cur_pn = None
    for row in ws.iter_rows(min_row=2, max_col=3):
        r, cat, code, val = row[0].row, row[0].value, row[1].value, row[2].value
        if cat == '篇名' and val:
            cur_pn = str(val).split('（')[0].strip()
        data.append((r, cat, code, val, cur_pn))
    # 每篇的最后数据行号
    last_of = {}
    for r, cat, code, val, pn in data:
        if pn:
            last_of[pn] = r
    # 在每篇末行后插入补注
    for pn, note in NOTES.items():
        if pn in last_of:
            anchor = last_of[pn]
            # 取该篇的节号（篇名行编码）
            sec_code = None
            for r, cat, code, val, ppn in data:
                if cat == '篇名' and ppn == pn and code:
                    sec_code = str(code).strip()
                    break
            if not sec_code:
                continue
            # 该篇已有注释的末序号: 找该篇内最后一个注释编码, 续接
            pian_codes = [str(code) for r, cat, code, val, ppn in data if ppn == pn and code and str(code).startswith(sec_code + '.')]
            # 取最大末段号
            last_sub = 0
            for c in pian_codes:
                parts = c.split('.')
                if len(parts) >= 4 and parts[0] == sec_code.split('.')[0]:
                    try:
                        last_sub = max(last_sub, int(parts[3]))
                    except (ValueError, IndexError):
                        pass
            new_code = f'{sec_code}.1.{last_sub + 1}' if last_sub else f'{sec_code}.1.1'
            insert_after(ws, anchor, '注释', new_code, note)
            added += 1
            print(f"  {s}: {pn} 补注→{new_code} (锚点R{anchor})")

wb.save(PATH)
print(f"修复完成: 重插 {added} 条补注")
