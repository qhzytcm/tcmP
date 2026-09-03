# -*- coding: utf-8 -*-
"""修复 ICD11统一命名 校准列（诊断学表头为合并单元格'平台病码/证码'）"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中医诊断学.xlsx'
wb = openpyxl.load_workbook(DST)
ws = wb['ICD11统一命名']

# 打印前4行定位
for r in range(1, 5):
    vals = [str(ws.cell(row=r, column=c).value) for c in range(1, 7) if ws.cell(row=r, column=c).value]
    print(f"R{r}: {vals}")

# 表头行: 含'平台病码'或'统一命名'的行
hdr_row = None
for r in range(1, 6):
    v1 = ws.cell(row=r, column=1).value
    v2 = ws.cell(row=r, column=2).value
    if v1 and ('平台病码' in str(v1) or '平台病码/证码' in str(v1)):
        hdr_row = r
        break
    if v2 and '统一命名' in str(v2):
        hdr_row = r
        break
print("表头行:", hdr_row)

calib = {
    'B1': ('CA00', 'Acute nasopharyngitis', '✅ 一致'),
    'B2': ('MD12', 'Cough', '✅ 一致'),
    'B3': ('CA23', 'Asthma', '✅ 一致'),
    'B4': ('MD11.5', 'Dyspnoea', '🔧 细化: MD11.5(原MD11大类)'),
    'B5': ('MC81.2', 'Palpitations', '🔧 细化: MC81.2(原MC81大类)'),
    'B6': ('BA40', 'Angina pectoris', '✅ 一致'),
    'B7': ('7A00', 'Chronic insomnia', '✅ 一致'),
    'B8': ('MB48.0', 'Vertigo', '🔧 修正: 原MB51=上肢瘫痪; 正确MB48.0 Vertigo'),
    'B9': ('8B11', 'Cerebral ischaemic stroke', '🔧 修正: 原8B20=未特指卒中; 缺血性=8B11'),
    'B10': ('MD81', 'Abdominal pain', '🔧 修正: 原MD90=恶心呕吐; 腹痛=MD81'),
    'B11': ('MD90', 'Nausea or vomiting', '🔧 修正: 原ME08=胃肠胀气; 恶心呕吐=MD90'),
    'B12': ('DD91.2', 'Functional diarrhoea', '🔧 修正: 原ME05=排便习惯改变; 功能性腹泻=DD91.2'),
    'B13': ('DD91.1', 'Functional constipation', '🔧 修正: 原ME06=慢性肠炎; 功能性便秘=DD91.1'),
    'B14': (None, '(待WHO-API回填)', '⏳ pending_who_api'),
    'B15': ('MG27', 'Oedema', '⚠️ db中MG27=出血未归类, 水肿码待复核'),
    'B16': ('5A11', 'Type 2 diabetes mellitus', '✅ 一致'),
    'B17': (None, '(待WHO-API回填)', '⏳ pending_who_api'),
    'B18': ('GC00', 'Cystitis', '✅ 一致'),
    'B19': ('ME84.2', 'Low back pain', '🔧 修正: 原ME83=风湿; 下背痛=ME84.2'),
    'B20': ('8A81', 'Tension-type headache', '🔧 细化: 原8A80=偏头痛; 头痛按型8A80/8A81/8A83'),
}

if hdr_row:
    ws.cell(row=hdr_row, column=6, value='平台db校准(icd11_mms.db 2026-01)')
    n = 0
    for r in range(hdr_row + 1, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        if code and str(code).strip() in calib:
            new_code, title, status = calib[str(code).strip()]
            ws.cell(row=r, column=6, value=f"{status} | db权威: {title}" + (f" | 建议码: {new_code}" if new_code else ""))
            n += 1
    print(f"校准列写入 {n} 行")
else:
    print("!! 仍找不到表头行")

wb.save(DST)
print("完成")
