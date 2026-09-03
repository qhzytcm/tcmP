# -*- coding: utf-8 -*-
"""ch01 fig1-1 图块下移修复: 插入行 + 全部后续锚点+1 + 图注写 R9
原: 图题R6/图体R7(图锚)/图注R8(截断) -> 新: 图题R7/图体R8(图锚)/图注R9(完整)"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
D = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl')

note = ''
for ln in (D / 'ch01' / '01-1.1.md').read_text(encoding='utf-8').splitlines():
    if ln.strip().startswith('图注：'):
        note = ln.strip()
        break

wb = load_workbook(X)
ws = wb['1cmrl-conceptions']

# 1) 图题 R6 -> R7, R6 清空
cap6 = str(ws.cell(6, 3).value or '')
print(f'图题 R6: {cap6[:20]}')
ws.cell(7, 3, cap6)
ws.cell(7, 3).font = Font(name='黑体', size=12)
ws.cell(7, 3).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[7].height = 22
ws.cell(6, 3, None)

# 2) 插入 1 行在 R9 前（正文段下移）
ws.insert_rows(9, 1)

# 3) 全部图片锚点 >= 8 的 +1（fig1-1 锚 R7 -> R8, 其余 +1）
for img in ws._images:
    if img.anchor._from.row + 1 >= 8:
        img.anchor._from.row += 1

# 4) R8 清空（图体行, 图片覆盖）
ws.cell(8, 3, None)

# 5) 图注写 R9（新空行）
ws.cell(9, 3, note)
ws.cell(9, 3).font = Font(name='黑体', size=10)
ws.cell(9, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[9].height = (math.ceil(len(note) / 44) + 1) * 14 + 6

wb.save(X)

# 校验
wb2 = load_workbook(X)
ws2 = wb2['1cmrl-conceptions']
v9 = str(ws2.cell(9, 3).value or '')
anchors = sorted(img.anchor._from.row + 1 for img in ws2._images)
cap7 = str(ws2.cell(7, 3).value or '')
wb2.close()
print(f'图注 R9: {len(v9)}/{len(note)} {"✅" if v9 == note else "⚠截断"}')
print(f'图题 R7: {cap7[:18]}')
print(f'锚点: {anchors}')
