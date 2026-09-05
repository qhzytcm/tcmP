# -*- coding: utf-8 -*-
"""导出 武夷山(1).xlsx 全部 wy sheet 字幕 → _wyshan_text/wyNN.txt；打印 sheet 清单"""
import openpyxl
import os
from pathlib import Path

path = r"C:\Users\DELL\Desktop\武夷山(1).xlsx"
out_dir = Path(r"C:\Users\DELL\tcmP\scripts\_wyshan_text")
out_dir.mkdir(exist_ok=True)
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

names = []
for name in wb.sheetnames:
    if name.lower().startswith("wy"):
        names.append(name)
print("wy sheets:", names)

for name in names:
    ws = wb[name]
    lines = []
    for row in ws.iter_rows(values_only=True):
        # 列: 序号/开始时间/结束时间/文本内容
        seq, t0, t1, text = (list(row) + [None] * 4)[:4]
        if text is None or str(text).strip() == "":
            continue
        lines.append(str(text).strip())
    (out_dir / f"{name}.txt").write_text("\n".join(lines), encoding="utf-8")
    print(f"{name}: {len(lines)} 条字幕 → {name}.txt")

# 也导出 0认知特色 全文
ws = wb["0认知特色"]
rows = []
for row in ws.iter_rows(values_only=True):
    a, b = (list(row) + [None] * 2)[:2]
    if a is None and b is None:
        continue
    rows.append(f"{a or ''}\t{b or ''}")
(out_dir / "0认知特色.txt").write_text("\n".join(rows), encoding="utf-8")
print("0认知特色.txt 导出完成")
