# -*- coding: utf-8 -*-
"""runner: tempfile.mkstemp OS-safe 动态生成 hermes-verify- 脚本 → 运行 → 自动清理（ad-hoc）"""
import os, subprocess, sys, tempfile

BODY = r'''# -*- coding: utf-8 -*-
"""hermes-verify-hjnj-ch16-final.py — qhzy-黄帝内经 第16章重写（全文照抄+悬疑注释）聚焦验证（ad-hoc, tempfile OS-safe）"""
import sys, re, os, json
import openpyxl

print(f"[hermes-verify] 脚本: {os.path.abspath(__file__)}")

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
fails, passes = [], []

def check(name, cond, detail=""):
    (passes if cond else fails).append(f"{name}{(' | ' + detail) if detail and not cond else ''}")

wb = openpyxl.load_workbook(PATH, read_only=True)

# 1. Sheet 顺序: 15→16→17
names = wb.sheetnames
i15, i16, i17 = names.index('15.素问·调经缪刺'), names.index('16.素问·运气七篇'), names.index('17.素问·医论杂篇')
check("16章位置(15-16-17)", i16 == i15 + 1 and i17 == i16 + 1, f"15@{i15} 16@{i16} 17@{i17}")

# 2. 16章结构: 9篇/原文299/通读9/注释9/悬疑/非零和
ws = wb['16.素问·运气七篇']
rows = []
for row in ws.iter_rows(min_row=2, max_col=3):
    cells = ['' if c.value is None else str(c.value).strip() for c in row]
    if any(cells):
        rows.append(cells)
pnames = [r[2] for r in rows if r[0] == '篇名']
check("篇名9", len(pnames) == 9, f"实际{len(pnames)}")
for kw in ['天元纪大论','五运行大论','六微旨大论','气交变大论','五常政大论','六元正纪大论','刺法论','本病论','至真要大论']:
    check(f"篇[{kw}]", any(kw in p for p in pnames))
n_or = sum(1 for r in rows if r[0] == '原文')
n_td = sum(1 for r in rows if r[0] == '通读')
n_no = sum(1 for r in rows if r[0] == '注释')
check(f"原文{n_or}段(全文≥200)", n_or >= 200, f"实际{n_or}")
check("通读9", n_td == 9, f"实际{n_td}")
check("注释9", n_no == 9, f"实际{n_no}")
zt = "\n".join(r[2] for r in rows if r[0] == '注释')
check("悬疑标注≥5", zt.count('悬疑') >= 5, f"悬疑{zt.count('悬疑')}处")
check("非零和标注", '非零和' in zt)

# 3. 素问81篇齐全
pian_total = 0
for s in wb.sheetnames:
    m = re.match(r'^(\d+)\.', s)
    if m and 1 <= int(m.group(1)) <= 17:
        w = wb[s]
        pian_total += sum(1 for row in w.iter_rows(min_row=2, max_col=3) if row[0].value == '篇名')
check("素问81篇", pian_total == 81, f"实际{pian_total}")

# 4. 全文照抄关键句（运气七篇核心, 不选取）
yw = "\n".join(r[2] for r in rows if r[0] == '原文')
for ks in ['甲己之岁，土运统之','丹天之气，经于牛女戊分','亢则害，承乃制',
           '岁木太过，风气流行','木曰敷和','木郁达之，火郁发之',
           '正气存内，邪不可干','人气不足，天气如虚','诸风掉眩，皆属于肝',
           '热因寒用，寒因热用','有故无殒','主病之谓君',
           '太虚寥廓，肇基化元','天枢之上，天气主之','根于中者，命曰神机',
           '西北之气散而寒之','发表不远热，攻里不远寒','塞因塞用，通因通用']:
    check(f"原文[{ks[:10]}]", ks in yw)

# 5. 0系列/引擎未破坏
t1 = "\n".join(str(c.value) for row in wb['0映射1'].iter_rows() for c in row if c.value)
t2 = "\n".join(str(c.value) for row in wb['0映射2'].iter_rows() for c in row if c.value)
t3 = "\n".join(str(c.value) for row in wb['0映射3'].iter_rows() for c in row if c.value)
tt = "\n".join(str(c.value) for row in wb['0目录'].iter_rows() for c in row if c.value)
check("0映射1邹纯朴", '邹纯朴' in t1)
check("0映射2翟双庆", '翟双庆' in t2)
check("0映射3李梢", '李梢' in t3)
check("0目录素问灵枢", '素问81篇' in tt and '灵枢81篇' in tt)

# 6. 进度文件
p = r'C:\Users\DELL\tcmP\scripts\progress\hjnj_writing_progress.json'
d = json.load(open(p, encoding='utf-8'))
check("进度17/31", d['done'] == 17, str(d['done']))

wb.close()
print(f"\nPASS {len(passes)} | FAIL {len(fails)}")
for p2 in passes:
    print("  ✔", p2)
for f in fails:
    print("  ✘", f)
sys.exit(1 if fails else 0)
'''

fd, tmp_path = tempfile.mkstemp(prefix='hermes-verify-hjnj-ch16-final-', suffix='.py', dir=tempfile.gettempdir())
with os.fdopen(fd, 'w', encoding='utf-8') as f:
    f.write(BODY)
print(f"[hermes-verify] 临时脚本: {tmp_path}")

PY = r'C:\Users\DELL\AppData\Local\hermes\hermes-agent\venv\Scripts\python.exe'
try:
    r = subprocess.run([PY, tmp_path], capture_output=True, text=True, encoding='utf-8', timeout=180)
    print(r.stdout)
    if r.stderr:
        print("STDERR:", r.stderr[-1200:])
    sys.exit(r.returncode)
finally:
    try:
        os.remove(tmp_path)
        print(f"[hermes-verify] 已清理: {tmp_path} 存在={os.path.exists(tmp_path)}")
    except OSError:
        pass
