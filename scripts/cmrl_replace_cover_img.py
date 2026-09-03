# -*- coding: utf-8 -*-
"""替换封面页知识树图片字节（锚点保持 R40, 用新 PNG）"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PImg

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
TREE = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\figures\book\cmrl-knowledge-tree.png')

wb = load_workbook(X)
ws = wb['封面页']
old_anchor = None
if ws._images:
    old_anchor = ws._images[0].anchor._from.row + 1
    ws._images.clear()
    print(f'旧图已移除（锚点 R{old_anchor}）')

w, h = PImg.open(TREE).size
img = XLImage(str(TREE))
img.width = 480
img.height = int(480 * h / w)
row = old_anchor if old_anchor else 40
ws.add_image(img, f'A{row}')
ws.row_dimensions[row].height = img.height * 0.75
wb.save(X)
print(f'新图已插入（锚点 R{row}, {img.width}x{img.height}）')

# 校验
wb2 = load_workbook(X)
ws2 = wb2['封面页']
anchors = [img.anchor._from.row + 1 for img in ws2._images]
n_img = len(ws2._images)
wb2.close()
print(f'封面页图片: {n_img} 幅, 锚点 {anchors}')
