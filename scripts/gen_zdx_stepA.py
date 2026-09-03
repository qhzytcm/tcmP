# -*- coding: utf-8 -*-
"""
Step A: 复制诊断学种子为 qhzy 底稿（保留 26 Sheet 原顺序与内容）
"""
import openpyxl
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\DELL\Desktop\textbooks\中医诊断学.xlsx'
DST = r'C:\Users\DELL\Desktop\qhzy-中医诊断学.xlsx'

src_wb = openpyxl.load_workbook(SRC)
dst_wb = openpyxl.Workbook()
dst_wb.remove(dst_wb.active)

def copy_sheet(src_ws, dst_wb, name):
    dst_ws = dst_wb.create_sheet(title=name)
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    for mc in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mc))
    for col_idx, dim in src_ws.column_dimensions.items():
        if dim.width:
            letter = col_idx if isinstance(col_idx, str) else get_column_letter(col_idx)
            dst_ws.column_dimensions[letter].width = dim.width
    return dst_ws

for name in src_wb.sheetnames:
    copy_sheet(src_wb[name], dst_wb, name)
    print(f"✔ 复制: {name} ({src_wb[name].max_row}行)")

dst_wb.save(DST)
print(f"\n底稿已保存: {DST}")
print(f"Sheet 数: {len(dst_wb.sheetnames)}")
