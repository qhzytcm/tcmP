# -*- coding: utf-8 -*-
"""隔离测试: 全新文件写 FULL62（含焠㶼 生僻字）是否截断"""
import openpyxl

FULL = '病在肾，愈在春；春不愈，甚于长夏；长夏不死，持于秋；起于冬。禁犯焠㶼热食温炙衣。肾病者，夜半慧，四季甚，下晡静。肾欲坚，急食苦以坚之，用苦补之，咸写之。'
print(f"FULL len={len(FULL)}")

# 测试1: 全新工作簿
PATH1 = r'C:\Users\DELL\AppData\Local\Temp\test_full1.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1, value=FULL)
wb.save(PATH1)
wb.close()

wb2 = openpyxl.load_workbook(PATH1)
v = wb2.active.cell(row=1, column=1).value
print(f"全新工作簿: len={len(v)} {'✔' if len(v) == len(FULL) else '✘截断'}")

# 测试2: 复制现有文件结构写入（load 现有 46 sheet 文件）
PATH2 = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb3 = openpyxl.load_workbook(PATH2)
ws3 = wb3['6.素问·经脉血气']
# 写入临时单元格 999 行
ws3.cell(row=1, column=5, value=FULL)
wb3.save(PATH2)
wb3.close()

wb4 = openpyxl.load_workbook(PATH2)
v2 = wb4['6.素问·经脉血气'].cell(row=1, column=5).value
print(f"现有文件: len={len(v2)} {'✔' if len(v2) == len(FULL) else '✘截断'}")

# 清理测试列
wb5 = openpyxl.load_workbook(PATH2)
ws5 = wb5['6.素问·经脉血气']
ws5.cell(row=1, column=5).value = None
wb5.save(PATH2)
wb5.close()
print("测试列已清理")
