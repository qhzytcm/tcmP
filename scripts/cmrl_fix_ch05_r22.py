# -*- coding: utf-8 -*-
"""ch05 第2幅图块下移一行（R16 模式）: 图题R21/图体R22(图片)/图注R23(完整)"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
f = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\ch05\03-1.3.md')
note = ''
for ln in f.read_text(encoding='utf-8').splitlines():
    if ln.strip().startswith('图注：'):
        note = ln.strip()
        break

wb = load_workbook(X)
ws = wb['5cmrl-montecarlo']
# 定位锚定 R21 的图片（图注 R22 → 图体 R21）
target = None
for img in ws._images:
    if img.anchor._from.row + 1 == 21:
        target = img
        break
if target is None:
    print('[MISS] 未找到锚定 R21 图片')
    raise SystemExit(1)
cap = str(ws.cell(20, 3).value or '')
# 1) 图题 R20 -> R21
ws.cell(21, 3, cap)
ws.cell(21, 3).font = Font(name='黑体', size=12)
ws.cell(21, 3).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[21].height = 22
ws.cell(20, 3, None)
# 2) 图片锚点 R21 -> R22
target.anchor._from.row += 1
# 3) 图注写 R23
ws.cell(23, 3, note)
ws.cell(23, 3).font = Font(name='黑体', size=10)
ws.cell(23, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[23].height = (len(note) // 44 + 1) * 14 + 6
# 4) R22 清空（图体行, 图片覆盖）
ws.cell(22, 3, None)
wb.save(X)

wb2 = load_workbook(X)
ws2 = wb2['5cmrl-montecarlo']
v23 = str(ws2.cell(23, 3).value or '')
anchors = sorted(img.anchor._from.row + 1 for img in ws2._images)
cap21 = str(ws2.cell(21, 3).value or '')
wb2.close()
print(f'图注 R23: {len(v23)}/{len(note)} {"✅" if v23 == note else "⚠截断"}')
print(f'图题 R21: {cap21[:24]}')
print(f'锚点: {anchors}')
