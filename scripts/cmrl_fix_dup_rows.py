# -*- coding: utf-8 -*-
"""清空 ch01 R8 残留 + ch10 R7 重复图注 → 重跑替换"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['1cmrl-conceptions']
v8 = str(ws.cell(8, 3).value or '')
if v8.startswith('图注：'):
    ws['C8'] = None
    print(f'ch01 R8 已清空（{len(v8)}字残留）')
else:
    print(f'ch01 R8 非图注残留（{len(v8)}字）, 未动')

ws = wb['10cmrl-acloop']
v7 = str(ws.cell(7, 3).value or '')
if v7.startswith('图注：'):
    ws['C7'] = None
    print(f'ch10 R7 已清空（{len(v7)}字重复图注）')
else:
    print(f'ch10 R7 非图注（{len(v7)}字）, 未动')
wb.save(X)

# 校验
wb2 = load_workbook(X)
for sn in ('1cmrl-conceptions', '10cmrl-acloop'):
    ws2 = wb2[sn]
    n = sum(1 for r in range(1, ws2.max_row + 1)
            if str(ws2.cell(r, 3).value or '').strip().startswith('图注：'))
    n_img = len(ws2._images)
    print(f'{sn}: 图注行 {n}, 图数 {n_img}')
wb2.close()
