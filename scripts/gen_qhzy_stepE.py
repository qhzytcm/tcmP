# -*- coding: utf-8 -*-
"""Step E: 可复用Skills 补齐到 10 个规格（原6 + 已增2 + 补2）"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-中医基础理论.xlsx'
wb = openpyxl.load_workbook(DST)
ws = wb['可复用Skills']

# 找到最后一行有内容的位置
start = ws.max_row + 1
while start > 1:
    has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, 7))
    if has:
        break
    start -= 1

new_rows = [
    ['tcm-zhongji-tizhi-bianbie', '体质辨识数字化: 以《灵枢·阴阳二十五人》五形人×五音=二十五人为纲, 王琦九分法为现代参考, 输出"体质→易感病证→养生方案"双轨教学链。', '用户要求"体质辨识/阴阳二十五人/王琦九分法/体质-病证关联"时使用(对应本教材第12章)。', '① 建立五形人×五音=二十五人编码(如"上角木形人"); ② 挂接王琦九分法九类体质; ③ 经典↔九分法映射表(标注"教学类比,非定义"); ④ 体质节点关联易感病证单元(BZU); ⑤ 输出体质辨识量表结构化字段。', '经典与九分法非严格一一对应, 须明示映射为教学类比; 不将体质分类当作诊断结论; 体质量表条目须注明来源。', '体质节点编码与知识图谱 schema 一致; 映射表含"非定义"标注; 与 12.5 AI 实现节呼应。'],
    ['tcm-wuyun-liuqi-calculator', '五运六气推算工具: 输入年份→输出岁运/主客运/六气司天在泉格局, 附运气与发病的经典对应(供教学与科研假设检验)。', '用户要求"五运六气/运气推算/干支纪年/时令病预测"时使用(对应本教材附篇)。', '① 干支纪年换算(天干化运: 甲己土/乙庚金/丙辛水/丁壬木/戊癸火; 地支定气); ② 输出大运/主运/客运/主气/客气/司天在泉; ③ 关联经典运气病候表述; ④ 可选对接气象-疾病数据做假设检验。', '运气推算是历法模型, 不可作为个体诊疗依据; 经典病候表述标注出处; 数据检验须声明样本与统计方法。', '推算结果与权威运气历表抽查一致; 病候关联标注《素问》篇目; AI 只做舟楫不作医理替代。'],
]
for i, row in enumerate(new_rows):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=start + i, column=c, value=v)

wb.save(DST)
print("Step E 完成")
cnt = sum(1 for r in ws.iter_rows(max_col=1) if r[0].value and str(r[0].value).startswith('tcm-'))
print("tcm-* 规格总数:", cnt)
