# -*- coding: utf-8 -*-
"""清空 ch6 R16 旧图题 + R18 截断图注残留（索引赋值）"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['6cmrl-approximation']
ws['C16'] = None
ws['C18'] = None
wb.save(X)

wb2 = load_workbook(X)
ws2 = wb2['6cmrl-approximation']
v16 = str(ws2.cell(16, 3).value or '')
v18 = str(ws2.cell(18, 3).value or '')
n_notes = sum(1 for r in range(1, ws2.max_row + 1)
              if str(ws2.cell(r, 3).value or '').strip().startswith('图注：'))
n_img = len(ws2._images)
cap17 = str(ws2.cell(17, 3).value or '')
note19 = str(ws2.cell(19, 3).value or '')
wb2.close()
print(f'R16: {len(v16)}字 R18: {len(v18)}字')
print(f'图注行 {n_notes} = 图数 {n_img}: {"✅" if n_notes == n_img else "⚠"}')
print(f'图题 R17: {cap17[:16]} 图注 R19: {len(note19)}字')
