# -*- coding: utf-8 -*-
"""修复 6.2.1.7 截断（早期遗留 sharedStrings 问题）"""
import openpyxl

PATH = r'C:\Users\DELL\Desktop\qhzy-黄帝内经(通读注释版).xlsx'
wb = openpyxl.load_workbook(PATH)
ws = wb['6.素问·经脉血气']

FULL = '病在肾，愈在春；春不愈，甚于长夏；长夏不死，持于秋；起于冬。禁犯焠㶼热食温炙衣。肾病者，夜半慧，四季甚，下晡静。肾欲坚，急食苦以坚之，用苦补之，咸写之。'
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[1].value == '6.2.1.7':
        old = str(row[2].value)
        print(f"修复前 len={len(old)}: {old}")
        ws.cell(row=row[0].row, column=3, value=FULL)
        print(f"修复后 len={len(FULL)}: {FULL[:50]}...")
        break

wb.save(PATH)
print("已保存")
