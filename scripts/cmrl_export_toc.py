# -*- coding: utf-8 -*-
"""读取 0cmrl目录 sheet：统计层级分布，导出每章四级编码树到 drafts/cmrl/toc/"""
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

SRC = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
OUT = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\toc')
OUT.mkdir(parents=True, exist_ok=True)

wb = load_workbook(SRC, read_only=True)
ws = wb['0cmrl目录']
rows = []
for row in ws.iter_rows(values_only=True):
    if row[0] in ('章', '节', '小节', '段落'):
        rows.append(row)
wb.close()

cnt = Counter(r[0] for r in rows)
print('层级统计:', dict(cnt), '总行:', len(rows))

# 按章分组导出
cur_ch = None
buf = []
def flush():
    if cur_ch is None:
        return
    f = OUT / f'ch{cur_ch:02d}_toc.md'
    f.write_text('\n'.join(buf) + '\n', encoding='utf-8')
    print(f'[write] {f.name} ({len(buf)} 行)')

for r in rows:
    lvl, code, title, note = r[0], r[1], r[2], r[3]
    words = r[4]
    if lvl == '章':
        flush()
        cur_ch = int(code.replace('第', '').replace('章', ''))
        buf = [f'# {code} {title}  预算 {words} 字']
        buf.append(f'  内容定位: {note}')
    elif lvl == '节':
        buf.append(f'## {code} {title}  ({words}字)')
        if note:
            buf.append(f'  定位: {note}')
    elif lvl == '小节':
        buf.append(f'### {code} {title}  ({words}字)')
        if note:
            buf.append(f'  定位: {note}')
    else:
        buf.append(f'  - {code} {title}  ({words}字)')
flush()
print('导出完成 →', OUT)
