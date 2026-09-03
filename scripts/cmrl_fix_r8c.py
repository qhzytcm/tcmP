# -*- coding: utf-8 -*-
"""修正 fig1-1 锚点 R7 -> R8（图体行应在图题 R7 下方）"""
from pathlib import Path
from openpyxl import load_workbook

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
wb = load_workbook(X)
ws = wb['1cmrl-conceptions']
fixed = 0
for img in ws._images:
    if img.anchor._from.row + 1 == 7:
        img.anchor._from.row += 1
        fixed += 1
wb.save(X)

wb2 = load_workbook(X)
ws2 = wb2['1cmrl-conceptions']
anchors = sorted(img.anchor._from.row + 1 for img in ws2._images)
cap7 = str(ws2.cell(7, 3).value or '')
note9 = str(ws2.cell(9, 3).value or '')
wb2.close()
print(f'修正 {fixed} 个锚点, 锚点列表: {anchors}')
print(f'图题 R7: {cap7[:16]}')
print(f'图注 R9: {len(note9)}字')
