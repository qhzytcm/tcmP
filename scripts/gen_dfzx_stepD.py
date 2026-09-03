# -*- coding: utf-8 -*-
"""
Step D: qhzy-东方哲学概论 缺陷修复
1. 0分布式体系 补充 node-coordinator 调度协议段（对齐系列）
2. 0历史 增补 2026 分布式体系记录行（呼应平台建设）
"""
import openpyxl

DST = r'C:\Users\DELL\Desktop\qhzy-东方哲学概论.xlsx'
wb = openpyxl.load_workbook(DST)

def append_rows(sheet_name, rows, ncols=5):
    ws = wb[sheet_name]
    start = ws.max_row + 1
    while start > 1:
        has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, ncols + 1))
        if has:
            break
        start -= 1
    for i, row in enumerate(rows):
        for c, v in enumerate(row, start=1):
            if v:
                ws.cell(row=start + i, column=c, value=v)
    return start

# ═══ 1. 0分布式体系 补 node-coordinator 协议段 ═══
append_rows('0分布式体系', [
    ['', '', '', '', ''],
    ['六、分布式任务调度（node-coordinator 协议, 对齐系列教材）', '', '', '', ''],
    ['调度规则', '内容', '', '', ''],
    ['任务分发', '一律经 node-coordinator: 章节编写→台式机/浪潮硬②; 人机实练数据采集→台式机T1-T4; 概念向量/决策数据集训练(A800)→浪潮硬②; 批量生成→华为硬①容器; 公网服务→华为云①; 版本门禁→GitHub', '', '', ''],
    ['节点心跳', '各节点60s心跳上报, 调度决策留痕', '', '', ''],
    ['故障降级', '节点故障→自动迁移就近节点; 任务重试', '', '', ''],
    ['算力池化', '轻量任务就近台式机, 批量上华为硬①容器, A800重训练上浪潮硬②, 公网服务上华为云①——各司其职, 不得"赢家通吃"', '', '', ''],
    ['兼容协议', '节点冲突先找就近调度/错峰方案, 不得宣布某节点"无用"而淘汰之(兼容性思维流第7局)', '', '', ''],
], ncols=3)
print("0分布式体系 node-coordinator 协议已补充")

# ═══ 2. 0历史 增补 2026 记录 ═══
ws = wb['0历史']
start = ws.max_row + 1
while start > 1:
    has = any(ws.cell(row=start - 1, column=c).value not in (None, '') for c in range(1, 5))
    if has:
        break
    start -= 1
hist_rows = [
    ['0000–0000', '0–1', '2026', '当代·岐黄智医', '平台(华为云/GitHub/华为硬/浪潮硬A800/四台式机)', '分布式网络编程体系、概念向量表、人机实练'],
]
for i, row in enumerate(hist_rows):
    for c, v in enumerate(row, start=1):
        if v:
            ws.cell(row=start + i, column=c, value=v)
print("0历史 2026 记录已增补")

wb.save(DST)
print("Step D 完成")
