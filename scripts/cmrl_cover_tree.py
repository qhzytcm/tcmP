# -*- coding: utf-8 -*-
"""封面页 sheet: 追加内容简介摘要 + 知识树三行（图题/图体/图注）"""
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image as PImg

X = Path(r'C:\Users\DELL\Desktop\强化学习的数学原理-赵世钰.xlsx')
TREE = Path(r'C:\Users\DELL\textbook-project\drafts\cmrl\figures\book\cmrl-knowledge-tree.png')

wb = load_workbook(X)
ws = wb['封面页']

# 1) 内容简介摘要（封面元素表之后）
start = ws.max_row + 2
INTRO = [
    ('内容简介', '本书以强化学习的数学理论为主线，完整移植到中医药临床与医院场景，三编 10 章：上编·提高认知（状态动作奖励/决策过程模型/统计估计随机逼近）；中编·确认信仰（价值估计方程/最优价值迭代/时序差分）；下编·追求幸福（奖励设计与四个最适度/函数近似/策略梯度/行动者-评价者与六者协同）。特色：中医场景全覆盖（45 幅插图）、数学与临床双向可验算、六者平台贯通、最适度≠最优化哲学贯穿。读者：中医药院校师生、临床医师、AI 研究者。', ''),
]
r = start
for a, b, c in INTRO:
    ws.cell(r, 1, a)
    ws.cell(r, 2, b)
    ws.cell(r, 3, c)
    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0,
                                      (len(b) // 90 + 1) * 14 + 6)
    r += 1
intro_end = r - 1

# 2) 知识树三行（图题/图体/图注）
cap = '图 K-1  cmrl 单门学科知识树：须—根—干—枝—叶（依据 0映射1 主编知识结构）'
note = ('图注：cmrl 单门学科知识树（依据 0映射1 主编知识结构）。须=数学基础'
        '（概率/期望/矩阵/收敛/梯度——知识须根，吸取数学养分）；根=状态-动作-奖励'
        '三要素（学科根基）；干=三编（提高认知/确认信仰/追求幸福）；枝=10 章'
        '（上编 1-3、中编 4-6、下编 7-10）；叶=单一终末知识点（共 41 叶，如'
        '状态S=四诊、贝尔曼方程=递归账本、Q学习=向最优学、六者协同——对应'
        '0cmrl目录 段落级四级编码，可展开至 480 叶）。树形隐喻：须根吸取数学养分，'
        '主干为三编哲学，枝为章节，叶为可单独考核的终末知识点。')

# 图题行
cap_row = intro_end + 2
c_cell = ws.cell(cap_row, 1, cap)
c_cell.font = Font(name='黑体', size=12, bold=False)
c_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells(start_row=cap_row, start_column=1, end_row=cap_row, end_column=3)
ws.row_dimensions[cap_row].height = 22

# 图体行（图片）
body_row = cap_row + 1
w, h = PImg.open(TREE).size
img = XLImage(str(TREE))
img.width = 480
img.height = int(480 * h / w)
ws.add_image(img, f'A{body_row}')
ws.row_dimensions[body_row].height = img.height * 0.75

# 图注行
note_row = body_row + 1
n_cell = ws.cell(note_row, 1, note)
n_cell.font = Font(name='黑体', size=10)
n_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
n_rows = max(1, math.ceil(len(note) / 100))
ws.row_dimensions[note_row].height = n_rows * 14 + 6

wb.save(X)
print(f'封面页追加: 内容简介 R{start}-R{intro_end}, 图题 R{cap_row}, 图体 R{body_row}, '
      f'图注 R{note_row} (图 {img.width}x{img.height})')
