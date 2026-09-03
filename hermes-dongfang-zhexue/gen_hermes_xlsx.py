# -*- coding: utf-8 -*-
"""
生成 hermes-东方哲学概论.xlsx（14个Sheet）
策略：保留初稿全部正文与档案内容，注入 hermes 教材工作流结构化增强。
"""
import openpyxl, json, os, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from hermes_content import (CHAPTER_META, PRACTICE_LOOP, EXERCISES, SUMMARY, GLOSSARY)

SRC = r'C:\Users\DELL\Desktop\东方哲学概论.xlsx'
HERE = os.path.dirname(os.path.abspath(__file__))
TREE = os.path.join(HERE, 'chapters_tree.json')
OUT = r'C:\Users\DELL\Desktop\hermes-东方哲学概论.xlsx'

# ---------- 样式 ----------
INDIGO = '1F3A5F'      # 深靛蓝（封面主色）
INDIGO_L = 'D6E4F0'    # 浅靛蓝
GOLD = 'C9A227'        # 金色点缀
GRAY_L = 'F2F2F2'      # 浅灰
WHITE = 'FFFFFF'
GREEN_L = 'E2EFDA'
BLUE_L = 'DDEBF7'
ORANGE_L = 'FCE4D6'
PURPLE_L = 'E4DFEC'

F_TITLE = Font(name='微软雅黑', size=16, bold=True, color=WHITE)
F_H1 = Font(name='微软雅黑', size=13, bold=True, color=INDIGO)
F_H2 = Font(name='微软雅黑', size=11, bold=True, color=INDIGO)
F_H3 = Font(name='微软雅黑', size=10, bold=True, color='2E5A87')
F_BODY = Font(name='微软雅黑', size=10, color='333333')
F_BODY_B = Font(name='微软雅黑', size=10, bold=True, color='333333')
F_TAG = Font(name='微软雅黑', size=9, bold=True, color=WHITE)
F_SMALL = Font(name='微软雅黑', size=9, color='666666')

def pf(color):
    return PatternFill(patternType='solid', fgColor=color)

FILL_TITLE = pf(INDIGO)
FILL_H1 = pf(INDIGO_L)
FILL_H2 = pf(GRAY_L)
FILL_GREEN = pf(GREEN_L)
FILL_BLUE = pf(BLUE_L)
FILL_ORANGE = pf(ORANGE_L)
FILL_PURPLE = pf(PURPLE_L)
FILL_WHITE = pf(WHITE)

AL_L = Alignment(horizontal='left', vertical='top', wrap_text=True)
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_CT = Alignment(horizontal='center', vertical='top', wrap_text=True)

THIN = Side(style='thin', color='BBBBBB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TAG_COLORS = {
    '章': FILL_TITLE, '元': FILL_BLUE, '节': FILL_H1, '小节': FILL_H2,
    '细目': FILL_ORANGE, '正文': FILL_GREEN, '实练': FILL_BLUE,
    '习题': FILL_ORANGE, '小结': FILL_PURPLE, '术语': FILL_GREEN,
    '区块': FILL_TITLE,
}

def tag_fill(tag):
    return TAG_COLORS.get(tag, FILL_H2)

class SheetWriter:
    def __init__(self, ws, widths=None):
        self.ws = ws
        self.r = 1
        if widths:
            for i, w in enumerate(widths, 1):
                self.ws.column_dimensions[get_column_letter(i)].width = w

    def row(self, cells, fills=None, fonts=None, aligns=None, height=None):
        """cells: list of values; fills/fonts/aligns optional lists (or single)."""
        for ci, val in enumerate(cells, 1):
            c = self.ws.cell(row=self.r, column=ci, value=val)
            c.border = BORDER
            if fills:
                f = fills[ci-1] if isinstance(fills, list) else fills
                if f: c.fill = f
            if fonts:
                f = fonts[ci-1] if isinstance(fonts, list) else fonts
                if f: c.font = f
            if aligns:
                a = aligns[ci-1] if isinstance(aligns, list) else aligns
                if a: c.alignment = a
        if height:
            self.ws.row_dimensions[self.r].height = height
        self.r += 1

    def title(self, text, span):
        """span: number of columns to merge."""
        self.row([text] + [''] * (span - 1), fills=FILL_TITLE, fonts=F_TITLE, aligns=AL_C)
        self.ws.merge_cells(start_row=self.r-1, start_column=1, end_row=self.r-1, end_column=span)
        self.ws.row_dimensions[self.r-1].height = 30

    def section(self, text, span, fill=FILL_H1, font=F_H1):
        self.row([text] + [''] * (span - 1), fills=fill, fonts=font, aligns=AL_L)
        self.ws.merge_cells(start_row=self.r-1, start_column=1, end_row=self.r-1, end_column=span)
        self.ws.row_dimensions[self.r-1].height = 22

    def kv(self, k, v, span, fill=FILL_BLUE, kfont=F_BODY_B, vfont=F_BODY):
        self.row([k, v] + [''] * (span - 2), fills=[fill, FILL_WHITE] + [FILL_WHITE]*(span-2),
                 fonts=[kfont, vfont] + [vfont]*(span-2), aligns=[AL_C, AL_L] + [AL_L]*(span-2))
        self.ws.row_dimensions[self.r-1].height = max(18, 14 * (len(str(v)) // 46 + 1))

    def tagged(self, tag, text, span):
        self.row([tag, text] + [''] * (span - 2), fills=[tag_fill(tag), FILL_WHITE] + [FILL_WHITE]*(span-2),
                 fonts=[F_TAG, F_BODY] + [F_BODY]*(span-2), aligns=[AL_C, AL_L] + [AL_L]*(span-2))
        self.ws.row_dimensions[self.r-1].height = max(18, 14 * (len(str(text)) // 44 + 1))

    def blank(self, n=1):
        for _ in range(n):
            self.row([''] * 2, fills=[FILL_WHITE, FILL_WHITE], fonts=[F_BODY, F_BODY])
            self.ws.row_dimensions[self.r-1].height = 6

# ---------- 载入数据 ----------
wb_src = openpyxl.load_workbook(SRC, data_only=True)
tree = json.load(open(TREE, encoding='utf-8'))

def copy_src_rows(sw, sheet_name, tag=''):
    """复制初稿某Sheet全部非空行：单列→tagged；多列→取非空值序列（防超宽列）。"""
    ws = wb_src[sheet_name]
    n = 0
    for row in ws.iter_rows(values_only=True):
        vals = ['' if v is None else str(v).replace('\n', ' ') for v in row]
        idx = [i for i, v in enumerate(vals) if v.strip() != '']
        if not idx:
            continue
        if len(idx) == 1:
            sw.tagged(tag or '正文', vals[idx[0]], 2)
        else:
            cells = [vals[i] for i in idx[:9]]  # 最多保留9个非空值，压缩中间空列
            sw.row(cells, fills=[FILL_H2] + [FILL_WHITE] * (len(cells) - 1),
                   fonts=[F_BODY_B] + [F_BODY] * (len(cells) - 1),
                   aligns=[AL_C] + [AL_L] * (len(cells) - 1))
            sw.ws.row_dimensions[sw.r - 1].height = max(
                18, 14 * (max(len(str(c)) for c in cells[1:]) // 44 + 1))
        n += 1
    return n

# ---------- Sheet 1: 0历史 ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '0历史'
sw = SheetWriter(ws, [14, 14, 14, 16, 30, 50])
sw.title('《东方哲学概论》历史概貌（hermes版 · 供作者subagent快速定位）', 6)
sw.row(['进展步长', '距今年数', '公元年份', '时代', '代表人物', '代表概念'],
       fills=[FILL_H1]*6, fonts=[F_BODY_B]*6, aligns=[AL_C]*6)
src_hist = wb_src['历史']
for row in src_hist.iter_rows(min_row=4, max_row=70, values_only=True):
    vals = ['' if v is None else str(v).strip() for v in row[:6]]
    if not any(vals):
        continue
    sw.row(vals, fills=[None]*6, fonts=[F_BODY]*6, aligns=[AL_CT, AL_CT, AL_CT, AL_C, AL_L, AL_L])
sw.blank()
sw.section('关键转折点（Top 10）', 6)
sw.row(['', '', '', '时间', '事件', '哲学意义'],
       fills=[FILL_H1]*6, fonts=[F_BODY_B]*6, aligns=[AL_C]*6)
for row in src_hist.iter_rows(min_row=74, max_row=83, values_only=True):
    # 初稿布局：D列=时间 E列=事件 F列=哲学意义
    d = '' if row[3] is None else str(row[3]).strip()
    e = '' if row[4] is None else str(row[4]).strip()
    f = '' if row[5] is None else str(row[5]).strip()
    if not any([d, e, f]):
        continue
    sw.row(['', '', '', d, e, f], fills=[None]*6, fonts=[F_BODY]*6, aligns=[AL_CT, AL_CT, AL_CT, AL_C, AL_L, AL_L])
sw.blank()
sw.section('hermes版使用说明（作者subagent必读）', 6, fill=FILL_GREEN, font=F_H1)
for line in [
    '① 本表为作者subagent的"历史上下文锚点"：写作任一章前，先定位该章概念在时间轴上的坐标（如"阴阳"萌芽于新石器~西周，"辨证"定型于东汉张仲景）。',
    '② 教材锚点：每格代表人物优先选取对中医哲学有直接影响者（张仲景/孙思邈/朱震亨/陈可冀等），对应120门教材中《中医基础理论》《伤寒论讲义》《各家学说》等的知识单元。',
    '③ 图谱接口：本表可导入"历代圣贤决策数据集"（附录B），作为知识图谱时间戳字段的基准索引；各格人物即图谱"医圣节点"候选。',
    '④ 双序呼应：当代格（0000-0050）标注"周志华AI × 陈可冀中医"= 全书"人机闭环"主线的历史落点，与封面双序形成闭环。',
    '⑤ 用法建议：可作全书卷首"历史长河"拉页；可拆为每章"时间卡片"配合第4章传承延绵的族谱记忆回放；可作"历代圣贤决策数据集"时间戳基准。',
]:
    sw.tagged('说明', line, 6)

# ---------- Sheet 2-4: 0映射1/2/3 ----------
def write_profile(sheet_name, tag_label, addon_lines, title_text, widths):
    ws = wb.create_sheet(sheet_name)
    sw = SheetWriter(ws, widths)
    sw.title(title_text, len(widths))
    copy_src_rows(sw, sheet_name, tag=tag_label)
    sw.blank()
    sw.section('【hermes工作流增强】', len(widths), fill=FILL_GREEN, font=F_H1)
    for line in addon_lines:
        sw.tagged('增强', line, len(widths))
    return sw

write_profile('0映射1', '档案', [
    'L4作者工作台：本档案即 author-subagent 的 User Profile 加载文件。作者写作时按 textbook-writing 规范执行——章-节-细目三级结构、习题60-30-10比例（基础/综合/拓展）、术语表随章维护。',
    '章节分工：第3章"抽象建模"（阴阳动力学/五行拓扑）以蔡大勇"证候能量代谢×多指标群耦合"研究为主锚点；第1章"感知在地"以"多尺度研究（分子→器官→整体）"为尺度伸缩案例。',
    '协作协议：作者初稿完成后交 L3 学科主编初审（G1）→ 中医审阅subagent（0映射2）与 AI审阅subagent（0映射3）联席复审（G2）→ L1 终验（G3）。',
    'backend 建议：local（Windows开发机）；交付物命名 chapter-{NN}-{title}-v{version}.md，存 drafts/。',
    '红线：凡 [待核实] 字段在本人书面确认前不得用于正式出版；不编造专利号/新药证书号。',
], '0映射1 · 作者Subagent现实映射（蔡大勇 Dayong Cai · 药理专家）', [16, 22, 26, 20, 30, 40])

write_profile('0映射2', '档案', [
    '四维评分卡（hermes质量门禁标准）：科学性40% + 系统性25% + 教学适切性20% + 可读性15%。中医审阅subagent按此对教材章节打分，≥85/100 通过 G1，≥90/100 通过 G2。',
    '中医审稿CheckList：① 经典引用规范（《内经》《伤寒》《金匮》不得断章取义）② 辨证逻辑自洽（八纲/十纲辨证不越位）③ 中药安全性（乌头类剂量严把关）④ 循证一致性（RCT与流行病学证据）⑤ 中西医结合合理性。',
    '章节审稿重点：第3章"气血运行状态"（活血化瘀/血瘀证I型II型）、第5章"逻辑升维"（辨证=约束优化）、第6章"伦理熔断"（治未病预防观）。',
    '协作：与作者subagent（0映射1）形成"临床-药理"互补；与AI审阅subagent（0映射3）联席仲裁——临床争议陈可冀终裁，AI方法论争议周志华终裁，跨学科争议三人联席。',
    '红线：拒绝违背经典原意、拒绝夸大疗效、拒绝忽视毒性、拒绝AI冒充人类专家原创。',
], '0映射2 · 中医审阅人Subagent现实映射（陈可冀 Keji Chen · 国医大师）', [16, 24, 30, 24, 34])

write_profile('0映射3', '档案', [
    'AI方法论审稿标准：① 算法正确性（伪代码/公式/复杂度严格可验证）② 数学严谨性（定理证明链完整、符号统一）③ 表述清晰性（"挑西瓜"式直觉先行→形式化殿后）④ 教学适配性（面向中医药学子，强调可解释性）⑤ 跨学科对齐（AI概念映射到东方哲学术语须逻辑可解释）。',
    '章节审稿重点：第1章"风水建模/聚类回归"、第3章"五行拓扑=GNN/卦象=状态机"、第5章"中庸=约束优化/定理核验"——重点查数学硬伤与伪精确表述。',
    '弱监督现实：中医标注数据稀缺（对应周志华弱监督学习方向）——教材实练设计须容忍小样本，不假设标签干净。',
    '资源约束：教材中的AI实验须考虑算力约束（CoRE-learning精神），推荐可复现的小规模仿真而非大模型依赖。',
    '红线：拒绝夸大AI能力（"AI替代中医辨证"）、拒绝伪造数据、拒绝AI冒充人类专家/医师/审稿人、拒绝忽视AI安全隐私伦理。',
], '0映射3 · AI审阅人Subagent现实映射（周志华 Zhi-Hua Zhou · AI院士）', [16, 24, 30, 24, 34])

# ---------- Sheet 5: 0封面 ----------
ws = wb.create_sheet('0封面')
sw = SheetWriter(ws, [16, 90])
sw.title('《东方哲学概论》封面与前置页（hermes版）', 2)
sw.section('【教材元数据卡 · hermes工作流登记】', 2, fill=FILL_GREEN, font=F_H1)
for k, v in [
    ('教材编号', '智能与管理子系列 · 第98号（120门系列教材之一）'),
    ('课程性质', '人文常识课程 · 本科公共课（可选硕士/博士通识）'),
    ('学分/学时', '2学分 / 共30学时（理论18 + 人机实练12）'),
    ('前置依赖', '无（系列教材之底层逻辑基座；为中医基础理论、诊断学、方剂学等提供世界观与方法论）'),
    ('编写轮次', '轮次1（文化通识，全院系共用，全并行）'),
    ('教学层级', '本科优先，跨层级复用（硕博通识选修）'),
    ('编写委员会', 'L1行业规划主编（签批总纲）→ L2智能与管理领域主编 → L3学科内容主编（样章）→ L4篇章作者（蔡大勇 profile 驱动）'),
    ('质量门禁', 'G1学科初审≥85 → G2领域复审≥90（陈可冀+周志华双审）→ G3终验≥95（L1签批）'),
    ('数字化资源', '双序导读 / 习题感知题·仿行题·推演题 / 五行拓扑·阴阳动力学源代码 / 历代圣贤决策数据集 / 概念向量表'),
]:
    sw.kv(k, v, 2)
sw.blank()
sw.section('封面设计说明（初稿保留）', 2)
copy_src_rows(sw, '0封面', tag='封面')
sw.blank()
sw.section('hermes版新增：编写委员会与CI/CD闭环', 2, fill=FILL_GREEN, font=F_H1)
for line in [
    '本教材作为120门系列教材的"底层操作系统"，其知识将注入病证知识图谱（60,000+病证单位节点）：每章细目与图谱"概念单位"双向锚定。',
    '六者平台对接：第6章"人机共治/伦理熔断"→ 规者/法者Agent；第3章"五行拓扑"→ 医者Agent病证推理；第4章"传承延绵"→ 医圣成长引擎（张仲景/孙思邈人格）。',
    'CI/CD闭环：教材出版后，六者平台使用数据（对话日志）反向反馈教材修订——本工作簿的 8通雅 预留Sheet即下一次迭代的入口。',
]:
    sw.tagged('闭环', line, 2)

# ---------- Sheet 6: 0目录 ----------
ws = wb.create_sheet('0目录')
sw = SheetWriter(ws, [10, 34, 58])
sw.title('《东方哲学概论》正文目录（hermes版 · 含教学目标与学时）', 3)
sw.section('全书结构：感知→仿行→数理→传承→逻辑→生态 六级递进闭环（人法地→地法天→天法道→道法自然→道体恒常→自然无为）', 3, fill=FILL_H2, font=F_BODY_B)
src_toc = wb_src['0目录']
for row in src_toc.iter_rows(values_only=True):
    v = row[2] if len(row) > 2 else None
    if v is None or str(v).strip() == '':
        continue
    t = str(v).strip()
    sw.row([t, '', ''], fills=[FILL_H2, FILL_WHITE, FILL_WHITE], fonts=[F_BODY_B, F_BODY, F_BODY], aligns=[AL_L, AL_L, AL_L])
sw.blank()
sw.section('各章教学目标 / 学时 / 门禁一览（hermes工作流登记）', 3, fill=FILL_GREEN, font=F_H1)
sw.row(['章', '教学目标（知识/能力/素养摘要）', '学时与门禁'], fills=[FILL_H1]*3, fonts=[F_BODY_B]*3, aligns=[AL_C]*3)
for num in range(1, 7):
    m = CHAPTER_META[num]
    sw.row(['第%d章 %s' % (num, m['title']),
            '知识：%s\n能力：%s\n素养：%s' % (m['goal_know'][:80], m['goal_skill'][:60], m['goal_value'][:40]),
            '%s\n%s' % (m['hours'], m['gate'][:50])],
           fills=[FILL_H2, FILL_WHITE, FILL_WHITE], fonts=[F_BODY_B, F_BODY, F_BODY], aligns=[AL_C, AL_L, AL_L])
    sw.ws.row_dimensions[sw.r-1].height = 60

# ---------- Sheet 7-12: 1感知~6生态 ----------
for num in range(1, 7):
    key = '%d' % num
    sheet_names = {1: '1感知', 2: '2行规', 3: '3数理', 4: '4传承', 5: '5逻辑', 6: '6生态'}
    ws = wb.create_sheet(sheet_names[num])
    sw = SheetWriter(ws, [10, 100])
    m = CHAPTER_META[num]
    sw.title('第%d章 %s（hermes版）——%s' % (num, m['title'], m['subtitle']), 2)
    # 章首元数据块
    sw.section('【章首目标块 · 教学元数据】', 2, fill=FILL_GREEN, font=F_H1)
    for k, v in [('教学层级', m['level']), ('学时分配', m['hours']), ('前置依赖', m['prereq']),
                 ('知识目标', m['goal_know']), ('能力目标', m['goal_skill']), ('素养目标', m['goal_value']),
                 ('人机闭环路径', m['loop']), ('病证单位锚点(DSU)', m['dsu']), ('质量门禁', m['gate'])]:
        sw.kv(k, v, 2)
    sw.blank()
    # 正文（来自初稿树）
    sw.section('【正文】', 2)
    ch = tree[sheet_names[num]]
    if ch['intro']:
        for p in ch['intro']:
            sw.tagged('正文', p, 2)
    for sec in ch['sections']:
        sw.tagged('节', '%s %s' % (sec['code'], sec['title']), 2)
        for p in sec['intro']:
            sw.tagged('正文', p, 2)
        for sb in sec['subsections']:
            sw.tagged('小节', '%s %s' % (sb['code'], sb['title']), 2)
            for p in sb['intro']:
                sw.tagged('正文', p, 2)
            for it in sb['items']:
                sw.tagged('细目', '%s %s' % (it['code'], it['title']), 2)
                for p in it['paras']:
                    sw.tagged('正文', p, 2)
    sw.blank()
    # 人机实练闭环
    sw.section('【人机实练 · 四步闭环】', 2, fill=FILL_BLUE, font=F_H1)
    for step, desc in PRACTICE_LOOP[num]:
        sw.tagged('实练', '%s：%s' % (step, desc), 2)
    sw.blank()
    # 习题
    sw.section('【习题（60-30-10：基础60% / 综合30% / 拓展10%）】', 2, fill=FILL_ORANGE, font=F_H1)
    ex = EXERCISES[num]
    for i, q in enumerate(ex['base'], 1):
        sw.tagged('习题', '【基础%d】%s' % (i, q), 2)
    for i, q in enumerate(ex['comp'], 1):
        sw.tagged('习题', '【综合%d】%s' % (i, q), 2)
    for i, q in enumerate(ex['adv'], 1):
        sw.tagged('习题', '【拓展%d】%s' % (i, q), 2)
    sw.blank()
    # 本章小结
    sw.section('【本章小结】', 2, fill=FILL_PURPLE, font=F_H1)
    sw.tagged('小结', SUMMARY[num], 2)
    sw.blank()
    # 术语表
    sw.section('【术语表（东方范畴 ↔ AI映射）】', 2, fill=FILL_GREEN, font=F_H1)
    sw.row(['范畴', 'AI映射', '释义'], fills=[FILL_H1]*3, fonts=[F_BODY_B]*3, aligns=[AL_C]*3)
    for term, ai, mean in GLOSSARY[num]:
        sw.row([term, ai, mean], fills=[FILL_H2, FILL_WHITE, FILL_WHITE], fonts=[F_BODY_B, F_BODY, F_BODY], aligns=[AL_C, AL_C, AL_L])
        sw.ws.row_dimensions[sw.r-1].height = 20

# ---------- Sheet 13: 7附件 ----------
ws = wb.create_sheet('7附件')
sw = SheetWriter(ws, [16, 90])
sw.title('附录（hermes版）', 2)
copy_src_rows(sw, '7附件', tag='附录')
sw.blank()
sw.section('【hermes增强 · 附录D 数据与图谱对接】', 2, fill=FILL_GREEN, font=F_H1)
for line in [
    'D.1 概念向量表对接：附录A的"东方哲学核心概念向量表"与tcmP平台的 embedding 引擎（TF-IDF+FTS5+RRF，内网零token）同构——本表即为该引擎在哲学域的"种子词库"。',
    'D.2 圣贤决策数据集对接：附录B的"历代圣贤决策数据集"与病证知识图谱（DiseaseSyndromeUnit）共享时间戳Schema，可双向检索"以病索证/以证溯病"扩展为"以事索理/以理溯事"。',
    'D.3 伦理规范对接：附录C的"人机闭环实验伦理规范"三级防御体系（不可篡改公理/可审计日志/伦理熔断）即六者平台 规者/法者 Agent 的规则库输入。',
    'D.4 ICD-11桥接：本教材病证单位锚点可经 tcmP 的 /icd 三端点（icd11_mms.db，31,838实体）映射西医诊断编码，实现"哲学概念-中医证候-西医编码"三层贯通。',
]:
    sw.tagged('增强', line, 2)

# ---------- Sheet 14: 8通雅 ----------
ws = wb.create_sheet('8通雅')
sw = SheetWriter(ws, [16, 90])
sw.title('8通雅（hermes版 · 预留Sheet）', 2)
sw.section('【Sheet用途说明】', 2, fill=FILL_GREEN, font=F_H1)
for line in [
    '本Sheet为"通雅"内容的预留位——按用户约定，通雅内容将在下一次迭代追加，本次不写入正文。',
    '"通雅"取方以智《通雅》"会通古今、雅正名实"之意：作为全书的融通总结层，将六章（感知/行规/数理/传承/逻辑/生态）的核心概念贯通为一，并以雅正语言重述东方哲学的总体精神。',
]:
    sw.tagged('说明', line, 2)
sw.blank()
sw.section('【下次迭代追加计划（TODO）】', 2, fill=FILL_ORANGE, font=F_H1)
for line in [
    '[ ] T1 通雅总论：以"道"为主线，贯通六章的六重闭环（感知闭环→行为闭环→数理闭环→代际闭环→推理闭环→共治闭环）为一整体。',
    '[ ] T2 概念雅训：对全书术语表（每章8-10条，共约54条）做雅正释义，形成"东方哲学核心概念雅训"总表。',
    '[ ] T3 人机共演跋：呼应封面双序（陈可冀"天人合一寻根励志" × 周志华"妙用迭代万物可期"），撰写全书跋文，落点"人机共演"。',
    '[ ] T4 图谱回写：将通雅总结层回写病证知识图谱，形成"概念单位-病证单位"跨域边，供六者Agent调用。',
    '[ ] T5 审核闭环：通雅内容完成后经 中医审阅（0映射2）与 AI审阅（0映射3）联席评审，达标后并入 v2.0。',
]:
    sw.tagged('TODO', line, 2)
sw.blank()
sw.section('【版本记录】', 2)
for line in ['v2.0（下次迭代）：追加通雅内容 + 六章修订（依据使用反馈）', 'v1.0-hermes（本次）：完成hermes工作流增强，14 Sheet齐备']:
    sw.tagged('版本', line, 2)

wb.save(OUT)
print('saved:', OUT)
print('sheets:', wb.sheetnames)
